"""
Routes per la gestione degli Onorari (Fees)
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
import uuid
import io

from .deps import get_db, get_current_user, require_commercialista, log_activity
from .models import FeeCreate, FeeUpdate, FeeResponse

router = APIRouter(tags=["fees"])


@router.get("/fees/all")
async def get_all_fees(
    search: Optional[str] = None,
    client_type: Optional[str] = None,
    status: Optional[str] = None,
    year: Optional[int] = None,
    user: dict = Depends(require_commercialista)
):
    """
    Recupera tutti gli onorari con filtri opzionali.
    - search: ricerca per descrizione o nome cliente
    - client_type: filtra per tipo cliente (societa, autonomo, vivienda_vacacional, persona_fisica)
    - status: filtra per stato (pending, paid, overdue)
    - year: filtra per anno scadenza
    """
    db = get_db()
    # Prima recupera tutti i clienti per join
    clients_cursor = db.users.find({"role": "cliente"}, {"_id": 0, "id": 1, "full_name": 1, "email": 1, "tipo_cliente": 1})
    clients_list = await clients_cursor.to_list(10000)
    clients_map = {c["id"]: c for c in clients_list}
    
    # Query per gli onorari
    query = {}
    
    if status:
        query["status"] = status
    
    if year:
        # Filtra per anno della data di scadenza
        query["due_date"] = {"$regex": f"^{year}"}
    
    # Recupera onorari
    fees = await db.fees.find(query, {"_id": 0}).sort("due_date", -1).to_list(10000)
    
    # Arricchisci con info cliente e applica filtri aggiuntivi
    result = []
    for fee in fees:
        client = clients_map.get(fee.get("client_id"), {})
        
        # Filtro per tipo cliente
        if client_type and client.get("tipo_cliente") != client_type:
            continue
        
        # Filtro per ricerca
        if search:
            search_lower = search.lower()
            match = False
            if search_lower in fee.get("description", "").lower():
                match = True
            if search_lower in client.get("full_name", "").lower():
                match = True
            if search_lower in client.get("email", "").lower():
                match = True
            if not match:
                continue
        
        # Aggiungi info cliente al fee
        fee["client_name"] = client.get("full_name", "N/A")
        fee["client_email"] = client.get("email", "")
        fee["client_type"] = client.get("tipo_cliente", "")
        result.append(fee)
    
    return result


@router.get("/fees/summary")
async def get_global_fees_summary(user: dict = Depends(require_commercialista)):
    """Ottiene un riepilogo globale degli onorari"""
    db = get_db()
    fees = await db.fees.find({}, {"_id": 0}).to_list(10000)
    
    total_pending = sum(f["amount"] for f in fees if f.get("status") == "pending")
    total_paid = sum(f["amount"] for f in fees if f.get("status") == "paid")
    total_overdue = sum(f["amount"] for f in fees if f.get("status") == "overdue")
    
    # Raggruppa per cliente
    by_client = {}
    for fee in fees:
        cid = fee.get("client_id")
        if cid not in by_client:
            by_client[cid] = {"pending": 0, "paid": 0, "count": 0}
        by_client[cid]["count"] += 1
        if fee.get("status") == "pending":
            by_client[cid]["pending"] += fee["amount"]
        elif fee.get("status") == "paid":
            by_client[cid]["paid"] += fee["amount"]
    
    return {
        "total_pending": total_pending,
        "total_paid": total_paid,
        "total_overdue": total_overdue,
        "total_count": len(fees),
        "clients_count": len(by_client)
    }


@router.get("/fees/by-client")
async def get_fees_grouped_by_client(user: dict = Depends(require_commercialista)):
    """Recupera tutti i clienti con i loro onorari raggruppati"""
    db = get_db()
    # Recupera tutti i clienti
    clients = await db.users.find(
        {"role": "cliente"}, 
        {"_id": 0, "id": 1, "full_name": 1, "email": 1, "tipo_cliente": 1}
    ).to_list(10000)
    
    # Recupera tutti gli onorari
    fees = await db.fees.find({}, {"_id": 0}).to_list(10000)
    
    # Raggruppa per cliente
    fees_by_client = {}
    for fee in fees:
        cid = fee.get("client_id")
        if cid not in fees_by_client:
            fees_by_client[cid] = []
        fees_by_client[cid].append(fee)
    
    # Costruisci risultato
    result = []
    for client in clients:
        client_fees = fees_by_client.get(client["id"], [])
        total_pending = sum(f["amount"] for f in client_fees if f.get("status") == "pending")
        total_paid = sum(f["amount"] for f in client_fees if f.get("status") == "paid")
        
        # Calcola totale Iguala mensile
        iguala_monthly = sum(
            f["amount"] for f in client_fees 
            if f.get("fee_type", "").startswith("iguala_") or f.get("is_recurring")
        )
        
        result.append({
            "id": client["id"],
            "full_name": client.get("full_name", "N/A"),
            "email": client.get("email", ""),
            "tipo_cliente": client.get("tipo_cliente", ""),
            "fees": client_fees,
            "fees_count": len(client_fees),
            "total_pending": total_pending,
            "total_paid": total_paid,
            "iguala_monthly": iguala_monthly
        })
    
    # Ordina per nome
    result.sort(key=lambda x: x["full_name"].lower())
    return result


@router.get("/fees/export-excel")
async def export_fees_excel(
    category: Optional[str] = None,
    fee_type: Optional[str] = None,
    user: dict = Depends(require_commercialista)
):
    """Esporta onorari in formato Excel con filtri"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installato")
    
    db = get_db()
    
    # Recupera clienti con filtro categoria
    client_query = {"role": "cliente"}
    if category and category != "all":
        client_query["tipo_cliente"] = category
    
    clients = await db.users.find(
        client_query, 
        {"_id": 0, "id": 1, "full_name": 1, "email": 1, "tipo_cliente": 1}
    ).to_list(10000)
    
    client_ids = [c["id"] for c in clients]
    client_map = {c["id"]: c for c in clients}
    
    # Recupera onorari
    fee_query = {"client_id": {"$in": client_ids}}
    if fee_type and fee_type != "all":
        if fee_type == "iguala":
            fee_query["$or"] = [
                {"fee_type": {"$regex": "^iguala_"}},
                {"is_recurring": True}
            ]
        else:
            fee_query["fee_type"] = fee_type
    
    fees = await db.fees.find(fee_query, {"_id": 0}).to_list(10000)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Onorari"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F855A", end_color="2F855A", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    headers = ["Cliente", "Tipo Cliente", "Email", "Descrizione Onorario", "Tipo Onorario", "Importo (€)", "Stato", "Scadenza", "Mese Rif."]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Dati
    row = 2
    for fee in fees:
        client = client_map.get(fee.get("client_id"), {})
        fee_type_labels = {
            "standard": "Standard",
            "consulenza": "Consulenza",
            "pratica": "Pratica/Procedura",
            "dichiarazione": "Dichiarazione Fiscale",
            "iguala_buste_paga": "Iguala - Buste Paga",
            "iguala_contabilita": "Iguala - Contabilità",
            "iguala_domicilio": "Iguala - Domicilio"
        }
        status_labels = {"pending": "In Attesa", "paid": "Pagato", "overdue": "Scaduto"}
        tipo_cliente_labels = {
            "societa": "Società",
            "autonomo": "Autonomo",
            "vivienda_vacacional": "Vivienda Vacacional",
            "persona_fisica": "Persona Fisica"
        }
        
        ws.cell(row=row, column=1, value=client.get("full_name", "N/A")).border = thin_border
        ws.cell(row=row, column=2, value=tipo_cliente_labels.get(client.get("tipo_cliente", ""), client.get("tipo_cliente", ""))).border = thin_border
        ws.cell(row=row, column=3, value=client.get("email", "")).border = thin_border
        ws.cell(row=row, column=4, value=fee.get("description", "")).border = thin_border
        ws.cell(row=row, column=5, value=fee_type_labels.get(fee.get("fee_type", "standard"), fee.get("fee_type", ""))).border = thin_border
        ws.cell(row=row, column=6, value=fee.get("amount", 0)).border = thin_border
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=status_labels.get(fee.get("status", ""), fee.get("status", ""))).border = thin_border
        ws.cell(row=row, column=8, value=fee.get("due_date", "-") or "-").border = thin_border
        ws.cell(row=row, column=9, value=fee.get("recurring_month", "-") or "-").border = thin_border
        row += 1
    
    # Aggiungi riga totale
    if fees:
        total_row = row + 1
        ws.cell(row=total_row, column=5, value="TOTALE:").font = Font(bold=True)
        total_amount = sum(f.get("amount", 0) for f in fees)
        ws.cell(row=total_row, column=6, value=total_amount).font = Font(bold=True)
        ws.cell(row=total_row, column=6).number_format = '#,##0.00'
    
    # Larghezza colonne
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Salva in buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Genera nome file
    filename = f"onorari"
    if category and category != "all":
        filename += f"_{category}"
    if fee_type and fee_type != "all":
        filename += f"_{fee_type}"
    filename += f"_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Client-specific fee routes
client_fees_router = APIRouter(prefix="/clients/{client_id}/fees", tags=["client-fees"])


@client_fees_router.get("", response_model=List[FeeResponse])
async def get_client_fees(client_id: str, user: dict = Depends(require_commercialista)):
    """Recupera tutti gli onorari di un cliente"""
    db = get_db()
    fees = await db.fees.find({"client_id": client_id}, {"_id": 0}).sort("due_date", -1).to_list(1000)
    return [FeeResponse(**fee) for fee in fees]


@client_fees_router.post("", response_model=FeeResponse)
async def create_fee(client_id: str, fee_data: FeeCreate, user: dict = Depends(require_commercialista)):
    """Crea un nuovo onorario per un cliente"""
    db = get_db()
    # Verifica che il cliente esista
    client = await db.users.find_one({"id": client_id, "role": "cliente"})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    # Determina se is_recurring in base al fee_type
    is_recurring = fee_data.is_recurring or fee_data.fee_type.startswith("iguala_")
    
    fee_id = str(uuid.uuid4())
    fee = {
        "id": fee_id,
        "client_id": client_id,
        "description": fee_data.description,
        "amount": fee_data.amount,
        "due_date": fee_data.due_date,
        "status": fee_data.status,
        "paid_date": None,
        "notes": fee_data.notes,
        "fee_type": fee_data.fee_type,
        "is_recurring": is_recurring,
        "recurring_month": fee_data.recurring_month,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.fees.insert_one(fee)
    
    await log_activity(
        "creazione_onorario",
        f"Onorario €{fee_data.amount} ({fee_data.fee_type}) creato per cliente {client_id}",
        user["id"]
    )
    
    return FeeResponse(**fee)


@client_fees_router.put("/{fee_id}", response_model=FeeResponse)
async def update_fee(client_id: str, fee_id: str, fee_data: FeeUpdate, user: dict = Depends(require_commercialista)):
    """Aggiorna un onorario"""
    db = get_db()
    update_dict = {k: v for k, v in fee_data.model_dump().items() if v is not None}
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    # Se stato diventa "paid", imposta paid_date
    if update_dict.get("status") == "paid" and not update_dict.get("paid_date"):
        update_dict["paid_date"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.fees.update_one(
        {"id": fee_id, "client_id": client_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Onorario non trovato")
    
    fee = await db.fees.find_one({"id": fee_id}, {"_id": 0})
    return FeeResponse(**fee)


@client_fees_router.delete("/{fee_id}")
async def delete_fee(client_id: str, fee_id: str, user: dict = Depends(require_commercialista)):
    """Elimina un onorario"""
    db = get_db()
    result = await db.fees.delete_one({"id": fee_id, "client_id": client_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Onorario non trovato")
    
    await log_activity("eliminazione_onorario", f"Onorario {fee_id} eliminato", user["id"])
    
    return {"message": "Onorario eliminato"}


@client_fees_router.get("/summary")
async def get_fees_summary(client_id: str, user: dict = Depends(require_commercialista)):
    """Ottiene un riepilogo degli onorari di un cliente"""
    db = get_db()
    fees = await db.fees.find({"client_id": client_id}, {"_id": 0}).to_list(1000)
    
    total_pending = sum(f["amount"] for f in fees if f["status"] == "pending")
    total_paid = sum(f["amount"] for f in fees if f["status"] == "paid")
    total = total_pending + total_paid
    
    return {
        "total": total,
        "total_paid": total_paid,
        "total_pending": total_pending,
        "count_total": len(fees),
        "count_paid": len([f for f in fees if f["status"] == "paid"]),
        "count_pending": len([f for f in fees if f["status"] == "pending"])
    }
