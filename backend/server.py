from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, BackgroundTasks, Response
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import base64

# Import AI service
from ai_service import extract_text_from_pdf, analyze_document_with_ai, generate_standard_filename, search_documents_semantic
from email_service import send_welcome_email, notify_document_uploaded, notify_deadline_reminder, notify_new_note, send_invitation_email, send_generic_email
from chatbot_service import chat_with_assistant
from signing_service import sign_pdf_with_p12, verify_pdf_signature, save_certificate, list_certificates, delete_certificate
from storage_service import upload_file as cloud_upload, download_file as cloud_download, is_storage_enabled, init_b2_storage
from backup_service import create_client_backup, create_full_backup, export_database_json
import secrets

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise ValueError("JWT_SECRET must be set in environment variables")
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Commercialista predefinito (da env)
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', 'admin@example.com')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD')
if not ADMIN_PASSWORD:
    raise ValueError("ADMIN_PASSWORD must be set in environment variables")

# Categorie cartelle predefinite per l'organizzazione documenti
DEFAULT_FOLDER_CATEGORIES = [
    {"id": "documenti", "name": "Documenti", "icon": "file-text", "color": "#6b7280", "is_default": True, "order": 1},
    {"id": "agencia_tributaria", "name": "Agencia Tributaria", "icon": "landmark", "color": "#dc2626", "is_default": True, "order": 2},
    {"id": "seguridad_social", "name": "Seguridad Social", "icon": "users", "color": "#2563eb", "is_default": True, "order": 3},
    {"id": "ayuntamiento", "name": "Ayuntamiento", "icon": "building-2", "color": "#16a34a", "is_default": True, "order": 4},
    {"id": "contratti", "name": "Contratti", "icon": "file-signature", "color": "#9333ea", "is_default": True, "order": 5},
    {"id": "atti", "name": "Atti", "icon": "scale", "color": "#ca8a04", "is_default": True, "order": 6},
    {"id": "registro_mercantil", "name": "Registro Mercantil", "icon": "briefcase", "color": "#0891b2", "is_default": True, "order": 7},
]

app = FastAPI(title="Fiscal Tax Canarie API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str
    phone: Optional[str] = None
    codice_fiscale: Optional[str] = None
    nie: Optional[str] = None  # Número de Identidad de Extranjero
    nif: Optional[str] = None  # Número de Identificación Fiscal
    cif: Optional[str] = None  # Código de Identificación Fiscal (società)
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    cap: Optional[str] = None
    provincia: Optional[str] = None
    iban: Optional[str] = None  # Conto bancario
    regime_fiscale: Optional[str] = None
    tipo_attivita: Optional[str] = None
    tipo_cliente: Optional[str] = "autonomo"  # autonomo, societa, privato

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    full_name: str
    phone: Optional[str] = None
    codice_fiscale: Optional[str] = None
    nie: Optional[str] = None
    nif: Optional[str] = None
    cif: Optional[str] = None
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    cap: Optional[str] = None
    provincia: Optional[str] = None
    iban: Optional[str] = None
    regime_fiscale: Optional[str] = None
    tipo_attivita: Optional[str] = None
    tipo_cliente: Optional[str] = "autonomo"
    role: str
    stato: str = "attivo"
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class ClientUpdate(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    codice_fiscale: Optional[str] = None
    nie: Optional[str] = None
    nif: Optional[str] = None
    cif: Optional[str] = None
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    cap: Optional[str] = None
    provincia: Optional[str] = None
    iban: Optional[str] = None
    regime_fiscale: Optional[str] = None
    tipo_attivita: Optional[str] = None
    tipo_cliente: Optional[str] = None
    stato: Optional[str] = None
    note_interne: Optional[str] = None
    additional_emails: Optional[List[str]] = None  # Email aggiuntive
    bank_credentials: Optional[List[dict]] = None  # Credenziali bancarie

class ClientListCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: str = "#3caca4"

class ClientListResponse(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    color: str
    client_count: int = 0
    created_at: str

class DocumentCreate(BaseModel):
    title: str
    description: Optional[str] = None
    category: str
    client_id: str
    folder_category: Optional[str] = "documenti"  # Categoria cartella
    document_year: Optional[int] = None  # Anno del documento

class DocumentResponse(BaseModel):
    id: str
    title: str
    description: Optional[str] = None
    category: str
    client_id: str
    file_name: str
    file_data: Optional[str] = None
    uploaded_by: str
    created_at: str
    ai_description: Optional[str] = None
    tags: List[str] = []
    folder_category: Optional[str] = "documenti"  # Categoria cartella
    document_year: Optional[int] = None  # Anno del documento

# Modelli per le categorie cartella
class FolderCategoryCreate(BaseModel):
    name: str
    icon: Optional[str] = "folder"
    color: Optional[str] = "#6b7280"

class FolderCategoryResponse(BaseModel):
    id: str
    name: str
    icon: str
    color: str
    is_default: bool
    order: int
    created_at: Optional[str] = None
    created_by: Optional[str] = None

class PayslipResponse(BaseModel):
    id: str
    title: str
    month: str
    year: int
    client_id: str
    file_name: str
    file_data: Optional[str] = None
    uploaded_by: str
    created_at: str

class NoteCreate(BaseModel):
    title: str
    content: str
    client_id: str
    is_internal: bool = False

class NoteResponse(BaseModel):
    id: str
    title: str
    content: str
    client_id: str
    is_internal: bool
    created_by: str
    created_at: str
    updated_at: str

# ==================== TICKET MODELS ====================

class TicketMessageCreate(BaseModel):
    content: str

class TicketMessage(BaseModel):
    id: str
    content: str
    sender_id: str
    sender_name: str
    sender_role: str  # "cliente" o "commercialista"
    created_at: str

class TicketCreate(BaseModel):
    subject: str
    content: str

class TicketUpdate(BaseModel):
    status: Optional[str] = None  # "aperto", "chiuso", "archiviato"

class TicketResponse(BaseModel):
    id: str
    subject: str
    client_id: str
    client_name: Optional[str] = None
    status: str  # "aperto", "chiuso", "archiviato"
    messages: List[TicketMessage] = []
    created_by: str
    created_at: str
    updated_at: str
    closed_at: Optional[str] = None

class DeadlineCreate(BaseModel):
    title: str
    description: str
    due_date: str
    category: str
    is_recurring: bool = False
    recurrence_type: Optional[str] = None  # mensile, trimestrale, annuale
    recurrence_end_date: Optional[str] = None  # Data fine ricorrenza
    applies_to_all: bool = False
    client_ids: List[str] = []
    list_ids: List[str] = []  # Assegna a liste di clienti
    status: str = "da_fare"  # da_fare, in_lavorazione, completata, scaduta
    priority: str = "normale"  # bassa, normale, alta, urgente
    modello_tributario_id: Optional[str] = None
    send_reminders: bool = True  # Invia promemoria automatici
    reminder_days: List[int] = [7, 3, 1, 0]  # Giorni prima della scadenza

class DeadlineResponse(BaseModel):
    id: str
    title: str
    description: str
    due_date: str
    category: str
    is_recurring: bool
    recurrence_type: Optional[str] = None
    recurrence_end_date: Optional[str] = None
    applies_to_all: bool
    client_ids: List[str]
    list_ids: List[str] = []
    status: str
    priority: str
    modello_tributario_id: Optional[str] = None
    send_reminders: bool = True
    reminder_days: List[int] = [7, 3, 1, 0]
    last_reminder_sent: Optional[str] = None
    next_occurrence: Optional[str] = None
    created_at: Optional[str] = None

class ClientInListResponse(BaseModel):
    id: str
    email: Optional[str] = None  # Può essere None per clienti invitati
    email_notifica: Optional[str] = None  # Email per notifiche (sempre presente)
    full_name: str
    phone: Optional[str] = None
    codice_fiscale: Optional[str] = None
    nie: Optional[str] = None
    nif: Optional[str] = None
    cif: Optional[str] = None
    tipo_cliente: Optional[str] = "autonomo"
    stato: str = "attivo"  # attivo, invitato, sospeso, inattivo
    created_at: str
    documents_count: int = 0
    payslips_count: int = 0
    notes_count: int = 0
    lists: List[str] = []  # IDs delle liste di appartenenza

# Modello Tributario
class ModelloTributarioCreate(BaseModel):
    codice: str  # es: "Modelo-303", "IGIC"
    nome: str
    descrizione: str
    a_cosa_serve: str
    chi_deve_presentarlo: str
    periodicita: str  # trimestrale, mensile, annuale
    scadenza_tipica: str
    documenti_necessari: List[str] = []
    note_operative: Optional[str] = None
    video_youtube: Optional[str] = None  # URL video YouTube
    link_approfondimento: Optional[str] = None  # URL pagina esterna di approfondimento

class ModelloTributarioResponse(BaseModel):
    id: str
    codice: str
    nome: str
    descrizione: str
    a_cosa_serve: str
    chi_deve_presentarlo: str
    periodicita: str
    scadenza_tipica: str
    documenti_necessari: List[str]
    note_operative: Optional[str] = None
    video_youtube: Optional[str] = None
    video_thumbnail: Optional[str] = None  # Thumbnail calcolata
    link_approfondimento: Optional[str] = None
    created_at: str

# ==================== FEE (ONORARI) MODELS ====================

class FeeCreate(BaseModel):
    description: str
    amount: float
    due_date: Optional[str] = None  # Opzionale - richiesto solo per certi tipi
    status: str = "pending"  # pending, paid, overdue
    notes: Optional[str] = None
    fee_type: str = "standard"  # standard, consulenza, pratica, dichiarazione, iguala_buste_paga, iguala_contabilita, iguala_domicilio
    is_recurring: bool = False
    recurring_month: Optional[str] = None  # YYYY-MM per Iguala

class FeeUpdate(BaseModel):
    description: Optional[str] = None
    amount: Optional[float] = None
    due_date: Optional[str] = None
    status: Optional[str] = None
    paid_date: Optional[str] = None
    notes: Optional[str] = None
    fee_type: Optional[str] = None
    is_recurring: Optional[bool] = None
    recurring_month: Optional[str] = None

class FeeResponse(BaseModel):
    id: str
    client_id: str
    description: str
    amount: float
    due_date: Optional[str] = None
    status: str
    paid_date: Optional[str] = None
    notes: Optional[str] = None
    fee_type: str = "standard"
    is_recurring: bool = False
    recurring_month: Optional[str] = None
    created_at: str

# ==================== CLIENT CREATION MODELS ====================

class ClientCreate(BaseModel):
    """Creazione cliente - crea immediatamente la cartella cliente"""
    full_name: str  # Nome obbligatorio
    email: Optional[EmailStr] = None  # Email opzionale - se fornita, invia invito
    tipo_cliente: Optional[str] = "autonomo"  # autonomo, societa, privato
    # Campi anagrafica opzionali
    phone: Optional[str] = None
    codice_fiscale: Optional[str] = None
    nie: Optional[str] = None
    nif: Optional[str] = None
    cif: Optional[str] = None
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    cap: Optional[str] = None
    provincia: Optional[str] = None
    iban: Optional[str] = None
    regime_fiscale: Optional[str] = None
    tipo_attivita: Optional[str] = None
    note_interne: Optional[str] = None
    send_invite: Optional[bool] = True  # Se true e email presente, invia invito

# Alias per retrocompatibilità
class ClientInvite(BaseModel):
    """Invito cliente - crea immediatamente la cartella cliente"""
    email: EmailStr  # Email per notifiche e matching
    full_name: Optional[str] = None
    tipo_cliente: Optional[str] = "autonomo"  # autonomo, societa, privato
    # Campi anagrafica opzionali (precompilabili dall'admin)
    phone: Optional[str] = None
    codice_fiscale: Optional[str] = None
    nie: Optional[str] = None  # Per matching
    nif: Optional[str] = None
    cif: Optional[str] = None  # Per matching
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    cap: Optional[str] = None
    provincia: Optional[str] = None
    iban: Optional[str] = None
    regime_fiscale: Optional[str] = None
    tipo_attivita: Optional[str] = None
    note_interne: Optional[str] = None

class CompleteRegistration(BaseModel):
    token: str
    email: EmailStr  # Email scelta dal cliente per l'account
    password: str
    full_name: Optional[str] = None

# ==================== BANK CREDENTIALS MODELS ====================

class BankCredential(BaseModel):
    bank_entity_id: str  # ID dell'entità bancaria
    username: str
    password: str

class BankCredentialUpdate(BaseModel):
    bank_entity_id: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None

class BankEntity(BaseModel):
    name: str  # Nome banca (Revolut, Caixa, etc.)

# ==================== CONSULENTE LAVORO MODELS ====================

class ConsulenteCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str

class ConsulenteInvite(BaseModel):
    """Invito per consulente del lavoro"""
    email: EmailStr
    full_name: str

class ClientAssignment(BaseModel):
    client_ids: List[str]  # Lista di ID clienti da assegnare

# ==================== EMPLOYEE (DIPENDENTI) MODELS ====================

class EmployeeHireRequest(BaseModel):
    """Richiesta di assunzione dipendente dal cliente"""
    full_name: str
    start_date: str  # Data inizio lavoro
    job_title: str  # Mansione
    work_hours: str  # Orario di lavoro (es. "8:00-17:00")
    work_location: str  # Luogo di lavoro
    work_days: str  # Giorni lavorativi (es. "Lunedì-Venerdì")
    weekly_hours: Optional[int] = None  # Ore settimanali di lavoro (max 40)
    notes: Optional[str] = None  # Note aggiuntive

class EmployeeUpdate(BaseModel):
    full_name: Optional[str] = None
    job_title: Optional[str] = None
    work_hours: Optional[str] = None
    work_location: Optional[str] = None
    work_days: Optional[str] = None
    weekly_hours: Optional[int] = None  # Ore settimanali di lavoro (max 40)
    status: Optional[str] = None  # active, terminated, pending
    termination_date: Optional[str] = None
    notes: Optional[str] = None

class EmployeeTerminationRequest(BaseModel):
    """Richiesta di licenziamento dipendente"""
    reason: Optional[str] = None
    termination_date: str

# ==================== NOTIFICATION MODELS ====================

class NotificationCreate(BaseModel):
    type: str  # hire_request, termination_request, document_upload
    title: str
    message: str
    client_id: str
    employee_id: Optional[str] = None

# Email destinatari per notifiche dipendenti
EMPLOYEE_NOTIFICATION_EMAILS = [
    "amministrazione@fiscaltaxcanarie.com",
    "segreteria@fiscaltaxcanarie.com",
    "bruno@fiscaltaxcanarie.com",
    "francesco@fiscaltaxcanarie.com"
]

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Utente non trovato")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token scaduto")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token non valido")

async def require_commercialista(user: dict = Depends(get_current_user)):
    if user["role"] != "commercialista":
        raise HTTPException(status_code=403, detail="Accesso riservato ai commercialisti")
    return user

async def require_commercialista_or_consulente(user: dict = Depends(get_current_user)):
    """Permette accesso a commercialista o consulente del lavoro"""
    if user["role"] not in ["commercialista", "consulente_lavoro"]:
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    return user

async def require_any_role(user: dict = Depends(get_current_user)):
    """Permette accesso a commercialista, consulente o cliente"""
    if user["role"] not in ["commercialista", "consulente_lavoro", "cliente"]:
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    return user

async def require_consulente_lavoro(user: dict = Depends(get_current_user)):
    """Solo per consulenti del lavoro"""
    if user["role"] != "consulente_lavoro":
        raise HTTPException(status_code=403, detail="Accesso riservato ai consulenti del lavoro")
    return user

async def get_accessible_clients(user: dict) -> List[str]:
    """Restituisce la lista di ID clienti accessibili per l'utente"""
    if user["role"] == "commercialista":
        # Commercialista vede tutti i clienti
        clients = await db.users.find({"role": "cliente"}, {"id": 1, "_id": 0}).to_list(10000)
        return [c["id"] for c in clients]
    elif user["role"] == "consulente_lavoro":
        # Consulente vede solo i clienti assegnati
        return user.get("assigned_clients", [])
    return []

# ==================== INIT ADMIN ====================

async def init_admin_user():
    """Crea l'account commercialista predefinito se non esiste"""
    existing = await db.users.find_one({"email": ADMIN_EMAIL})
    if not existing:
        admin_id = str(uuid.uuid4())
        admin_doc = {
            "id": admin_id,
            "email": ADMIN_EMAIL,
            "password": hash_password(ADMIN_PASSWORD),
            "full_name": "Fiscal Tax Canarie",
            "phone": None,
            "codice_fiscale": None,
            "indirizzo": None,
            "regime_fiscale": None,
            "tipo_attivita": None,
            "role": "commercialista",
            "stato": "attivo",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_doc)
        logger.info(f"Account commercialista creato: {ADMIN_EMAIL}")

async def init_default_bank_entities():
    """Crea le entità bancarie predefinite se non esistono"""
    default_banks = ["Revolut", "Caixa", "Santander", "BBVA", "Cajamar"]
    for bank_name in default_banks:
        existing = await db.bank_entities.find_one({"name": bank_name})
        if not existing:
            await db.bank_entities.insert_one({
                "id": str(uuid.uuid4()),
                "name": bank_name,
                "is_default": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    logger.info("Entità bancarie predefinite inizializzate")

@app.on_event("startup")
async def startup_event():
    await init_admin_user()
    await init_default_modelli_tributari()
    await init_default_bank_entities()

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    """Registrazione SOLO per clienti"""
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email già registrata")
    
    # Sempre cliente - nessuna opzione per commercialista
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": user_data.email,
        "password": hash_password(user_data.password),
        "full_name": user_data.full_name,
        "phone": user_data.phone,
        "codice_fiscale": user_data.codice_fiscale,
        "indirizzo": user_data.indirizzo,
        "regime_fiscale": user_data.regime_fiscale,
        "tipo_attivita": user_data.tipo_attivita,
        "role": "cliente",  # SEMPRE cliente
        "stato": "attivo",
        "note_interne": "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Log attività
    await log_activity("registrazione", f"Nuovo cliente registrato: {user_data.email}", user_id)
    
    # Invia email di benvenuto (in background)
    try:
        await send_welcome_email(user_data.email, user_data.full_name)
    except Exception as e:
        logger.error(f"Errore invio email benvenuto: {e}")
    
    token = create_token(user_id, user_data.email, "cliente")
    user_response = UserResponse(
        id=user_id,
        email=user_data.email,
        full_name=user_data.full_name,
        phone=user_data.phone,
        codice_fiscale=user_data.codice_fiscale,
        indirizzo=user_data.indirizzo,
        regime_fiscale=user_data.regime_fiscale,
        tipo_attivita=user_data.tipo_attivita,
        role="cliente",
        stato="attivo",
        created_at=user_doc["created_at"]
    )
    return TokenResponse(access_token=token, user=user_response)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    
    # Log attività
    await log_activity("login", f"Accesso utente: {credentials.email}", user["id"])
    
    token = create_token(user["id"], user["email"], user["role"])
    user_response = UserResponse(
        id=user["id"],
        email=user["email"],
        full_name=user["full_name"],
        phone=user.get("phone"),
        codice_fiscale=user.get("codice_fiscale"),
        indirizzo=user.get("indirizzo"),
        regime_fiscale=user.get("regime_fiscale"),
        tipo_attivita=user.get("tipo_attivita"),
        role=user["role"],
        stato=user.get("stato", "attivo"),
        created_at=user["created_at"]
    )
    return TokenResponse(access_token=token, user=user_response)

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(
        id=user["id"],
        email=user["email"],
        full_name=user["full_name"],
        phone=user.get("phone"),
        codice_fiscale=user.get("codice_fiscale"),
        nie=user.get("nie"),
        nif=user.get("nif"),
        cif=user.get("cif"),
        indirizzo=user.get("indirizzo"),
        citta=user.get("citta"),
        cap=user.get("cap"),
        provincia=user.get("provincia"),
        iban=user.get("iban"),
        regime_fiscale=user.get("regime_fiscale"),
        tipo_attivita=user.get("tipo_attivita"),
        tipo_cliente=user.get("tipo_cliente", "autonomo"),
        role=user["role"],
        stato=user.get("stato", "attivo"),
        created_at=user["created_at"]
    )

class ClientSelfUpdate(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    codice_fiscale: Optional[str] = None
    nie: Optional[str] = None
    nif: Optional[str] = None
    cif: Optional[str] = None
    indirizzo: Optional[str] = None
    citta: Optional[str] = None
    cap: Optional[str] = None
    provincia: Optional[str] = None
    iban: Optional[str] = None

@api_router.put("/auth/me")
async def update_my_profile(update_data: ClientSelfUpdate, user: dict = Depends(get_current_user)):
    """Permette al cliente di modificare la propria anagrafica"""
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    await db.users.update_one({"id": user["id"]}, {"$set": update_dict})
    
    await log_activity("aggiornamento_profilo", f"Profilo aggiornato da {user['full_name']}", user["id"])
    
    # Restituisci i dati aggiornati
    updated_user = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password": 0})
    return {"message": "Profilo aggiornato", "user": updated_user}

# ==================== PASSWORD RESET ====================

class PasswordResetRequest(BaseModel):
    email: EmailStr

class PasswordResetConfirm(BaseModel):
    token: str
    new_password: str

@api_router.post("/auth/forgot-password")
async def forgot_password(request: PasswordResetRequest):
    """Richiesta di reset password - invia email con link"""
    user = await db.users.find_one({"email": request.email.lower()}, {"_id": 0})
    
    # Non rivelare se l'email esiste o meno per sicurezza
    if not user:
        return {"message": "Se l'email esiste nel sistema, riceverai un link per reimpostare la password"}
    
    # Genera token di reset (valido 1 ora)
    reset_token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
    
    # Salva il token nel database
    await db.password_resets.delete_many({"user_id": user["id"]})  # Rimuovi token precedenti
    await db.password_resets.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "email": user["email"],
        "token": reset_token,
        "expires_at": expires_at.isoformat(),
        "used": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Invia email con link di reset
    frontend_url = os.environ.get("FRONTEND_URL", "https://app.fiscaltaxcanarie.com")
    reset_link = f"{frontend_url}/reset-password?token={reset_token}"
    
    try:
        email_subject = "Recupero Password - Fiscal Tax Canarie"
        email_content = f"""
        <h2>Richiesta di Recupero Password</h2>
        <p>Ciao {user.get('full_name', '')},</p>
        <p>Hai richiesto di reimpostare la password del tuo account Fiscal Tax Canarie.</p>
        <p>Clicca sul link seguente per creare una nuova password:</p>
        <p><a href="{reset_link}" style="display: inline-block; background-color: #3caca4; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; font-weight: bold;">Reimposta Password</a></p>
        <p>Oppure copia e incolla questo link nel browser:</p>
        <p>{reset_link}</p>
        <hr>
        <p><strong>Importante:</strong> Questo link è valido per 1 ora.</p>
        <p>Se non hai richiesto tu il reset della password, ignora questa email.</p>
        <br>
        <p>Fiscal Tax Canarie</p>
        """
        send_generic_email(user["email"], email_subject, email_content)
    except Exception as e:
        logger.error(f"Errore invio email reset password: {e}")
    
    return {"message": "Se l'email esiste nel sistema, riceverai un link per reimpostare la password"}

@api_router.post("/auth/reset-password")
async def reset_password(request: PasswordResetConfirm):
    """Conferma reset password con token"""
    # Trova il token
    reset_record = await db.password_resets.find_one({
        "token": request.token,
        "used": False
    }, {"_id": 0})
    
    if not reset_record:
        raise HTTPException(status_code=400, detail="Link non valido o già utilizzato")
    
    # Verifica scadenza
    expires_at = datetime.fromisoformat(reset_record["expires_at"].replace("Z", "+00:00"))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Il link è scaduto. Richiedi un nuovo reset")
    
    # Valida la nuova password
    if len(request.new_password) < 6:
        raise HTTPException(status_code=400, detail="La password deve essere di almeno 6 caratteri")
    
    # Aggiorna la password
    hashed_password = bcrypt.hashpw(request.new_password.encode('utf-8'), bcrypt.gensalt())
    await db.users.update_one(
        {"id": reset_record["user_id"]},
        {"$set": {"password": hashed_password.decode('utf-8')}}
    )
    
    # Marca il token come usato
    await db.password_resets.update_one(
        {"token": request.token},
        {"$set": {"used": True, "used_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    await log_activity("reset_password", f"Password reimpostata per {reset_record['email']}", reset_record["user_id"])
    
    return {"message": "Password reimpostata con successo. Ora puoi accedere con la nuova password"}

@api_router.get("/auth/verify-reset-token")
async def verify_reset_token(token: str):
    """Verifica se un token di reset è valido"""
    reset_record = await db.password_resets.find_one({
        "token": token,
        "used": False
    }, {"_id": 0})
    
    if not reset_record:
        raise HTTPException(status_code=400, detail="Link non valido o già utilizzato")
    
    expires_at = datetime.fromisoformat(reset_record["expires_at"].replace("Z", "+00:00"))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Il link è scaduto")
    
    return {"valid": True, "email": reset_record["email"]}

# ==================== CLIENTS ROUTES (COMMERCIALISTA) ====================

@api_router.get("/clients", response_model=List[ClientInListResponse])
async def get_clients(
    tipo_cliente: Optional[str] = None,
    list_id: Optional[str] = None,
    user: dict = Depends(require_commercialista)
):
    query = {"role": "cliente"}
    if tipo_cliente:
        query["tipo_cliente"] = tipo_cliente
    if list_id:
        query["lists"] = list_id
    
    clients = await db.users.find(query, {"_id": 0, "password": 0}).to_list(1000)
    
    # Ottimizzazione: Usa aggregation pipeline per contare documenti/payslips/notes in batch
    # invece di N+1 query separate per ogni cliente
    client_ids = [c["id"] for c in clients]
    
    # Aggregazione per documenti
    docs_counts = {}
    docs_agg = await db.documents.aggregate([
        {"$match": {"client_id": {"$in": client_ids}}},
        {"$group": {"_id": "$client_id", "count": {"$sum": 1}}}
    ]).to_list(1000)
    for item in docs_agg:
        docs_counts[item["_id"]] = item["count"]
    
    # Aggregazione per payslips
    payslips_counts = {}
    payslips_agg = await db.payslips.aggregate([
        {"$match": {"client_id": {"$in": client_ids}}},
        {"$group": {"_id": "$client_id", "count": {"$sum": 1}}}
    ]).to_list(1000)
    for item in payslips_agg:
        payslips_counts[item["_id"]] = item["count"]
    
    # Aggregazione per notes
    notes_counts = {}
    notes_agg = await db.notes.aggregate([
        {"$match": {"client_id": {"$in": client_ids}}},
        {"$group": {"_id": "$client_id", "count": {"$sum": 1}}}
    ]).to_list(1000)
    for item in notes_agg:
        notes_counts[item["_id"]] = item["count"]
    
    result = []
    for client in clients:
        result.append(ClientInListResponse(
            id=client["id"],
            email=client.get("email"),  # Può essere None per invitati
            email_notifica=client.get("email_notifica"),  # Email per notifiche
            full_name=client.get("full_name", ""),
            phone=client.get("phone"),
            codice_fiscale=client.get("codice_fiscale"),
            nie=client.get("nie"),
            nif=client.get("nif"),
            cif=client.get("cif"),
            tipo_cliente=client.get("tipo_cliente", "autonomo"),
            stato=client.get("stato", "attivo"),
            created_at=client.get("created_at", datetime.now(timezone.utc).isoformat()),
            documents_count=docs_counts.get(client["id"], 0),
            payslips_count=payslips_counts.get(client["id"], 0),
            notes_count=notes_counts.get(client["id"], 0),
            lists=client.get("lists", [])
        ))
    return result

@api_router.get("/clients/{client_id}")
async def get_client(client_id: str, user: dict = Depends(require_commercialista)):
    client = await db.users.find_one({"id": client_id, "role": "cliente"}, {"_id": 0, "password": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    return client

@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, update_data: ClientUpdate, user: dict = Depends(require_commercialista)):
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    # Se viene cambiato il tipo_cliente, aggiorna automaticamente le liste
    if "tipo_cliente" in update_dict:
        tipo_cliente = update_dict["tipo_cliente"]
        
        # Trova o crea la lista corrispondente
        list_name_map = {
            "autonomo": "Autonomi",
            "societa": "Società", 
            "privato": "Privati"
        }
        
        list_name = list_name_map.get(tipo_cliente)
        if list_name:
            # Cerca la lista esistente
            existing_list = await db.client_lists.find_one({"name": list_name})
            
            if not existing_list:
                # Crea la lista automaticamente
                list_colors = {"Autonomi": "#3b82f6", "Società": "#8b5cf6", "Privati": "#10b981"}
                new_list_id = str(uuid.uuid4())
                await db.client_lists.insert_one({
                    "id": new_list_id,
                    "name": list_name,
                    "description": f"Lista automatica per clienti {tipo_cliente}",
                    "color": list_colors.get(list_name, "#3caca4"),
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
                existing_list = {"id": new_list_id}
            
            # Rimuovi il cliente da altre liste di tipo (autonomi, società, privati)
            all_type_lists = await db.client_lists.find(
                {"name": {"$in": list(list_name_map.values())}}
            ).to_list(10)
            for lst in all_type_lists:
                await db.users.update_one(
                    {"id": client_id},
                    {"$pull": {"lists": lst["id"]}}
                )
            
            # Aggiungi il cliente alla lista corretta
            await db.users.update_one(
                {"id": client_id},
                {"$addToSet": {"lists": existing_list["id"]}}
            )
    
    result = await db.users.update_one({"id": client_id, "role": "cliente"}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    await log_activity("modifica_cliente", f"Cliente {client_id} modificato", user["id"])
    return {"message": "Cliente aggiornato"}

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, permanent: bool = False, user: dict = Depends(require_commercialista)):
    if permanent:
        # Eliminazione permanente - elimina anche tutti i dati correlati
        client = await db.users.find_one({"id": client_id, "role": "cliente"})
        if not client:
            raise HTTPException(status_code=404, detail="Cliente non trovato")
        
        # Elimina documenti, buste paga, note e scadenze del cliente
        await db.documents.delete_many({"client_id": client_id})
        await db.payslips.delete_many({"client_id": client_id})
        await db.notes.delete_many({"client_id": client_id})
        await db.deadlines.update_many(
            {"client_ids": client_id},
            {"$pull": {"client_ids": client_id}}
        )
        
        # Elimina l'utente
        await db.users.delete_one({"id": client_id})
        await log_activity("eliminazione_permanente_cliente", f"Cliente {client['full_name']} eliminato permanentemente", user["id"])
        return {"message": "Cliente e tutti i dati correlati eliminati permanentemente"}
    else:
        # Archivia invece di eliminare
        result = await db.users.update_one(
            {"id": client_id, "role": "cliente"},
            {"$set": {"stato": "cessato"}}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Cliente non trovato")
        
        await log_activity("archiviazione_cliente", f"Cliente {client_id} archiviato", user["id"])
        return {"message": "Cliente archiviato"}

# ==================== CLIENT LISTS (CATEGORIE) ====================

@api_router.get("/client-lists")
async def get_client_lists(user: dict = Depends(require_commercialista)):
    lists = await db.client_lists.find({}, {"_id": 0}).to_list(100)
    
    # Aggiungi conteggio clienti per lista
    for lst in lists:
        count = await db.users.count_documents({"lists": lst["id"], "role": "cliente"})
        lst["client_count"] = count
    
    return lists

@api_router.post("/client-lists")
async def create_client_list(list_data: ClientListCreate, user: dict = Depends(require_commercialista)):
    list_id = str(uuid.uuid4())
    new_list = {
        "id": list_id,
        "name": list_data.name,
        "description": list_data.description,
        "color": list_data.color,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.client_lists.insert_one(new_list)
    await log_activity("creazione_lista", f"Lista '{list_data.name}' creata", user["id"])
    return {"id": list_id, "message": "Lista creata"}

@api_router.put("/client-lists/{list_id}")
async def update_client_list(list_id: str, list_data: ClientListCreate, user: dict = Depends(require_commercialista)):
    result = await db.client_lists.update_one(
        {"id": list_id},
        {"$set": {"name": list_data.name, "description": list_data.description, "color": list_data.color}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lista non trovata")
    return {"message": "Lista aggiornata"}

@api_router.delete("/client-lists/{list_id}")
async def delete_client_list(list_id: str, user: dict = Depends(require_commercialista)):
    # Rimuovi la lista da tutti i clienti
    await db.users.update_many({"lists": list_id}, {"$pull": {"lists": list_id}})
    # Elimina la lista
    result = await db.client_lists.delete_one({"id": list_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lista non trovata")
    await log_activity("eliminazione_lista", f"Lista {list_id} eliminata", user["id"])
    return {"message": "Lista eliminata"}

@api_router.post("/client-lists/{list_id}/clients/{client_id}")
async def add_client_to_list(list_id: str, client_id: str, user: dict = Depends(require_commercialista)):
    # Verifica che la lista esista
    lst = await db.client_lists.find_one({"id": list_id})
    if not lst:
        raise HTTPException(status_code=404, detail="Lista non trovata")
    
    # Aggiungi cliente alla lista
    result = await db.users.update_one(
        {"id": client_id, "role": "cliente"},
        {"$addToSet": {"lists": list_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    return {"message": "Cliente aggiunto alla lista"}

@api_router.delete("/client-lists/{list_id}/clients/{client_id}")
async def remove_client_from_list(list_id: str, client_id: str, user: dict = Depends(require_commercialista)):
    result = await db.users.update_one(
        {"id": client_id, "role": "cliente"},
        {"$pull": {"lists": list_id}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    return {"message": "Cliente rimosso dalla lista"}

@api_router.post("/client-lists/{list_id}/send-notification")
async def send_notification_to_list(
    list_id: str,
    subject: str = Form(...),
    content: str = Form(...),
    user: dict = Depends(require_commercialista)
):
    """Invia una notifica email a tutti i clienti di una lista"""
    # Verifica che la lista esista
    lst = await db.client_lists.find_one({"id": list_id})
    if not lst:
        raise HTTPException(status_code=404, detail="Lista non trovata")
    
    # Trova tutti i clienti nella lista
    clients = await db.users.find(
        {"lists": list_id, "role": "cliente", "stato": "attivo"},
        {"_id": 0, "email": 1, "full_name": 1}
    ).to_list(1000)
    
    if not clients:
        return {"success": False, "error": "Nessun cliente attivo nella lista"}
    
    # Invia email a tutti i clienti
    sent_count = 0
    errors = []
    for client in clients:
        try:
            result = await notify_new_note(
                client_email=client["email"],
                client_name=client["full_name"],
                note_title=subject,
                note_content=content
            )
            if result.get("success"):
                sent_count += 1
            else:
                errors.append(f"{client['email']}: {result.get('error')}")
        except Exception as e:
            errors.append(f"{client['email']}: {str(e)}")
    
    await log_activity(
        "notifica_lista",
        f"Notifica '{subject}' inviata a lista '{lst['name']}': {sent_count}/{len(clients)} email inviate",
        user["id"]
    )
    
    return {
        "success": True,
        "sent_count": sent_count,
        "total_clients": len(clients),
        "list_name": lst["name"],
        "errors": errors if errors else None
    }

# ==================== FOLDER CATEGORIES ROUTES ====================

@api_router.get("/folder-categories")
async def get_folder_categories(user: dict = Depends(get_current_user)):
    """Ottiene tutte le categorie cartella (predefinite + personalizzate)"""
    # Carica le categorie personalizzate dal database
    custom_categories = await db.folder_categories.find({}, {"_id": 0}).to_list(100)
    
    # Unisci con le categorie predefinite
    all_categories = []
    
    # Aggiungi le predefinite
    for cat in DEFAULT_FOLDER_CATEGORIES:
        all_categories.append({
            **cat,
            "created_at": None,
            "created_by": None
        })
    
    # Aggiungi le personalizzate
    for cat in custom_categories:
        all_categories.append(cat)
    
    # Ordina per 'order'
    all_categories.sort(key=lambda x: x.get("order", 999))
    
    return all_categories

@api_router.post("/folder-categories")
async def create_folder_category(category: FolderCategoryCreate, user: dict = Depends(require_commercialista)):
    """Crea una nuova categoria cartella personalizzata"""
    # Verifica che non esista già una categoria con lo stesso nome
    existing = await db.folder_categories.find_one({"name": {"$regex": f"^{category.name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Una categoria con questo nome esiste già")
    
    # Verifica che non sia una categoria predefinita
    for default_cat in DEFAULT_FOLDER_CATEGORIES:
        if default_cat["name"].lower() == category.name.lower():
            raise HTTPException(status_code=400, detail="Non puoi creare una categoria con un nome predefinito")
    
    # Trova l'ordine più alto esistente
    max_order = len(DEFAULT_FOLDER_CATEGORIES)
    last_custom = await db.folder_categories.find_one(sort=[("order", -1)])
    if last_custom:
        max_order = max(max_order, last_custom.get("order", 0))
    
    cat_id = str(uuid.uuid4())
    # Crea ID slug dal nome
    slug_id = category.name.lower().replace(" ", "_").replace("-", "_")
    
    cat_doc = {
        "id": slug_id,
        "name": category.name,
        "icon": category.icon or "folder",
        "color": category.color or "#6b7280",
        "is_default": False,
        "order": max_order + 1,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    
    await db.folder_categories.insert_one(cat_doc)
    
    await log_activity(
        "categoria_creata",
        f"Creata categoria cartella: {category.name}",
        user["id"]
    )
    
    return {**cat_doc, "_id": None}

@api_router.delete("/folder-categories/{category_id}")
async def delete_folder_category(category_id: str, user: dict = Depends(require_commercialista)):
    """Elimina una categoria cartella personalizzata (non le predefinite)"""
    # Verifica che non sia una categoria predefinita
    for default_cat in DEFAULT_FOLDER_CATEGORIES:
        if default_cat["id"] == category_id:
            raise HTTPException(status_code=400, detail="Non puoi eliminare una categoria predefinita")
    
    result = await db.folder_categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoria non trovata")
    
    # Sposta i documenti di questa categoria in "documenti"
    await db.documents.update_many(
        {"folder_category": category_id},
        {"$set": {"folder_category": "documenti"}}
    )
    
    return {"success": True, "message": "Categoria eliminata"}

@api_router.put("/documents/{doc_id}/category")
async def update_document_category(
    doc_id: str,
    folder_category: str = Form(...),
    document_year: Optional[int] = Form(None),
    user: dict = Depends(get_current_user)
):
    """Aggiorna la categoria cartella e/o l'anno di un documento"""
    # Verifica che il documento esista
    document = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    # Verifica permessi: admin può modificare tutto, cliente solo i propri
    if user["role"] == "cliente" and document["client_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    update_data = {"folder_category": folder_category}
    if document_year:
        update_data["document_year"] = document_year
    
    await db.documents.update_one({"id": doc_id}, {"$set": update_data})
    
    await log_activity(
        "documento_categoria_modificata",
        f"Documento {doc_id} spostato in categoria {folder_category}",
        user["id"]
    )
    
    return {"success": True, "message": "Categoria aggiornata"}

@api_router.get("/clients/{client_id}/documents/by-folder")
async def get_client_documents_by_folder(
    client_id: str,
    year: Optional[int] = None,
    user: dict = Depends(get_current_user)
):
    """
    Ottiene i documenti di un cliente organizzati per cartelle.
    Restituisce la struttura delle cartelle con conteggio documenti.
    """
    # Verifica permessi
    if user["role"] == "cliente" and user["id"] != client_id:
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    if user["role"] == "consulente_lavoro":
        # Verifica che il cliente sia assegnato al consulente
        if client_id not in user.get("assigned_clients", []):
            raise HTTPException(status_code=403, detail="Cliente non assegnato")
    
    # Query base
    query = {"client_id": client_id}
    if year:
        query["document_year"] = year
    
    # Ottieni tutti i documenti del cliente
    documents = await db.documents.find(
        query, 
        {"_id": 0, "file_data": 0, "versions_history": 0}
    ).to_list(1000)
    
    # Ottieni tutte le categorie
    custom_categories = await db.folder_categories.find({}, {"_id": 0}).to_list(100)
    all_categories = list(DEFAULT_FOLDER_CATEGORIES) + custom_categories
    
    # Organizza documenti per cartella
    folders = {}
    for cat in all_categories:
        cat_id = cat["id"]
        folders[cat_id] = {
            "id": cat_id,
            "name": cat["name"],
            "icon": cat["icon"],
            "color": cat["color"],
            "is_default": cat.get("is_default", False),
            "order": cat.get("order", 999),
            "documents": [],
            "document_count": 0,
            "years": set()
        }
    
    # Aggiungi documenti alle cartelle
    for doc in documents:
        folder_cat = doc.get("folder_category", "documenti")
        if folder_cat not in folders:
            folder_cat = "documenti"  # Fallback
        
        folders[folder_cat]["documents"].append(doc)
        folders[folder_cat]["document_count"] += 1
        
        # Traccia gli anni
        doc_year = doc.get("document_year")
        if doc_year:
            folders[folder_cat]["years"].add(doc_year)
    
    # Converti sets in liste e ordina
    result = []
    for folder_id, folder_data in folders.items():
        folder_data["years"] = sorted(list(folder_data["years"]), reverse=True)
        result.append(folder_data)
    
    # Ordina per 'order'
    result.sort(key=lambda x: x["order"])
    
    # Calcola anni disponibili globalmente
    all_years = set()
    for doc in documents:
        if doc.get("document_year"):
            all_years.add(doc["document_year"])
    
    return {
        "folders": result,
        "total_documents": len(documents),
        "available_years": sorted(list(all_years), reverse=True)
    }

# ==================== DOCUMENTS ROUTES ====================

@api_router.post("/documents")
async def upload_document(
    title: str = Form(...),
    description: str = Form(None),
    category: str = Form(...),
    client_id: str = Form(...),
    file: UploadFile = File(...),
    user: dict = Depends(require_commercialista)
):
    file_content = await file.read()
    file_base64 = base64.b64encode(file_content).decode('utf-8')
    
    doc_id = str(uuid.uuid4())
    document = {
        "id": doc_id,
        "title": title,
        "description": description,
        "category": category,
        "client_id": client_id,
        "file_name": file.filename,
        "file_data": file_base64,
        "file_type": file.content_type,
        "uploaded_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "ai_description": None,
        "tags": [],
        "version": 1,
        "versions_history": []
    }
    await db.documents.insert_one(document)
    
    await log_activity("caricamento_documento", f"Documento {title} caricato per cliente {client_id}", user["id"])
    
    return {"id": doc_id, "message": "Documento caricato con successo"}

@api_router.get("/documents", response_model=List[DocumentResponse])
async def get_documents(client_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if user["role"] == "cliente":
        query["client_id"] = user["id"]
    elif client_id:
        query["client_id"] = client_id
    
    documents = await db.documents.find(query, {"_id": 0, "file_data": 0, "versions_history": 0}).to_list(1000)
    return [DocumentResponse(**doc) for doc in documents]

@api_router.get("/documents/search")
async def search_documents(
    q: str,
    client_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Ricerca semantica nei documenti usando AI"""
    query = {}
    if user["role"] == "cliente":
        query["client_id"] = user["id"]
    elif client_id:
        query["client_id"] = client_id
    
    # Recupera documenti (senza file_data per performance)
    documents = await db.documents.find(
        query, 
        {"_id": 0, "file_data": 0, "versions_history": 0, "extracted_text": 0}
    ).to_list(100)
    
    if not documents:
        return []
    
    # Esegui ricerca semantica
    results = await search_documents_semantic(q, documents)
    
    return results

@api_router.get("/documents/pending-verification")
async def get_documents_pending_verification(user: dict = Depends(require_commercialista)):
    """Recupera documenti che necessitano verifica manuale"""
    documents = await db.documents.find(
        {"needs_verification": True},
        {"_id": 0, "file_data": 0, "extracted_text": 0}
    ).to_list(100)
    return documents

@api_router.get("/documents/{doc_id}")
async def get_document(doc_id: str, user: dict = Depends(get_current_user)):
    document = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    if user["role"] == "cliente" and document["client_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Se il file è su B2, scaricalo e aggiungi file_data
    if document.get("storage_path") and not document.get("file_data"):
        file_content, content_type = await cloud_download(document["storage_path"])
        if file_content:
            document["file_data"] = base64.b64encode(file_content).decode('utf-8')
            document["file_type"] = content_type
    
    return document

@api_router.delete("/documents/{doc_id}")
async def delete_document(doc_id: str, user: dict = Depends(get_current_user)):
    """Elimina un documento. Admin può eliminare tutti, cliente solo i propri."""
    document = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    
    if not document:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    # Verifica permessi
    if user["role"] == "cliente" and document.get("client_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Non puoi eliminare questo documento")
    
    # Se il file è su B2, eliminalo
    if document.get("storage_path"):
        from storage_service import delete_file as cloud_delete
        await cloud_delete(document["storage_path"])
    
    result = await db.documents.delete_one({"id": doc_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    await log_activity("eliminazione_documento", f"Documento {doc_id} eliminato", user["id"])
    return {"message": "Documento eliminato"}

@api_router.get("/documents/{doc_id}/preview")
async def preview_document(
    doc_id: str, 
    token: Optional[str] = None,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(lambda: None)
):
    """
    Restituisce il documento per la preview inline.
    Supporta PDF, immagini e file di testo.
    Accetta il token sia come header Authorization che come query parameter.
    """
    from fastapi.responses import Response
    
    # Ottieni l'utente dal token (query param o header)
    auth_token = token
    if not auth_token:
        # Prova a ottenere dal header
        from fastapi import Request
        raise HTTPException(status_code=401, detail="Token non fornito")
    
    try:
        payload = jwt.decode(auth_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Token non valido")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Utente non trovato")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token scaduto")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Token non valido")
    
    document = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    # Verifica permessi
    if user["role"] == "cliente" and document.get("client_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Recupera il contenuto del file
    file_content = None
    content_type = "application/octet-stream"
    
    if document.get("storage_path"):
        # File su B2
        file_content, content_type = await cloud_download(document["storage_path"])
    elif document.get("file_data"):
        # File in base64 nel DB
        file_content = base64.b64decode(document["file_data"])
        content_type = document.get("file_type", "application/octet-stream")
    
    if not file_content:
        raise HTTPException(status_code=404, detail="Contenuto del file non trovato")
    
    # Determina il content type dal nome file se non disponibile
    file_name = document.get("file_name", "").lower()
    if file_name.endswith(".pdf"):
        content_type = "application/pdf"
    elif file_name.endswith((".jpg", ".jpeg")):
        content_type = "image/jpeg"
    elif file_name.endswith(".png"):
        content_type = "image/png"
    elif file_name.endswith(".gif"):
        content_type = "image/gif"
    elif file_name.endswith(".webp"):
        content_type = "image/webp"
    elif file_name.endswith(".txt"):
        content_type = "text/plain"
    elif file_name.endswith((".doc", ".docx")):
        content_type = "application/msword"
    elif file_name.endswith((".xls", ".xlsx")):
        content_type = "application/vnd.ms-excel"
    
    # Restituisce il file con header per visualizzazione inline
    return Response(
        content=file_content,
        media_type=content_type,
        headers={
            "Content-Disposition": f'inline; filename="{document.get("file_name", "document")}"',
            "Cache-Control": "private, max-age=3600"
        }
    )

# ==================== AI DOCUMENT ANALYSIS ====================

@api_router.post("/documents/upload-auto")
async def upload_document_with_ai(
    file: UploadFile = File(...),
    client_id: Optional[str] = Form(None),
    user: dict = Depends(require_commercialista)
):
    """
    Carica un documento e lo analizza automaticamente con AI.
    - Estrae testo dal PDF
    - Identifica tipo documento, modello tributario
    - Suggerisce cliente se non specificato
    - Genera descrizione e tag
    - Rinomina automaticamente il file
    """
    file_content = await file.read()
    file_base64 = base64.b64encode(file_content).decode('utf-8')
    
    # Estrai testo dal PDF
    extracted_text = ""
    if file.content_type == "application/pdf" or file.filename.lower().endswith(".pdf"):
        extracted_text = await extract_text_from_pdf(file_base64)
    
    # Recupera lista clienti per matching
    clients = await db.users.find({"role": "cliente"}, {"_id": 0, "password": 0}).to_list(1000)
    
    # Analizza con AI
    ai_result = await analyze_document_with_ai(
        file_content=extracted_text or f"File: {file.filename}",
        file_name=file.filename,
        clients_list=clients
    )
    
    # Determina cliente
    final_client_id = client_id
    client_confidence = "manuale"
    
    if not final_client_id and ai_result.get("success") and ai_result.get("cliente_identificato", {}).get("id"):
        final_client_id = ai_result["cliente_identificato"]["id"]
        client_confidence = ai_result["cliente_identificato"].get("confidenza", "bassa")
    
    # Se ancora nessun cliente, metti in "da verificare"
    needs_verification = not final_client_id or client_confidence == "bassa"
    
    # Ottieni info cliente per nome file
    client_name_for_file = None
    if final_client_id:
        client_info = await db.users.find_one({"id": final_client_id}, {"full_name": 1, "_id": 0})
        if client_info:
            client_name_for_file = client_info.get("full_name")
    
    # Genera nome file standardizzato automaticamente
    original_ext = "." + file.filename.rsplit(".", 1)[-1] if "." in file.filename else ".pdf"
    
    standardized_filename = generate_standard_filename(
        tipo_documento=ai_result.get("tipo_documento") if ai_result.get("success") else None,
        data_documento=ai_result.get("data_documento") if ai_result.get("success") else None,
        cliente_nome=client_name_for_file,
        riferimento=ai_result.get("periodo_riferimento") if ai_result.get("success") else None,
        original_extension=original_ext
    )
    
    # Usa nome suggerito dall'AI se diverso e valido
    suggested_filename = standardized_filename
    if ai_result.get("success") and ai_result.get("nome_file_suggerito"):
        ai_suggested = ai_result["nome_file_suggerito"]
        if ai_suggested and len(ai_suggested) > 10:  # Se nome AI è valido
            suggested_filename = ai_suggested
    
    # Determina categoria
    category = "altro"
    if ai_result.get("success") and ai_result.get("categoria_suggerita"):
        category = ai_result["categoria_suggerita"]
    
    # Determina folder_category (per organizzazione cartelle)
    folder_category = "documenti"  # Default
    if ai_result.get("success") and ai_result.get("folder_category"):
        folder_category = ai_result["folder_category"]
    
    # Determina anno documento
    document_year = None
    if ai_result.get("success"):
        # Prima prova anno_documento diretto dall'AI
        if ai_result.get("anno_documento"):
            document_year = ai_result["anno_documento"]
        # Poi prova a estrarre dalla data_documento
        elif ai_result.get("data_documento"):
            try:
                document_year = int(ai_result["data_documento"][:4])
            except (ValueError, TypeError):
                pass
        # Infine usa anno corrente come fallback
        if not document_year:
            document_year = datetime.now().year
    
    # Verifica se è una busta paga - se sì, salvala nella collection payslips
    is_payslip = ai_result.get("success") and ai_result.get("is_busta_paga", False)
    
    if is_payslip and final_client_id:
        # Crea record busta paga invece di documento
        payslip_id = str(uuid.uuid4())
        
        # Determina mese e anno dalla risposta AI
        mese = ai_result.get("mese_busta_paga", "Non specificato")
        anno = ai_result.get("anno_busta_paga", datetime.now().year)
        
        payslip = {
            "id": payslip_id,
            "title": ai_result.get("descrizione", file.filename),
            "month": mese,
            "year": anno,
            "client_id": final_client_id,
            "file_name": standardized_filename,
            "file_name_original": file.filename,
            "file_data": file_base64,
            "file_type": file.content_type,
            "uploaded_by": user["id"],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "auto_classified": True,  # Indica che è stato classificato automaticamente dall'AI
            "ai_analysis": ai_result
        }
        await db.payslips.insert_one(payslip)
        
        await log_activity(
            "caricamento_busta_paga_ai", 
            f"Busta paga {file.filename} riconosciuta e archiviata automaticamente. Mese: {mese} {anno}",
            user["id"]
        )
        
        return {
            "id": payslip_id,
            "message": "Busta paga riconosciuta e archiviata automaticamente nella sezione Buste Paga",
            "document_type": "payslip",
            "ai_analysis": ai_result,
            "client_id": final_client_id,
            "month": mese,
            "year": anno
        }
    
    # Crea documento con nome file rinominato
    doc_id = str(uuid.uuid4())
    
    # Prova upload su B2, fallback su MongoDB se non disponibile
    storage_result = await cloud_upload(
        file_data=file_content,
        client_id=final_client_id or "unassigned",
        original_filename=standardized_filename,
        folder="documents"
    )
    
    document = {
        "id": doc_id,
        "title": ai_result.get("descrizione", file.filename) if ai_result.get("success") else file.filename,
        "description": ai_result.get("descrizione_estesa") if ai_result.get("success") else None,
        "category": category,
        "folder_category": folder_category,  # Categoria cartella per organizzazione
        "document_year": document_year,  # Anno documento estratto dall'AI
        "client_id": final_client_id,
        "file_name": standardized_filename,  # Nome rinominato automaticamente
        "file_name_original": file.filename,  # Nome originale per riferimento
        "file_name_suggested": suggested_filename,
        "file_type": file.content_type,
        "uploaded_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "ai_analysis": ai_result if ai_result.get("success") else None,
        "ai_description": ai_result.get("descrizione_estesa") if ai_result.get("success") else None,
        "tags": ai_result.get("tags", []) if ai_result.get("success") else [],
        "modello_tributario": ai_result.get("modello_tributario") if ai_result.get("success") else None,
        "data_documento": ai_result.get("data_documento") if ai_result.get("success") else None,
        "periodo_riferimento": ai_result.get("periodo_riferimento") if ai_result.get("success") else None,
        "needs_verification": needs_verification,
        "client_confidence": client_confidence,
        "version": 1,
        "versions_history": [],
        "extracted_text": extracted_text[:2000] if extracted_text else None
    }
    
    # Salva su B2 se disponibile, altrimenti su MongoDB
    if storage_result.get("success"):
        document["storage_path"] = storage_result["storage_path"]
        document["storage_provider"] = "b2"
        document["file_size"] = storage_result.get("size", len(file_content))
    else:
        # Fallback: salva in MongoDB
        document["file_data"] = file_base64
        document["storage_provider"] = "mongodb"
    
    await db.documents.insert_one(document)
    
    await log_activity(
        "caricamento_documento_ai", 
        f"Documento {file.filename} caricato con analisi AI. Cliente: {final_client_id or 'da assegnare'}, Confidenza: {client_confidence}",
        user["id"]
    )
    
    # NOTA: Notifiche email per upload documenti DISABILITATE su richiesta utente
    # Le notifiche automatiche per caricamento documenti non vengono inviate
    # Rimangono attive solo: notifiche scadenze, comunicazioni globali/amministrative
    
    return {
        "id": doc_id,
        "message": "Documento caricato e analizzato con successo",
        "ai_analysis": ai_result if ai_result.get("success") else None,
        "needs_verification": needs_verification,
        "client_id": final_client_id,
        "client_confidence": client_confidence,
        "suggested_filename": suggested_filename,
        "standardized_filename": standardized_filename,
        "original_filename": file.filename,
        "category": category,
        "folder_category": folder_category,
        "document_year": document_year
    }

@api_router.put("/documents/{doc_id}/verify")
async def verify_document(
    doc_id: str,
    client_id: str = Form(...),
    title: str = Form(None),
    category: str = Form(None),
    user: dict = Depends(require_commercialista)
):
    """Verifica e corregge l'assegnazione di un documento"""
    update_data = {
        "client_id": client_id,
        "needs_verification": False,
        "client_confidence": "verificato"
    }
    if title:
        update_data["title"] = title
    if category:
        update_data["category"] = category
    
    result = await db.documents.update_one({"id": doc_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    await log_activity("verifica_documento", f"Documento {doc_id} verificato e assegnato a {client_id}", user["id"])
    
    return {"message": "Documento verificato con successo"}

@api_router.post("/documents/upload-batch")
async def upload_documents_batch(
    files: List[UploadFile] = File(...),
    user: dict = Depends(require_commercialista)
):
    """
    Carica più documenti contemporaneamente e li analizza con AI.
    L'IA identifica automaticamente il cliente per ogni documento.
    I file vengono salvati su Backblaze B2.
    """
    if len(files) > 15:
        raise HTTPException(status_code=400, detail="Massimo 15 file per upload")
    
    results = []
    clients = await db.users.find({"role": "cliente"}, {"_id": 0, "password": 0}).to_list(1000)
    
    for file in files:
        try:
            file_content = await file.read()
            file_base64 = base64.b64encode(file_content).decode('utf-8')
            
            # Estrai testo dal PDF
            extracted_text = ""
            if file.content_type == "application/pdf" or file.filename.lower().endswith(".pdf"):
                extracted_text = await extract_text_from_pdf(file_base64)
            
            # Analizza con AI
            ai_result = await analyze_document_with_ai(
                file_content=extracted_text or f"File: {file.filename}",
                file_name=file.filename,
                clients_list=clients
            )
            
            # Determina cliente
            final_client_id = None
            client_confidence = "non_identificato"
            
            if ai_result.get("success") and ai_result.get("cliente_identificato", {}).get("id"):
                final_client_id = ai_result["cliente_identificato"]["id"]
                client_confidence = ai_result["cliente_identificato"].get("confidenza", "bassa")
            
            needs_verification = not final_client_id or client_confidence == "bassa"
            
            # Ottieni info cliente per nome file
            client_name_for_file = None
            if final_client_id:
                client_info = await db.users.find_one({"id": final_client_id}, {"full_name": 1, "_id": 0})
                if client_info:
                    client_name_for_file = client_info.get("full_name")
            
            # Genera nome file standardizzato
            original_ext = "." + file.filename.rsplit(".", 1)[-1] if "." in file.filename else ".pdf"
            
            standardized_filename = generate_standard_filename(
                tipo_documento=ai_result.get("tipo_documento") if ai_result.get("success") else None,
                data_documento=ai_result.get("data_documento") if ai_result.get("success") else None,
                cliente_nome=client_name_for_file,
                riferimento=ai_result.get("periodo_riferimento") if ai_result.get("success") else None,
                original_extension=original_ext
            )
            
            category = ai_result.get("categoria_suggerita", "altro") if ai_result.get("success") else "altro"
            
            # Upload su B2
            storage_result = {"success": False}
            try:
                storage_result = await cloud_upload(
                    file_content=file_content,
                    filename=standardized_filename,
                    content_type=file.content_type or "application/octet-stream",
                    folder=f"documents/{final_client_id or 'unassigned'}"
                )
            except Exception as e:
                logger.warning(f"Errore upload B2 per {file.filename}: {e}")
            
            # Crea documento
            doc_id = str(uuid.uuid4())
            document = {
                "id": doc_id,
                "title": ai_result.get("descrizione", file.filename) if ai_result.get("success") else file.filename,
                "description": ai_result.get("descrizione_estesa") if ai_result.get("success") else None,
                "category": category,
                "folder_category": category,
                "document_year": ai_result.get("anno_documento") if ai_result.get("success") else str(datetime.now().year),
                "client_id": final_client_id,
                "file_name": standardized_filename,
                "file_name_original": file.filename,
                "file_type": file.content_type,
                "uploaded_by": user["id"],
                "created_at": datetime.now(timezone.utc).isoformat(),
                "ai_analysis": ai_result if ai_result.get("success") else None,
                "ai_description": ai_result.get("descrizione_estesa") if ai_result.get("success") else None,
                "tags": ai_result.get("tags", []) if ai_result.get("success") else [],
                "modello_tributario": ai_result.get("modello_tributario") if ai_result.get("success") else None,
                "needs_verification": needs_verification,
                "client_confidence": client_confidence,
                "version": 1,
                "extracted_text": extracted_text[:2000] if extracted_text else None
            }
            
            # Salva su B2 se disponibile, altrimenti su MongoDB
            if storage_result.get("success"):
                document["storage_path"] = storage_result["storage_path"]
                document["storage_provider"] = "b2"
                document["file_size"] = storage_result.get("size", len(file_content))
            else:
                # Fallback: salva in MongoDB
                document["file_data"] = file_base64
                document["storage_provider"] = "mongodb"
            
            await db.documents.insert_one(document)
            
            results.append({
                "id": doc_id,
                "original_filename": file.filename,
                "standardized_filename": standardized_filename,
                "client_id": final_client_id,
                "client_name": client_name_for_file,
                "client_confidence": client_confidence,
                "needs_verification": needs_verification,
                "category": category,
                "success": True
            })
            
            # NOTA: Notifiche email per upload documenti DISABILITATE su richiesta utente
                    
        except Exception as e:
            logger.error(f"Errore elaborazione file {file.filename}: {e}")
            results.append({
                "original_filename": file.filename,
                "success": False,
                "error": str(e)
            })
    
    await log_activity(
        "caricamento_batch",
        f"Caricati {len([r for r in results if r.get('success')])} documenti su {len(files)}",
        user["id"]
    )
    
    return {
        "message": f"Elaborati {len(files)} documenti",
        "results": results,
        "needs_verification_count": len([r for r in results if r.get("needs_verification")]),
        "assigned_count": len([r for r in results if r.get("client_id") and not r.get("needs_verification")])
    }

@api_router.put("/documents/{doc_id}/rename")
async def rename_document(
    doc_id: str,
    new_filename: str = Form(...),
    user: dict = Depends(get_current_user)
):
    """Rinomina un documento - accessibile a commercialista e cliente (per i propri documenti)"""
    # Recupera il documento
    document = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    # Verifica permessi
    if user["role"] == "cliente":
        # Il cliente può rinominare solo i propri documenti
        if document.get("client_id") != user["id"]:
            raise HTTPException(status_code=403, detail="Non hai i permessi per rinominare questo documento")
    elif user["role"] not in ["commercialista", "consulente_lavoro"]:
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Mantieni l'estensione originale
    original_extension = ""
    if "." in document.get("file_name", ""):
        original_extension = "." + document["file_name"].rsplit(".", 1)[-1]
    
    # Se il nuovo nome non ha estensione, aggiungi quella originale
    if original_extension and not new_filename.endswith(original_extension):
        # Rimuovi eventuale estensione diversa dal nuovo nome
        if "." in new_filename:
            new_filename = new_filename.rsplit(".", 1)[0]
        new_filename = new_filename + original_extension
    
    result = await db.documents.update_one({"id": doc_id}, {"$set": {"file_name": new_filename}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    await log_activity("rinomina_documento", f"Documento {doc_id} rinominato in {new_filename}", user["id"])
    
    return {"message": "Documento rinominato con successo", "new_filename": new_filename}

# ==================== PAYSLIPS ROUTES ====================

@api_router.post("/payslips")
async def upload_payslip(
    title: str = Form(...),
    month: str = Form(...),
    year: int = Form(...),
    client_id: str = Form(...),
    file: UploadFile = File(...),
    user: dict = Depends(require_commercialista)
):
    file_content = await file.read()
    file_base64 = base64.b64encode(file_content).decode('utf-8')
    
    payslip_id = str(uuid.uuid4())
    payslip = {
        "id": payslip_id,
        "title": title,
        "month": month,
        "year": year,
        "client_id": client_id,
        "file_name": file.filename,
        "file_data": file_base64,
        "file_type": file.content_type,
        "uploaded_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payslips.insert_one(payslip)
    
    await log_activity("caricamento_busta_paga", f"Busta paga {month}/{year} caricata per cliente {client_id}", user["id"])
    
    return {"id": payslip_id, "message": "Busta paga caricata con successo"}

@api_router.get("/payslips", response_model=List[PayslipResponse])
async def get_payslips(client_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if user["role"] == "cliente":
        query["client_id"] = user["id"]
    elif client_id:
        query["client_id"] = client_id
    
    payslips = await db.payslips.find(query, {"_id": 0, "file_data": 0}).to_list(1000)
    return [PayslipResponse(**p) for p in payslips]

@api_router.get("/payslips/{payslip_id}")
async def get_payslip(payslip_id: str, user: dict = Depends(get_current_user)):
    payslip = await db.payslips.find_one({"id": payslip_id}, {"_id": 0})
    if not payslip:
        raise HTTPException(status_code=404, detail="Busta paga non trovata")
    
    if user["role"] == "cliente" and payslip["client_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    return payslip

@api_router.delete("/payslips/{payslip_id}")
async def delete_payslip(payslip_id: str, user: dict = Depends(require_commercialista)):
    result = await db.payslips.delete_one({"id": payslip_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Busta paga non trovata")
    return {"message": "Busta paga eliminata"}

# ==================== NOTES ROUTES ====================

@api_router.post("/notes", response_model=NoteResponse)
async def create_note(note_data: NoteCreate, user: dict = Depends(require_commercialista)):
    note_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    note = {
        "id": note_id,
        "title": note_data.title,
        "content": note_data.content,
        "client_id": note_data.client_id,
        "is_internal": note_data.is_internal,
        "created_by": user["id"],
        "created_at": now,
        "updated_at": now
    }
    await db.notes.insert_one(note)
    return NoteResponse(**note)

@api_router.get("/notes", response_model=List[NoteResponse])
async def get_notes(client_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if user["role"] == "cliente":
        query["client_id"] = user["id"]
        query["is_internal"] = False
    elif client_id:
        query["client_id"] = client_id
    
    notes = await db.notes.find(query, {"_id": 0}).to_list(1000)
    return [NoteResponse(**n) for n in notes]

@api_router.put("/notes/{note_id}", response_model=NoteResponse)
async def update_note(note_id: str, note_data: NoteCreate, user: dict = Depends(require_commercialista)):
    update_data = {
        "title": note_data.title,
        "content": note_data.content,
        "is_internal": note_data.is_internal,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.notes.update_one({"id": note_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Nota non trovata")
    
    note = await db.notes.find_one({"id": note_id}, {"_id": 0})
    return NoteResponse(**note)

@api_router.delete("/notes/{note_id}")
async def delete_note(note_id: str, user: dict = Depends(require_commercialista)):
    result = await db.notes.delete_one({"id": note_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nota non trovata")
    return {"message": "Nota eliminata"}

# ==================== TICKET ROUTES ====================

@api_router.post("/tickets", response_model=TicketResponse)
async def create_ticket(ticket_data: TicketCreate, user: dict = Depends(get_current_user)):
    """Crea un nuovo ticket - accessibile sia a clienti che admin"""
    ticket_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # Determina client_id
    if user["role"] == "cliente":
        client_id = user["id"]
    else:
        raise HTTPException(status_code=400, detail="Solo i clienti possono aprire ticket")
    
    # Primo messaggio del ticket
    first_message = {
        "id": str(uuid.uuid4()),
        "content": ticket_data.content,
        "sender_id": user["id"],
        "sender_name": user.get("full_name", "Cliente"),
        "sender_role": user["role"],
        "created_at": now
    }
    
    ticket = {
        "id": ticket_id,
        "subject": ticket_data.subject,
        "client_id": client_id,
        "status": "aperto",
        "messages": [first_message],
        "created_by": user["id"],
        "created_at": now,
        "updated_at": now,
        "closed_at": None
    }
    
    await db.tickets.insert_one(ticket)
    
    # Crea notifica per admin
    await db.admin_notifications.insert_one({
        "id": str(uuid.uuid4()),
        "type": "new_ticket",
        "title": "Nuovo Ticket",
        "message": f"Nuovo ticket da {user.get('full_name', 'Cliente')}: {ticket_data.subject}",
        "ticket_id": ticket_id,
        "client_id": client_id,
        "read": False,
        "created_at": now
    })
    
    # Get client name
    client = await db.users.find_one({"id": client_id}, {"_id": 0, "full_name": 1})
    ticket["client_name"] = client.get("full_name") if client else None
    
    return TicketResponse(**ticket)

@api_router.get("/tickets", response_model=List[TicketResponse])
async def get_tickets(
    client_id: Optional[str] = None,
    status: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Recupera tickets - filtrati per ruolo utente"""
    query = {}
    
    if user["role"] == "cliente":
        # Cliente vede solo i propri ticket
        query["client_id"] = user["id"]
    elif client_id:
        # Admin può filtrare per cliente
        query["client_id"] = client_id
    
    # Filtro per stato
    if status and status != "tutti":
        query["status"] = status
    
    tickets = await db.tickets.find(query, {"_id": 0}).sort("updated_at", -1).to_list(1000)
    
    # Aggiungi nomi clienti
    client_ids = list(set(t["client_id"] for t in tickets))
    clients = await db.users.find({"id": {"$in": client_ids}}, {"_id": 0, "id": 1, "full_name": 1}).to_list(1000)
    client_map = {c["id"]: c.get("full_name") for c in clients}
    
    for ticket in tickets:
        ticket["client_name"] = client_map.get(ticket["client_id"])
    
    return [TicketResponse(**t) for t in tickets]

@api_router.get("/tickets/{ticket_id}", response_model=TicketResponse)
async def get_ticket(ticket_id: str, user: dict = Depends(get_current_user)):
    """Recupera singolo ticket"""
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket non trovato")
    
    # Verifica accesso
    if user["role"] == "cliente" and ticket["client_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Aggiungi nome cliente
    client = await db.users.find_one({"id": ticket["client_id"]}, {"_id": 0, "full_name": 1})
    ticket["client_name"] = client.get("full_name") if client else None
    
    return TicketResponse(**ticket)

@api_router.post("/tickets/{ticket_id}/messages", response_model=TicketResponse)
async def add_ticket_message(ticket_id: str, message_data: TicketMessageCreate, user: dict = Depends(get_current_user)):
    """Aggiungi messaggio a un ticket - sia cliente che admin"""
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket non trovato")
    
    # Verifica accesso
    if user["role"] == "cliente" and ticket["client_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Verifica che il ticket non sia chiuso o archiviato
    if ticket["status"] in ["chiuso", "archiviato"]:
        raise HTTPException(status_code=400, detail="Non è possibile rispondere a un ticket chiuso o archiviato")
    
    now = datetime.now(timezone.utc).isoformat()
    
    new_message = {
        "id": str(uuid.uuid4()),
        "content": message_data.content,
        "sender_id": user["id"],
        "sender_name": user.get("full_name", "Utente"),
        "sender_role": user["role"],
        "created_at": now
    }
    
    await db.tickets.update_one(
        {"id": ticket_id},
        {
            "$push": {"messages": new_message},
            "$set": {"updated_at": now}
        }
    )
    
    # Notifica all'altra parte
    if user["role"] == "cliente":
        # Notifica admin
        await db.admin_notifications.insert_one({
            "id": str(uuid.uuid4()),
            "type": "ticket_reply",
            "title": "Risposta Ticket",
            "message": f"Nuova risposta da {user.get('full_name', 'Cliente')} al ticket: {ticket['subject']}",
            "ticket_id": ticket_id,
            "client_id": ticket["client_id"],
            "read": False,
            "created_at": now
        })
    
    # Recupera ticket aggiornato
    updated_ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    client = await db.users.find_one({"id": updated_ticket["client_id"]}, {"_id": 0, "full_name": 1})
    updated_ticket["client_name"] = client.get("full_name") if client else None
    
    return TicketResponse(**updated_ticket)

@api_router.put("/tickets/{ticket_id}/status", response_model=TicketResponse)
async def update_ticket_status(ticket_id: str, status_data: TicketUpdate, user: dict = Depends(require_commercialista)):
    """Aggiorna stato ticket - solo admin"""
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket non trovato")
    
    now = datetime.now(timezone.utc).isoformat()
    update_data = {"updated_at": now}
    
    if status_data.status:
        update_data["status"] = status_data.status
        if status_data.status == "chiuso":
            update_data["closed_at"] = now
    
    await db.tickets.update_one({"id": ticket_id}, {"$set": update_data})
    
    updated_ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    client = await db.users.find_one({"id": updated_ticket["client_id"]}, {"_id": 0, "full_name": 1})
    updated_ticket["client_name"] = client.get("full_name") if client else None
    
    return TicketResponse(**updated_ticket)

@api_router.delete("/tickets/{ticket_id}")
async def delete_ticket(ticket_id: str, user: dict = Depends(require_commercialista)):
    """Elimina ticket - solo admin"""
    result = await db.tickets.delete_one({"id": ticket_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ticket non trovato")
    return {"message": "Ticket eliminato"}

@api_router.get("/tickets/{ticket_id}/export-pdf")
async def export_ticket_pdf(ticket_id: str, user: dict = Depends(require_commercialista)):
    """Esporta ticket in PDF con storico completo"""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket non trovato")
    
    # Get client info
    client = await db.users.find_one({"id": ticket["client_id"]}, {"_id": 0})
    client_name = client.get("full_name", "N/A") if client else "N/A"
    client_email = client.get("email", "N/A") if client else "N/A"
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#0d9488'), spaceAfter=20, alignment=TA_CENTER)
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#64748b'), alignment=TA_CENTER)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#1e293b'), spaceBefore=15, spaceAfter=10)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#334155'), spaceAfter=6)
    message_client_style = ParagraphStyle('MessageClient', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#0f766e'), leftIndent=20, spaceAfter=10)
    message_admin_style = ParagraphStyle('MessageAdmin', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#1e40af'), leftIndent=20, spaceAfter=10)
    date_style = ParagraphStyle('Date', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94a3b8'))
    
    elements = []
    
    # Header
    elements.append(Paragraph("FISCAL TAX CANARIE", header_style))
    elements.append(Paragraph("Copia Certificata Ticket di Assistenza", header_style))
    elements.append(Spacer(1, 20))
    
    # Title
    elements.append(Paragraph(f"Ticket: {ticket['subject']}", title_style))
    elements.append(Spacer(1, 10))
    
    # Status badge
    status_colors = {"aperto": "#22c55e", "chiuso": "#64748b", "archiviato": "#ef4444"}
    status_labels = {"aperto": "APERTO", "chiuso": "CHIUSO", "archiviato": "ARCHIVIATO"}
    status = ticket.get("status", "aperto")
    elements.append(Paragraph(f"<font color='{status_colors.get(status, '#64748b')}'>● {status_labels.get(status, status.upper())}</font>", ParagraphStyle('Status', parent=styles['Normal'], fontSize=12, alignment=TA_CENTER)))
    elements.append(Spacer(1, 20))
    
    # Info table
    info_data = [
        ["Cliente:", client_name],
        ["Email:", client_email],
        ["Data Apertura:", ticket.get("created_at", "N/A")[:19].replace("T", " ") if ticket.get("created_at") else "N/A"],
        ["Ultimo Aggiornamento:", ticket.get("updated_at", "N/A")[:19].replace("T", " ") if ticket.get("updated_at") else "N/A"],
    ]
    if ticket.get("closed_at"):
        info_data.append(["Data Chiusura:", ticket["closed_at"][:19].replace("T", " ")])
    
    info_table = Table(info_data, colWidths=[4*cm, 12*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Separator
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
    elements.append(Spacer(1, 15))
    
    # Conversation history
    elements.append(Paragraph("STORICO CONVERSAZIONE", section_style))
    
    messages = ticket.get("messages", [])
    if messages:
        for i, msg in enumerate(messages):
            sender_name = msg.get("sender_name", "Utente")
            sender_role = msg.get("sender_role", "")
            created_at = msg.get("created_at", "")[:19].replace("T", " ") if msg.get("created_at") else ""
            content = msg.get("content", "")
            
            # Message header
            if sender_role == "cliente":
                elements.append(Paragraph(f"<b>🟢 {sender_name} (Cliente)</b> - {created_at}", date_style))
                elements.append(Paragraph(content, message_client_style))
            else:
                elements.append(Paragraph(f"<b>🔵 {sender_name} (Fiscal Tax Canarie)</b> - {created_at}", date_style))
                elements.append(Paragraph(content, message_admin_style))
            
            if i < len(messages) - 1:
                elements.append(Spacer(1, 5))
    else:
        elements.append(Paragraph("Nessun messaggio nel ticket.", normal_style))
    
    # Footer
    elements.append(Spacer(1, 30))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
    elements.append(Spacer(1, 10))
    footer_text = f"Documento generato il {datetime.now(timezone.utc).strftime('%d/%m/%Y alle %H:%M:%S')} UTC"
    elements.append(Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER)))
    elements.append(Paragraph("Fiscal Tax Canarie - Gestionale Studio Professionale", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER)))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Filename
    safe_subject = "".join(c for c in ticket['subject'] if c.isalnum() or c in (' ', '-', '_')).strip()[:30]
    filename = f"ticket_{safe_subject}_{ticket_id[:8]}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/admin/ticket-notifications")
async def get_ticket_notifications(user: dict = Depends(require_commercialista)):
    """Recupera notifiche ticket non lette per admin"""
    notifications = await db.admin_notifications.find(
        {"type": {"$in": ["new_ticket", "ticket_reply"]}, "read": False},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return notifications

@api_router.put("/admin/ticket-notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(require_commercialista)):
    """Segna notifica come letta"""
    await db.admin_notifications.update_one({"id": notification_id}, {"$set": {"read": True}})
    return {"message": "Notifica segnata come letta"}

# ==================== DEADLINES ROUTES ====================

@api_router.get("/deadlines", response_model=List[DeadlineResponse])
async def get_deadlines(user: dict = Depends(get_current_user)):
    query = {}
    if user["role"] == "cliente":
        query = {
            "$or": [
                {"applies_to_all": True},
                {"client_ids": user["id"]}
            ]
        }
    
    deadlines = await db.deadlines.find(query, {"_id": 0}).to_list(1000)
    
    # Aggiorna stato scadenze scadute e aggiungi campi mancanti
    now = datetime.now(timezone.utc).date()
    result = []
    for deadline in deadlines:
        # Aggiungi valori di default per campi mancanti
        deadline.setdefault("status", "da_fare")
        deadline.setdefault("priority", "normale")
        deadline.setdefault("modello_tributario_id", None)
        deadline.setdefault("created_at", None)
        
        due_date = datetime.fromisoformat(deadline["due_date"]).date() if isinstance(deadline["due_date"], str) else deadline["due_date"]
        if due_date < now and deadline.get("status") not in ["completata", "scaduta"]:
            deadline["status"] = "scaduta"
            await db.deadlines.update_one({"id": deadline["id"]}, {"$set": {"status": "scaduta", "priority": deadline["priority"]}})
        
        result.append(DeadlineResponse(**deadline))
    
    return result

class DeadlineCreateWithNotify(DeadlineCreate):
    send_notification: bool = False

async def get_clients_for_deadline(client_ids: List[str], list_ids: List[str]) -> List[dict]:
    """Ottieni tutti i clienti associati a una scadenza (da ID diretti e liste)"""
    all_client_ids = set(client_ids)
    
    # Aggiungi clienti dalle liste
    if list_ids:
        for list_id in list_ids:
            clients_in_list = await db.users.find(
                {"lists": list_id, "role": "cliente", "stato": "attivo"},
                {"id": 1, "_id": 0}
            ).to_list(1000)
            for c in clients_in_list:
                all_client_ids.add(c["id"])
    
    # Ottieni dettagli clienti
    clients = []
    for cid in all_client_ids:
        client = await db.users.find_one({"id": cid, "stato": "attivo"}, {"_id": 0, "password": 0})
        if client:
            clients.append(client)
    
    return clients

def calculate_next_occurrence(current_date: str, recurrence_type: str) -> str:
    """Calcola la prossima occorrenza di una scadenza ricorrente"""
    from dateutil.relativedelta import relativedelta
    
    current = datetime.fromisoformat(current_date.replace('Z', '+00:00') if 'Z' in current_date else current_date)
    if isinstance(current, str):
        current = datetime.strptime(current[:10], "%Y-%m-%d")
    
    if recurrence_type == "mensile":
        next_date = current + relativedelta(months=1)
    elif recurrence_type == "trimestrale":
        next_date = current + relativedelta(months=3)
    elif recurrence_type == "annuale":
        next_date = current + relativedelta(years=1)
    else:
        return current_date
    
    return next_date.strftime("%Y-%m-%d")

@api_router.post("/deadlines", response_model=DeadlineResponse)
async def create_deadline(deadline_data: DeadlineCreateWithNotify, user: dict = Depends(require_commercialista)):
    deadline_id = str(uuid.uuid4())
    
    # Calcola prossima occorrenza se ricorrente
    next_occurrence = None
    if deadline_data.is_recurring and deadline_data.recurrence_type:
        next_occurrence = calculate_next_occurrence(deadline_data.due_date, deadline_data.recurrence_type)
    
    deadline = {
        "id": deadline_id,
        "title": deadline_data.title,
        "description": deadline_data.description,
        "due_date": deadline_data.due_date,
        "category": deadline_data.category,
        "is_recurring": deadline_data.is_recurring,
        "recurrence_type": deadline_data.recurrence_type,
        "recurrence_end_date": deadline_data.recurrence_end_date,
        "applies_to_all": deadline_data.applies_to_all,
        "client_ids": deadline_data.client_ids,
        "list_ids": deadline_data.list_ids,
        "status": deadline_data.status,
        "priority": deadline_data.priority,
        "modello_tributario_id": deadline_data.modello_tributario_id,
        "send_reminders": deadline_data.send_reminders,
        "reminder_days": deadline_data.reminder_days,
        "last_reminder_sent": None,
        "next_occurrence": next_occurrence,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.deadlines.insert_one(deadline)
    
    await log_activity("creazione_scadenza", f"Scadenza {deadline_data.title} creata", user["id"])
    
    # Invia notifica email ai clienti assegnati (se richiesto)
    if deadline_data.send_notification:
        clients = await get_clients_for_deadline(deadline_data.client_ids, deadline_data.list_ids)
        for client_data in clients:
            try:
                if client_data.get("email"):
                    await notify_deadline_reminder(
                        client_email=client_data["email"],
                        client_name=client_data["full_name"],
                        deadline_title=deadline_data.title,
                        deadline_date=deadline_data.due_date,
                        deadline_description=deadline_data.description
                    )
                    logger.info(f"Email notifica scadenza inviata a {client_data['email']}")
            except Exception as e:
                logger.error(f"Errore invio email notifica scadenza: {e}")
    
    return DeadlineResponse(**deadline)

@api_router.post("/deadlines/send-reminders")
async def trigger_deadline_reminders(user: dict = Depends(require_commercialista)):
    """Invia manualmente i promemoria per le scadenze imminenti"""
    result = await send_deadline_reminders_batch()
    return result

async def send_deadline_reminders_batch():
    """Task per inviare promemoria automatici delle scadenze"""
    today = datetime.now(timezone.utc).date()
    sent_count = 0
    errors = []
    
    # Trova tutte le scadenze attive con promemoria abilitati
    deadlines = await db.deadlines.find({
        "status": {"$in": ["da_fare", "in_lavorazione"]},
        "send_reminders": True
    }, {"_id": 0}).to_list(1000)
    
    for deadline in deadlines:
        try:
            due_date = datetime.strptime(deadline["due_date"][:10], "%Y-%m-%d").date()
            days_until = (due_date - today).days
            reminder_days = deadline.get("reminder_days", [7, 3, 1, 0])
            
            # Controlla se oggi è un giorno di promemoria
            if days_until in reminder_days and days_until >= 0:
                # Controlla se già inviato oggi
                last_sent = deadline.get("last_reminder_sent")
                if last_sent:
                    last_sent_date = datetime.fromisoformat(last_sent.replace('Z', '+00:00')).date()
                    if last_sent_date == today:
                        continue
                
                # Ottieni clienti associati
                clients = await get_clients_for_deadline(
                    deadline.get("client_ids", []),
                    deadline.get("list_ids", [])
                )
                
                # Invia promemoria a tutti i clienti
                for client in clients:
                    try:
                        if client.get("email"):
                            days_text = "oggi" if days_until == 0 else f"tra {days_until} giorni"
                            await notify_deadline_reminder(
                                client_email=client["email"],
                                client_name=client["full_name"],
                                deadline_title=f"[PROMEMORIA] {deadline['title']}",
                                deadline_date=deadline["due_date"],
                                deadline_description=f"La scadenza è {days_text}. {deadline.get('description', '')}"
                            )
                            sent_count += 1
                    except Exception as e:
                        errors.append(f"{client.get('email')}: {str(e)}")
                
                # Aggiorna last_reminder_sent
                await db.deadlines.update_one(
                    {"id": deadline["id"]},
                    {"$set": {"last_reminder_sent": datetime.now(timezone.utc).isoformat()}}
                )
                
        except Exception as e:
            logger.error(f"Errore elaborazione scadenza {deadline.get('id')}: {e}")
            errors.append(str(e))
    
    # Gestisci scadenze ricorrenti - crea nuove occorrenze
    await handle_recurring_deadlines()
    
    return {
        "success": True,
        "reminders_sent": sent_count,
        "deadlines_processed": len(deadlines),
        "errors": errors if errors else None
    }

async def handle_recurring_deadlines():
    """Gestisce la creazione di nuove occorrenze per scadenze ricorrenti"""
    today = datetime.now(timezone.utc).date()
    
    # Trova scadenze ricorrenti completate che devono essere rigenerate
    recurring_deadlines = await db.deadlines.find({
        "is_recurring": True,
        "status": "completata",
        "next_occurrence": {"$ne": None}
    }, {"_id": 0}).to_list(100)
    
    for deadline in recurring_deadlines:
        try:
            # Verifica se c'è una data di fine ricorrenza
            if deadline.get("recurrence_end_date"):
                end_date = datetime.strptime(deadline["recurrence_end_date"][:10], "%Y-%m-%d").date()
                if today > end_date:
                    continue
            
            next_date = deadline.get("next_occurrence")
            if not next_date:
                continue
            
            next_occurrence_date = datetime.strptime(next_date[:10], "%Y-%m-%d").date()
            
            # Se la prossima occorrenza è passata o oggi, crea la nuova scadenza
            if next_occurrence_date <= today:
                new_deadline_id = str(uuid.uuid4())
                new_next = calculate_next_occurrence(next_date, deadline.get("recurrence_type", "mensile"))
                
                new_deadline = {
                    "id": new_deadline_id,
                    "title": deadline["title"],
                    "description": deadline.get("description", ""),
                    "due_date": next_date,
                    "category": deadline.get("category", "altro"),
                    "is_recurring": True,
                    "recurrence_type": deadline.get("recurrence_type"),
                    "recurrence_end_date": deadline.get("recurrence_end_date"),
                    "applies_to_all": deadline.get("applies_to_all", False),
                    "client_ids": deadline.get("client_ids", []),
                    "list_ids": deadline.get("list_ids", []),
                    "status": "da_fare",
                    "priority": deadline.get("priority", "normale"),
                    "modello_tributario_id": deadline.get("modello_tributario_id"),
                    "send_reminders": deadline.get("send_reminders", True),
                    "reminder_days": deadline.get("reminder_days", [7, 3, 1, 0]),
                    "last_reminder_sent": None,
                    "next_occurrence": new_next,
                    "parent_deadline_id": deadline["id"],
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.deadlines.insert_one(new_deadline)
                
                # Aggiorna la scadenza originale
                await db.deadlines.update_one(
                    {"id": deadline["id"]},
                    {"$set": {"next_occurrence": new_next}}
                )
                
                logger.info(f"Creata nuova occorrenza scadenza ricorrente: {deadline['title']}")
                
        except Exception as e:
            logger.error(f"Errore gestione scadenza ricorrente {deadline.get('id')}: {e}")

@api_router.put("/deadlines/{deadline_id}", response_model=DeadlineResponse)
async def update_deadline(deadline_id: str, deadline_data: DeadlineCreate, user: dict = Depends(require_commercialista)):
    update_dict = deadline_data.model_dump()
    result = await db.deadlines.update_one({"id": deadline_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Scadenza non trovata")
    
    deadline = await db.deadlines.find_one({"id": deadline_id}, {"_id": 0})
    return DeadlineResponse(**deadline)

@api_router.patch("/deadlines/{deadline_id}/status")
async def update_deadline_status(deadline_id: str, status: str = Form(...), user: dict = Depends(get_current_user)):
    if status not in ["da_fare", "in_lavorazione", "completata", "scaduta"]:
        raise HTTPException(status_code=400, detail="Stato non valido")
    
    result = await db.deadlines.update_one({"id": deadline_id}, {"$set": {"status": status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Scadenza non trovata")
    
    await log_activity("modifica_stato_scadenza", f"Scadenza {deadline_id} stato cambiato a {status}", user["id"])
    
    return {"message": "Stato aggiornato"}

@api_router.delete("/deadlines/{deadline_id}")
async def delete_deadline(deadline_id: str, user: dict = Depends(require_commercialista)):
    result = await db.deadlines.delete_one({"id": deadline_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Scadenza non trovata")
    return {"message": "Scadenza eliminata"}

# ==================== MODELLI TRIBUTARI ROUTES ====================

async def init_default_modelli_tributari():
    """Inizializza i modelli tributari predefiniti"""
    count = await db.modelli_tributari.count_documents({})
    if count > 0:
        return
    
    modelli = [
        {
            "id": str(uuid.uuid4()),
            "codice": "Modelo-303",
            "nome": "Dichiarazione IVA Trimestrale",
            "descrizione": "Modello per la dichiarazione trimestrale dell'IVA (Impuesto sobre el Valor Añadido)",
            "a_cosa_serve": "Dichiarare l'IVA a debito o a credito del trimestre. Permette di calcolare la differenza tra IVA incassata e IVA pagata.",
            "chi_deve_presentarlo": "Tutti i soggetti passivi IVA: imprenditori, professionisti e società che realizzano operazioni soggette ad IVA.",
            "periodicita": "Trimestrale",
            "scadenza_tipica": "20 aprile, 20 luglio, 20 ottobre, 30 gennaio",
            "documenti_necessari": ["Fatture emesse", "Fatture ricevute", "Libro IVA vendite", "Libro IVA acquisti"],
            "conseguenze_mancata_presentazione": "Sanzioni dal 50% al 150% dell'importo non dichiarato. Interessi di mora. Possibili controlli fiscali.",
            "note_operative": "Verificare sempre la corrispondenza tra registri IVA e dichiarazione.",
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "codice": "Modelo-111",
            "nome": "Ritenute IRPF su Lavoro",
            "descrizione": "Dichiarazione trimestrale delle ritenute effettuate su redditi da lavoro dipendente e autonomo.",
            "a_cosa_serve": "Dichiarare e versare le ritenute IRPF effettuate su stipendi, compensi professionali e altri redditi.",
            "chi_deve_presentarlo": "Datori di lavoro, imprese e professionisti che effettuano pagamenti soggetti a ritenuta.",
            "periodicita": "Trimestrale",
            "scadenza_tipica": "20 aprile, 20 luglio, 20 ottobre, 20 gennaio",
            "documenti_necessari": ["Buste paga", "Fatture professionisti", "Registro ritenute"],
            "conseguenze_mancata_presentazione": "Sanzioni dal 50% al 150% delle ritenute non dichiarate. Responsabilità solidale.",
            "note_operative": "Conservare tutta la documentazione relativa ai pagamenti effettuati.",
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "codice": "Modelo-130",
            "nome": "Pagamento Frazionato IRPF",
            "descrizione": "Pagamento frazionato dell'IRPF per autonomi in regime di stima diretta.",
            "a_cosa_serve": "Anticipare trimestralmente l'IRPF dovuto sui redditi da attività economica.",
            "chi_deve_presentarlo": "Lavoratori autonomi e professionisti in regime di stima diretta (normale o semplificata).",
            "periodicita": "Trimestrale",
            "scadenza_tipica": "20 aprile, 20 luglio, 20 ottobre, 30 gennaio",
            "documenti_necessari": ["Libro entrate e uscite", "Fatture emesse", "Fatture ricevute"],
            "conseguenze_mancata_presentazione": "Sanzioni per mancato pagamento. Interessi di mora. Maggiori controlli nella dichiarazione annuale.",
            "note_operative": "Il 20% del rendimento netto trimestrale, con deduzione delle ritenute subite.",
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "codice": "IGIC",
            "nome": "Imposta Generale Indiretta Canarie",
            "descrizione": "Imposta indiretta delle Isole Canarie equivalente all'IVA peninsulare.",
            "a_cosa_serve": "Dichiarare l'IGIC a debito o credito. Le Canarie hanno un regime fiscale speciale con aliquote ridotte.",
            "chi_deve_presentarlo": "Tutti i soggetti passivi IGIC che operano nelle Isole Canarie.",
            "periodicita": "Trimestrale/Mensile",
            "scadenza_tipica": "20 del mese successivo al periodo",
            "documenti_necessari": ["Fatture con IGIC", "Registri IGIC", "Documenti import/export"],
            "conseguenze_mancata_presentazione": "Sanzioni equivalenti a quelle IVA. Perdita benefici fiscali canari.",
            "note_operative": "Aliquota generale 7%. Aliquote ridotte 0%, 3%. Verificare esenzioni specifiche Canarie.",
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "codice": "Modelo-390",
            "nome": "Riepilogo Annuale IVA",
            "descrizione": "Dichiarazione riepilogativa annuale delle operazioni IVA.",
            "a_cosa_serve": "Riepilogare tutte le operazioni IVA dell'anno. Confronto con le dichiarazioni trimestrali.",
            "chi_deve_presentarlo": "Tutti i soggetti passivi IVA obbligati a presentare il Modelo 303.",
            "periodicita": "Annuale",
            "scadenza_tipica": "30 gennaio dell'anno successivo",
            "documenti_necessari": ["Modelli 303 trimestrali", "Libri IVA completi", "Fatture annuali"],
            "conseguenze_mancata_presentazione": "Sanzioni fisse e proporzionali. Impossibilità di ottenere certificati fiscali.",
            "note_operative": "Deve coincidere con la somma dei quattro trimestri. Verificare operazioni intracomunitarie.",
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "codice": "Modelo-200",
            "nome": "Imposta sulle Società",
            "descrizione": "Dichiarazione annuale dell'Imposta sulle Società.",
            "a_cosa_serve": "Dichiarare il risultato fiscale della società e calcolare l'imposta dovuta.",
            "chi_deve_presentarlo": "Tutte le società e entità giuridiche residenti in Spagna.",
            "periodicita": "Annuale",
            "scadenza_tipica": "25 luglio (esercizio coincidente con anno solare)",
            "documenti_necessari": ["Bilancio", "Conto economico", "Libri contabili", "Documentazione fiscale"],
            "conseguenze_mancata_presentazione": "Sanzioni dal 50% al 150% della quota non dichiarata. Responsabilità amministratori.",
            "note_operative": "Aliquota generale 25%. Verificare incentivi ZEC per Canarie.",
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "codice": "Modelo-347",
            "nome": "Operazioni con Terzi",
            "descrizione": "Dichiarazione annuale delle operazioni con terzi superiori a 3.005,06€.",
            "a_cosa_serve": "Informare l'Agenzia delle Entrate delle operazioni significative con clienti e fornitori.",
            "chi_deve_presentarlo": "Tutti i soggetti che hanno realizzato operazioni superiori a 3.005,06€ con un singolo cliente/fornitore.",
            "periodicita": "Annuale",
            "scadenza_tipica": "28 febbraio",
            "documenti_necessari": ["Elenco clienti/fornitori", "Fatture superiori alla soglia", "NIF delle controparti"],
            "conseguenze_mancata_presentazione": "Sanzioni da 20€ per dato omesso fino a 20.000€. Controlli incrociati.",
            "note_operative": "Include operazioni in contanti superiori a 6.000€. Escluse operazioni intracomunitarie.",
            "created_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "id": str(uuid.uuid4()),
            "codice": "IRPF-Renta",
            "nome": "Dichiarazione Redditi Persone Fisiche",
            "descrizione": "Dichiarazione annuale dei redditi delle persone fisiche (IRPF).",
            "a_cosa_serve": "Dichiarare tutti i redditi percepiti nell'anno e calcolare l'imposta definitiva.",
            "chi_deve_presentarlo": "Persone fisiche residenti con redditi superiori alle soglie di esenzione.",
            "periodicita": "Annuale",
            "scadenza_tipica": "30 giugno",
            "documenti_necessari": ["Certificati ritenute", "Dati immobiliari", "Spese deducibili", "Dati familiari"],
            "conseguenze_mancata_presentazione": "Sanzioni dal 50% al 150%. Perdita deduzioni. Maggiorazione interessi.",
            "note_operative": "Verificare sempre residenza fiscale. Deduzioni per investimenti Canarie.",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    
    for modello in modelli:
        await db.modelli_tributari.insert_one(modello)
    
    # NON creare scadenze predefinite - il commercialista le assegna per tipo cliente

async def init_default_deadlines():
    """Inizializza le scadenze predefinite"""
    count = await db.deadlines.count_documents({})
    if count > 0:
        return
    
    deadlines = [
        {"title": "Modelo 303 - IVA Trimestrale Q1", "description": "Dichiarazione trimestrale IVA primo trimestre", "due_date": "2025-04-20", "category": "IVA", "status": "da_fare", "priority": "alta"},
        {"title": "Modelo 111 - Ritenute IRPF Q1", "description": "Dichiarazione trimestrale ritenute primo trimestre", "due_date": "2025-04-20", "category": "IRPF", "status": "da_fare", "priority": "alta"},
        {"title": "Modelo 130 - Pagamento Frazionato Q1", "description": "Pagamento frazionato IRPF primo trimestre", "due_date": "2025-04-20", "category": "IRPF", "status": "da_fare", "priority": "normale"},
        {"title": "IGIC Trimestrale Q1", "description": "Dichiarazione IGIC primo trimestre", "due_date": "2025-04-20", "category": "IGIC", "status": "da_fare", "priority": "alta"},
        {"title": "Modelo 303 - IVA Trimestrale Q2", "description": "Dichiarazione trimestrale IVA secondo trimestre", "due_date": "2025-07-20", "category": "IVA", "status": "da_fare", "priority": "alta"},
        {"title": "Modelo 200 - Imposta Società", "description": "Dichiarazione annuale imposta sulle società", "due_date": "2025-07-25", "category": "Società", "status": "da_fare", "priority": "urgente"},
        {"title": "IRPF - Dichiarazione Annuale", "description": "Dichiarazione dei redditi persone fisiche", "due_date": "2025-06-30", "category": "IRPF", "status": "da_fare", "priority": "urgente"},
        {"title": "Modelo 347 - Operazioni con Terzi", "description": "Dichiarazione annuale operazioni con terzi", "due_date": "2025-02-28", "category": "Informativa", "status": "da_fare", "priority": "normale"},
        {"title": "Modelo 390 - Riepilogo Annuale IVA", "description": "Dichiarazione riepilogativa annuale IVA", "due_date": "2025-01-30", "category": "IVA", "status": "da_fare", "priority": "alta"},
    ]
    
    for d in deadlines:
        deadline = {
            "id": str(uuid.uuid4()),
            "title": d["title"],
            "description": d["description"],
            "due_date": d["due_date"],
            "category": d["category"],
            "is_recurring": True,
            "applies_to_all": True,
            "client_ids": [],
            "status": d["status"],
            "priority": d["priority"],
            "modello_tributario_id": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.deadlines.insert_one(deadline)

def get_youtube_thumbnail(video_url: Optional[str]) -> Optional[str]:
    """Estrae l'ID del video YouTube e restituisce l'URL della thumbnail"""
    if not video_url:
        return None
    
    import re
    # Pattern per diversi formati URL YouTube
    patterns = [
        r'(?:youtube\.com/watch\?v=|youtu\.be/|youtube\.com/embed/)([a-zA-Z0-9_-]{11})',
        r'youtube\.com/v/([a-zA-Z0-9_-]{11})',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, video_url)
        if match:
            video_id = match.group(1)
            return f"https://img.youtube.com/vi/{video_id}/hqdefault.jpg"
    
    return None

@api_router.get("/modelli-tributari", response_model=List[ModelloTributarioResponse])
async def get_modelli_tributari(user: dict = Depends(get_current_user)):
    modelli = await db.modelli_tributari.find({}, {"_id": 0}).to_list(100)
    result = []
    for m in modelli:
        # Aggiungi thumbnail per video YouTube
        m["video_thumbnail"] = get_youtube_thumbnail(m.get("video_youtube"))
        # Rimuovi campo obsoleto se presente
        m.pop("conseguenze_mancata_presentazione", None)
        result.append(ModelloTributarioResponse(**m))
    return result

@api_router.get("/modelli-tributari/{modello_id}", response_model=ModelloTributarioResponse)
async def get_modello_tributario(modello_id: str, user: dict = Depends(get_current_user)):
    modello = await db.modelli_tributari.find_one({"id": modello_id}, {"_id": 0})
    if not modello:
        raise HTTPException(status_code=404, detail="Modello tributario non trovato")
    modello["video_thumbnail"] = get_youtube_thumbnail(modello.get("video_youtube"))
    modello.pop("conseguenze_mancata_presentazione", None)
    return ModelloTributarioResponse(**modello)

@api_router.post("/modelli-tributari", response_model=ModelloTributarioResponse)
async def create_modello_tributario(modello_data: ModelloTributarioCreate, user: dict = Depends(require_commercialista)):
    modello_id = str(uuid.uuid4())
    modello = {
        "id": modello_id,
        **modello_data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.modelli_tributari.insert_one(modello)
    modello["video_thumbnail"] = get_youtube_thumbnail(modello.get("video_youtube"))
    return ModelloTributarioResponse(**modello)

@api_router.put("/modelli-tributari/{modello_id}", response_model=ModelloTributarioResponse)
async def update_modello_tributario(modello_id: str, modello_data: ModelloTributarioCreate, user: dict = Depends(require_commercialista)):
    update_dict = modello_data.model_dump()
    result = await db.modelli_tributari.update_one({"id": modello_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Modello tributario non trovato")
    
    modello = await db.modelli_tributari.find_one({"id": modello_id}, {"_id": 0})
    modello["video_thumbnail"] = get_youtube_thumbnail(modello.get("video_youtube"))
    modello.pop("conseguenze_mancata_presentazione", None)
    return ModelloTributarioResponse(**modello)

@api_router.delete("/modelli-tributari/{modello_id}")
async def delete_modello_tributario(modello_id: str, user: dict = Depends(require_commercialista)):
    result = await db.modelli_tributari.delete_one({"id": modello_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Modello tributario non trovato")
    return {"message": "Modello tributario eliminato"}

# ==================== ACTIVITY LOG ====================

async def log_activity(action: str, description: str, user_id: str):
    """Registra un'attività nel log"""
    log_entry = {
        "id": str(uuid.uuid4()),
        "action": action,
        "description": description,
        "user_id": user_id,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.activity_logs.insert_one(log_entry)

@api_router.get("/activity-logs")
async def get_activity_logs(limit: int = 50, user: dict = Depends(require_commercialista)):
    logs = await db.activity_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return logs

# ==================== STATS ROUTES ====================

@api_router.get("/stats")
async def get_stats(user: dict = Depends(get_current_user)):
    if user["role"] == "commercialista":
        clients_count = await db.users.count_documents({"role": "cliente"})
        clients_active = await db.users.count_documents({"role": "cliente", "stato": "attivo"})
        documents_count = await db.documents.count_documents({})
        payslips_count = await db.payslips.count_documents({})
        notes_count = await db.notes.count_documents({})
        
        # Scadenze per stato
        deadlines_da_fare = await db.deadlines.count_documents({"status": "da_fare"})
        deadlines_in_lavorazione = await db.deadlines.count_documents({"status": "in_lavorazione"})
        deadlines_completate = await db.deadlines.count_documents({"status": "completata"})
        deadlines_scadute = await db.deadlines.count_documents({"status": "scaduta"})
        
        # Documenti da verificare (senza AI description)
        docs_da_verificare = await db.documents.count_documents({"ai_description": None})
        
        return {
            "clients_count": clients_count,
            "clients_active": clients_active,
            "documents_count": documents_count,
            "payslips_count": payslips_count,
            "notes_count": notes_count,
            "deadlines_da_fare": deadlines_da_fare,
            "deadlines_in_lavorazione": deadlines_in_lavorazione,
            "deadlines_completate": deadlines_completate,
            "deadlines_scadute": deadlines_scadute,
            "docs_da_verificare": docs_da_verificare
        }
    else:
        documents_count = await db.documents.count_documents({"client_id": user["id"]})
        payslips_count = await db.payslips.count_documents({"client_id": user["id"]})
        notes_count = await db.notes.count_documents({"client_id": user["id"], "is_internal": False})
        
        # Scadenze del cliente
        deadlines_query = {"$or": [{"applies_to_all": True}, {"client_ids": user["id"]}]}
        deadlines_da_fare = await db.deadlines.count_documents({**deadlines_query, "status": "da_fare"})
        deadlines_completate = await db.deadlines.count_documents({**deadlines_query, "status": "completata"})
        
        return {
            "documents_count": documents_count,
            "payslips_count": payslips_count,
            "notes_count": notes_count,
            "deadlines_da_fare": deadlines_da_fare,
            "deadlines_completate": deadlines_completate
        }

# ==================== CHATBOT AI ROUTES ====================

class ChatMessage(BaseModel):
    message: str
    conversation_id: Optional[str] = None

class ChatResponse(BaseModel):
    response: str
    conversation_id: str
    success: bool

@api_router.post("/chat", response_model=ChatResponse)
async def chat_endpoint(chat_msg: ChatMessage, user: dict = Depends(get_current_user)):
    """Endpoint per chattare con l'assistente AI"""
    
    # Recupera contesto del cliente
    client_docs = []
    client_deadlines_list = []
    
    if user["role"] == "cliente":
        # Recupera documenti del cliente
        docs = await db.documents.find({"client_id": user["id"]}, {"title": 1, "_id": 0}).to_list(20)
        client_docs = [d["title"] for d in docs]
        
        # Recupera scadenze
        deadlines = await db.deadlines.find({
            "$or": [{"applies_to_all": True}, {"client_ids": user["id"]}]
        }, {"title": 1, "due_date": 1, "_id": 0}).to_list(10)
        client_deadlines_list = [f"{d['title']} ({d['due_date']})" for d in deadlines]
    
    # Recupera modelli tributari
    modelli = await db.modelli_tributari.find({}, {"_id": 0}).to_list(20)
    
    # Recupera storico conversazione se esiste
    conversation_history = []
    conv_id = chat_msg.conversation_id or str(uuid.uuid4())
    
    if chat_msg.conversation_id:
        conv = await db.conversations.find_one({"id": chat_msg.conversation_id, "user_id": user["id"]})
        if conv:
            conversation_history = conv.get("messages", [])[-10:]
    
    # Chiama assistente AI
    result = await chat_with_assistant(
        user_message=chat_msg.message,
        client_name=user["full_name"],
        conversation_history=conversation_history,
        client_documents=client_docs,
        client_deadlines=client_deadlines_list,
        modelli_tributari=modelli
    )
    
    # Salva conversazione
    new_messages = conversation_history + [
        {"role": "user", "content": chat_msg.message},
        {"role": "assistant", "content": result.get("response", "")}
    ]
    
    await db.conversations.update_one(
        {"id": conv_id},
        {
            "$set": {
                "id": conv_id,
                "user_id": user["id"],
                "messages": new_messages[-20:],  # Mantieni ultimi 20 messaggi
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        },
        upsert=True
    )
    
    return ChatResponse(
        response=result.get("response", "Errore nel processare la richiesta"),
        conversation_id=conv_id,
        success=result.get("success", False)
    )

@api_router.delete("/chat/{conversation_id}")
async def clear_conversation(conversation_id: str, user: dict = Depends(get_current_user)):
    """Cancella una conversazione"""
    await db.conversations.delete_one({"id": conversation_id, "user_id": user["id"]})
    return {"message": "Conversazione cancellata"}

# ==================== NOTIFICATION ROUTES ====================

# NOTA: Endpoint /notifications/send-document RIMOSSO su richiesta utente
# Le notifiche automatiche per caricamento documenti non vengono più inviate

@api_router.post("/notifications/send-deadline-reminder")
async def send_deadline_reminder(
    client_id: str = Form(...),
    deadline_id: str = Form(...),
    user: dict = Depends(require_commercialista)
):
    """Invia promemoria scadenza al cliente"""
    client = await db.users.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    deadline = await db.deadlines.find_one({"id": deadline_id}, {"_id": 0})
    if not deadline:
        raise HTTPException(status_code=404, detail="Scadenza non trovata")
    
    result = await notify_deadline_reminder(
        client_email=client["email"],
        client_name=client["full_name"],
        deadline_title=deadline["title"],
        deadline_date=deadline["due_date"],
        deadline_description=deadline.get("description")
    )
    
    # Log notifica
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "type": "deadline_reminder",
        "client_id": client_id,
        "deadline_id": deadline_id,
        "client_email": client["email"],
        "subject": f"Promemoria: {deadline['title']}",
        "sent_by": user["id"],
        "result": result,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return result

@api_router.post("/notifications/send-note")
async def send_note_notification(
    client_id: str = Form(...),
    note_title: str = Form(...),
    note_content: str = Form(...),
    user: dict = Depends(require_commercialista)
):
    """Invia notifica per nuova comunicazione"""
    client = await db.users.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    result = await notify_new_note(
        client_email=client["email"],
        client_name=client["full_name"],
        note_title=note_title,
        note_content=note_content
    )
    
    # Log notifica
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "type": "new_note",
        "client_id": client_id,
        "client_email": client["email"],
        "subject": f"Comunicazione: {note_title}",
        "sent_by": user["id"],
        "result": result,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return result

@api_router.get("/notifications/history")
async def get_notifications_history(limit: int = 50, user: dict = Depends(require_commercialista)):
    """Recupera storico notifiche inviate"""
    notifications = await db.notifications.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return notifications

@api_router.get("/")
async def root():
    return {"message": "Fiscal Tax Canarie API"}

# ==================== CLIENT CREATION ROUTES ====================

@api_router.post("/clients/create")
async def create_client(client_data: ClientCreate, user: dict = Depends(require_commercialista)):
    """
    Crea un nuovo cliente con cartella immediatamente accessibile.
    Se viene fornita l'email e send_invite=True, invia anche l'invito per la registrazione.
    """
    # Verifica duplicati (email, NIE, CIF)
    match_conditions = []
    if client_data.email:
        match_conditions.append({"email": client_data.email})
        match_conditions.append({"email_notifica": client_data.email})
    if client_data.nie:
        match_conditions.append({"nie": client_data.nie})
    if client_data.cif:
        match_conditions.append({"cif": client_data.cif})
    
    if match_conditions:
        existing = await db.users.find_one({"$or": match_conditions, "role": "cliente"})
        if existing:
            raise HTTPException(
                status_code=400, 
                detail="Esiste già un cliente con questa email, NIE o CIF"
            )
    
    client_id = str(uuid.uuid4())
    invitation_token = None
    invitation_link = None
    
    # Se c'è email e send_invite è True, genera token di invito
    if client_data.email and client_data.send_invite:
        invitation_token = secrets.token_urlsafe(32)
        frontend_url = os.environ.get('FRONTEND_URL', 'https://app.fiscaltaxcanarie.com')
        invitation_link = f"{frontend_url}/complete-registration?token={invitation_token}"
    
    # Crea il cliente
    client_doc = {
        "id": client_id,
        "email": None,  # Sarà impostata quando il cliente completa la registrazione
        "email_notifica": client_data.email,  # Email per notifiche (opzionale)
        "password": None,
        "full_name": client_data.full_name,
        "phone": client_data.phone,
        "codice_fiscale": client_data.codice_fiscale,
        "nie": client_data.nie,
        "nif": client_data.nif,
        "cif": client_data.cif,
        "indirizzo": client_data.indirizzo,
        "citta": client_data.citta,
        "cap": client_data.cap,
        "provincia": client_data.provincia,
        "iban": client_data.iban,
        "regime_fiscale": client_data.regime_fiscale,
        "tipo_attivita": client_data.tipo_attivita,
        "tipo_cliente": client_data.tipo_cliente or "autonomo",
        "role": "cliente",
        "stato": "invitato" if client_data.email else "attivo",  # Se no email, è già attivo (gestito solo da admin)
        "invited_by": user["id"],
        "note_interne": client_data.note_interne or "",
        "lists": [],
        "bank_credentials": [],
        "additional_emails": [],
        "invitation_token": invitation_token,
        "invitation_sent_at": datetime.now(timezone.utc).isoformat() if invitation_token else None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(client_doc)
    
    # Se c'è email e invito, salva anche nella collection invitations e invia email
    email_sent = False
    if client_data.email and invitation_token:
        invite_doc = {
            "id": str(uuid.uuid4()),
            "client_id": client_id,
            "notification_email": client_data.email,
            "suggested_name": client_data.full_name,
            "tipo_cliente": client_data.tipo_cliente or "autonomo",
            "invitation_token": invitation_token,
            "invitation_sent_at": datetime.now(timezone.utc).isoformat(),
            "invited_by": user["id"],
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.invitations.insert_one(invite_doc)
        
        # Invia email di invito
        try:
            email_result = await send_invitation_email(
                client_email=client_data.email,
                client_name=client_data.full_name,
                invitation_link=invitation_link
            )
            email_sent = email_result.get("success", False)
            if not email_sent:
                logger.warning(f"Email invito non inviata: {email_result.get('error')}")
        except Exception as e:
            logger.error(f"Errore invio email invito: {e}")
    
    await log_activity(
        "creazione_cliente",
        f"Nuovo cliente creato: {client_data.full_name}" + (f" - Invito inviato a {client_data.email}" if email_sent else ""),
        user["id"]
    )
    
    return {
        "success": True,
        "message": f"Cliente '{client_data.full_name}' creato con successo" + (f". Invito inviato a {client_data.email}" if email_sent else ""),
        "client_id": client_id,
        "invitation_link": invitation_link,
        "email_sent": email_sent
    }

# ==================== CLIENT INVITATION ROUTES (legacy) ====================

@api_router.post("/clients/invite")
async def invite_client(invite_data: ClientInvite, user: dict = Depends(require_commercialista)):
    """
    Crea un invito per un nuovo cliente E crea immediatamente la sua cartella.
    L'admin può subito caricare documenti, compilare anagrafica, aggiungere note.
    Quando il cliente completa la registrazione, viene matchato per email/NIE/CIF.
    """
    # Verifica che non esista già un cliente con questa email/NIE/CIF
    match_query = {"$or": [{"email": invite_data.email}]}
    if invite_data.nie:
        match_query["$or"].append({"nie": invite_data.nie})
    if invite_data.cif:
        match_query["$or"].append({"cif": invite_data.cif})
    
    existing = await db.users.find_one(match_query)
    if existing:
        raise HTTPException(
            status_code=400, 
            detail="Esiste già un cliente con questa email, NIE o CIF"
        )
    
    # Genera token di invito
    invitation_token = secrets.token_urlsafe(32)
    client_id = str(uuid.uuid4())
    
    # Crea il cliente completo IMMEDIATAMENTE (stato: invitato)
    client_doc = {
        "id": client_id,
        "email": None,  # Sarà impostata quando il cliente completa la registrazione
        "email_notifica": invite_data.email,  # Email per notifiche
        "password": None,  # Sarà impostata quando il cliente completa la registrazione
        "full_name": invite_data.full_name or "",
        "phone": invite_data.phone,
        "codice_fiscale": invite_data.codice_fiscale,
        "nie": invite_data.nie,
        "nif": invite_data.nif,
        "cif": invite_data.cif,
        "indirizzo": invite_data.indirizzo,
        "citta": invite_data.citta,
        "cap": invite_data.cap,
        "provincia": invite_data.provincia,
        "iban": invite_data.iban,
        "regime_fiscale": invite_data.regime_fiscale,
        "tipo_attivita": invite_data.tipo_attivita,
        "tipo_cliente": invite_data.tipo_cliente or "autonomo",
        "role": "cliente",
        "stato": "invitato",  # Stato speciale: invitato ma non ancora registrato
        "invited_by": user["id"],
        "note_interne": invite_data.note_interne or "",
        "lists": [],
        "bank_credentials": [],
        "additional_emails": [],
        # Token per il completamento registrazione
        "invitation_token": invitation_token,
        "invitation_sent_at": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(client_doc)
    
    # Salva anche nella collection invitations per compatibilità
    invite_doc = {
        "id": str(uuid.uuid4()),
        "client_id": client_id,  # Riferimento al cliente creato
        "notification_email": invite_data.email,
        "suggested_name": invite_data.full_name or "",
        "tipo_cliente": invite_data.tipo_cliente or "autonomo",
        "invitation_token": invitation_token,
        "invitation_sent_at": datetime.now(timezone.utc).isoformat(),
        "invited_by": user["id"],
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.invitations.insert_one(invite_doc)
    
    # Genera link di invito
    frontend_url = os.environ.get('FRONTEND_URL', 'https://app.fiscaltaxcanarie.com')
    invitation_link = f"{frontend_url}/complete-registration?token={invitation_token}"
    
    # Invia email di invito
    try:
        email_result = await send_invitation_email(
            client_email=invite_data.email,
            client_name=invite_data.full_name or "",
            invitation_link=invitation_link
        )
        
        if not email_result.get("success"):
            logger.warning(f"Email invito non inviata: {email_result.get('error')}")
    except Exception as e:
        logger.error(f"Errore invio email invito: {e}")
    
    await log_activity(
        "invito_cliente",
        f"Invito e cartella creati per {invite_data.email} - Cliente {invite_data.full_name}",
        user["id"]
    )
    
    return {
        "success": True,
        "message": f"Cartella cliente creata e invito inviato a {invite_data.email}",
        "client_id": client_id,  # ID della cartella cliente creata
        "invite_id": invite_doc["id"],
        "invitation_link": invitation_link
    }

@api_router.post("/auth/complete-registration")
async def complete_registration(data: CompleteRegistration):
    """
    Completa la registrazione di un utente invitato.
    Per i CLIENTI: Trova la cartella esistente (creata all'invito) e la attiva con le credenziali.
    Il matching avviene per invitation_token, email_notifica, NIE o CIF.
    Per i CONSULENTI: Crea un nuovo account.
    """
    # Prima cerca il cliente già creato con questo token
    existing_client = await db.users.find_one(
        {"invitation_token": data.token, "role": "cliente"},
        {"_id": 0}
    )
    
    if existing_client:
        # CASO 1: Cliente con cartella già creata - aggiorna con credenziali
        
        # Verifica che il token non sia scaduto (7 giorni)
        invitation_sent = existing_client.get("invitation_sent_at")
        if invitation_sent:
            sent_date = datetime.fromisoformat(invitation_sent.replace('Z', '+00:00'))
            if datetime.now(timezone.utc) - sent_date > timedelta(days=7):
                raise HTTPException(status_code=400, detail="Invito scaduto. Richiedi un nuovo invito.")
        
        # Verifica che non sia già registrato
        if existing_client.get("stato") == "attivo" and existing_client.get("password"):
            raise HTTPException(status_code=400, detail="Registrazione già completata. Usa il Login.")
        
        # Verifica che l'email scelta non sia già usata da un altro utente
        email_check = await db.users.find_one({"email": data.email, "id": {"$ne": existing_client["id"]}})
        if email_check:
            raise HTTPException(status_code=400, detail="Questa email è già registrata. Scegli un'altra email.")
        
        # Aggiorna il cliente esistente con le credenziali
        update_data = {
            "email": data.email,  # Email scelta per il login
            "password": hash_password(data.password),
            "stato": "attivo",  # Da "invitato" a "attivo"
            "registration_completed_at": datetime.now(timezone.utc).isoformat(),
            "invitation_token": None  # Rimuovi il token usato
        }
        
        # Aggiorna il nome solo se fornito
        if data.full_name:
            update_data["full_name"] = data.full_name
        
        await db.users.update_one(
            {"id": existing_client["id"]},
            {"$set": update_data}
        )
        
        # Aggiorna anche l'invito nella collection invitations
        await db.invitations.update_one(
            {"client_id": existing_client["id"]},
            {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Invia email di benvenuto
        try:
            await send_welcome_email(data.email, data.full_name or existing_client.get("full_name", ""))
        except Exception as e:
            logger.error(f"Errore invio email benvenuto: {e}")
        
        await log_activity(
            "registrazione_completata",
            f"Cliente {data.email} ha completato la registrazione (cartella esistente)",
            existing_client["id"]
        )
        
        # Genera token di accesso
        token = create_token(existing_client["id"], data.email, "cliente")
        
        return {
            "success": True,
            "message": "Registrazione completata! Puoi accedere alla tua area personale.",
            "access_token": token,
            "token_type": "bearer",
            "user": {
                "id": existing_client["id"],
                "email": data.email,
                "full_name": data.full_name or existing_client.get("full_name", ""),
                "role": "cliente"
            }
        }
    
    # CASO 2: Cerca nella collection invitations (per consulenti o inviti vecchio stile)
    invitation = await db.invitations.find_one({"invitation_token": data.token}, {"_id": 0})
    
    if not invitation:
        raise HTTPException(
            status_code=400, 
            detail="Token non valido o già utilizzato. Se hai già completato la registrazione, usa il Login."
        )
    
    if invitation.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Invito già utilizzato. Usa il Login per accedere.")
    
    # Verifica scadenza
    invitation_sent = invitation.get("invitation_sent_at")
    if invitation_sent:
        sent_date = datetime.fromisoformat(invitation_sent.replace('Z', '+00:00'))
        if datetime.now(timezone.utc) - sent_date > timedelta(days=7):
            raise HTTPException(status_code=400, detail="Invito scaduto. Richiedi un nuovo invito.")
    
    # Verifica email non già usata
    existing_user = await db.users.find_one({"email": data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Questa email è già registrata. Scegli un'altra email o usa il Login.")
    
    invited_role = invitation.get("role", "cliente")
    is_consulente = invited_role == "consulente_lavoro"
    
    user_id = str(uuid.uuid4())
    
    if is_consulente:
        # Crea documento consulente del lavoro
        user_doc = {
            "id": user_id,
            "email": data.email,
            "password": hash_password(data.password),
            "full_name": data.full_name or invitation.get("suggested_name", ""),
            "role": "consulente_lavoro",
            "stato": "attivo",
            "assigned_clients": [],
            "created_by": invitation.get("invited_by"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "registration_completed_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user_doc)
    else:
        # Per clienti vecchio stile (senza cartella pre-creata)
        user_doc = {
            "id": user_id,
            "email": data.email,
            "email_notifica": invitation.get("notification_email"),
            "additional_emails": [],
            "password": hash_password(data.password),
            "full_name": data.full_name or invitation.get("suggested_name", ""),
            "phone": None,
            "codice_fiscale": None,
            "nie": None,
            "nif": None,
            "cif": None,
            "indirizzo": None,
            "citta": None,
            "cap": None,
            "provincia": None,
            "iban": None,
            "regime_fiscale": None,
            "tipo_attivita": None,
            "tipo_cliente": invitation.get("tipo_cliente", "autonomo"),
            "role": "cliente",
            "stato": "attivo",
            "invited_by": invitation.get("invited_by"),
            "note_interne": "",
            "lists": [],
            "bank_credentials": [],
            "created_at": datetime.now(timezone.utc).isoformat(),
            "registration_completed_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user_doc)
    
    # Aggiorna l'invito come completato
    await db.invitations.update_one(
        {"id": invitation["id"]},
        {"$set": {
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "user_id": user_id
        }}
    )
    
    # Invia email di benvenuto
    try:
        await send_welcome_email(data.email, data.full_name or invitation.get("suggested_name", ""))
    except Exception as e:
        logger.error(f"Errore invio email benvenuto: {e}")
    
    role_label = "Consulente del lavoro" if is_consulente else "Cliente"
    await log_activity(
        "registrazione_completata",
        f"{role_label} {data.email} ha completato la registrazione",
        user_id
    )
    
    # Genera token di accesso
    final_role = "consulente_lavoro" if is_consulente else "cliente"
    token = create_token(user_id, data.email, final_role)
    
    return {
        "success": True,
        "message": "Registrazione completata con successo!",
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user_id,
            "email": data.email,
            "full_name": data.full_name or invitation.get("suggested_name", ""),
            "role": final_role
        }
    }

@api_router.post("/clients/resend-invite/{invite_id}")
async def resend_invitation(invite_id: str, user: dict = Depends(require_commercialista)):
    """Reinvia l'invito"""
    # Cerca l'invito
    invitation = await db.invitations.find_one({"id": invite_id}, {"_id": 0})
    
    if not invitation:
        raise HTTPException(status_code=404, detail="Invito non trovato")
    
    if invitation.get("status") != "pending":
        raise HTTPException(status_code=400, detail="L'invito è già stato utilizzato")
    
    # Genera nuovo token
    new_token = secrets.token_urlsafe(32)
    
    await db.invitations.update_one(
        {"id": invite_id},
        {"$set": {
            "invitation_token": new_token,
            "invitation_sent_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Genera link e invia email - usa dominio produzione
    frontend_url = os.environ.get('FRONTEND_URL', 'https://app.fiscaltaxcanarie.com')
    invitation_link = f"{frontend_url}/complete-registration?token={new_token}"
    
    try:
        await send_invitation_email(
            client_email=invitation["notification_email"],
            client_name=invitation.get("suggested_name", ""),
            invitation_link=invitation_link
        )
    except Exception as e:
        logger.error(f"Errore reinvio email invito: {e}")
        raise HTTPException(status_code=500, detail="Errore nell'invio dell'email")
    
    return {"success": True, "message": f"Invito reinviato a {invitation['notification_email']}", "invitation_link": invitation_link}

# ==================== INVITATIONS LIST ====================

@api_router.get("/invitations")
async def get_invitations(user: dict = Depends(require_commercialista)):
    """
    Lista gli inviti pendenti.
    Include client_id per permettere al frontend di navigare alla cartella del cliente.
    """
    invitations = await db.invitations.find(
        {"invited_by": user["id"], "status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Per ogni invito con client_id, verifica che il cliente esista ancora
    # e aggiungi informazioni utili
    for inv in invitations:
        if inv.get("client_id"):
            client = await db.users.find_one(
                {"id": inv["client_id"], "role": "cliente"},
                {"_id": 0, "full_name": 1, "stato": 1}
            )
            if client:
                inv["has_client_folder"] = True
                inv["client_name"] = client.get("full_name", "")
                inv["client_stato"] = client.get("stato", "")
    
    return invitations

@api_router.delete("/invitations/{invitation_id}")
async def delete_invitation(invitation_id: str, user: dict = Depends(require_commercialista)):
    """
    Elimina un invito in attesa di registrazione.
    Se l'invito ha un client_id associato (cliente già creato con stato 'invitato'),
    elimina anche il profilo cliente.
    """
    # Trova l'invito
    invitation = await db.invitations.find_one({"id": invitation_id, "invited_by": user["id"]})
    if not invitation:
        raise HTTPException(status_code=404, detail="Invito non trovato")
    
    # Se l'invito ha un client_id, elimina anche il cliente associato
    if invitation.get("client_id"):
        client = await db.users.find_one({"id": invitation["client_id"], "stato": "invitato"})
        if client:
            # Elimina documenti del cliente
            await db.documents.delete_many({"client_id": invitation["client_id"]})
            # Elimina il cliente
            await db.users.delete_one({"id": invitation["client_id"]})
            logger.info(f"Eliminato cliente invitato {invitation['client_id']} insieme all'invito")
    
    # Elimina l'invito
    result = await db.invitations.delete_one({"id": invitation_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Errore durante l'eliminazione")
    
    await log_activity(
        "eliminazione_invito",
        f"Invito eliminato: {invitation.get('notification_email', invitation.get('suggested_name', 'N/A'))}",
        user["id"]
    )
    
    return {"message": "Invito eliminato con successo"}

# ==================== FEES (ONORARI) ROUTES ====================

@api_router.get("/fees/all")
async def get_all_fees(
    search: Optional[str] = None,
    client_type: Optional[str] = None,
    status: Optional[str] = None,
    year: Optional[int] = None,
    user: dict = Depends(require_commercialista)
):
    """
    Recupera tutti gli onorari con filtri opzionali.
    - search: ricerca per descrizione o nome cliente
    - client_type: filtra per tipo cliente (societa, autonomo, vivienda_vacacional, persona_fisica)
    - status: filtra per stato (pending, paid, overdue)
    - year: filtra per anno scadenza
    """
    # Prima recupera tutti i clienti per join
    clients_cursor = db.users.find({"role": "cliente"}, {"_id": 0, "id": 1, "full_name": 1, "email": 1, "tipo_cliente": 1})
    clients_list = await clients_cursor.to_list(10000)
    clients_map = {c["id"]: c for c in clients_list}
    
    # Query per gli onorari
    query = {}
    
    if status:
        query["status"] = status
    
    if year:
        # Filtra per anno della data di scadenza
        query["due_date"] = {"$regex": f"^{year}"}
    
    # Recupera onorari
    fees = await db.fees.find(query, {"_id": 0}).sort("due_date", -1).to_list(10000)
    
    # Arricchisci con info cliente e applica filtri aggiuntivi
    result = []
    for fee in fees:
        client = clients_map.get(fee.get("client_id"), {})
        
        # Filtro per tipo cliente
        if client_type and client.get("tipo_cliente") != client_type:
            continue
        
        # Filtro per ricerca
        if search:
            search_lower = search.lower()
            match = False
            if search_lower in fee.get("description", "").lower():
                match = True
            if search_lower in client.get("full_name", "").lower():
                match = True
            if search_lower in client.get("email", "").lower():
                match = True
            if not match:
                continue
        
        # Aggiungi info cliente al fee
        fee["client_name"] = client.get("full_name", "N/A")
        fee["client_email"] = client.get("email", "")
        fee["client_type"] = client.get("tipo_cliente", "")
        result.append(fee)
    
    return result

@api_router.get("/fees/summary")
async def get_global_fees_summary(user: dict = Depends(require_commercialista)):
    """Ottiene un riepilogo globale degli onorari"""
    fees = await db.fees.find({}, {"_id": 0}).to_list(10000)
    
    total_pending = sum(f["amount"] for f in fees if f.get("status") == "pending")
    total_paid = sum(f["amount"] for f in fees if f.get("status") == "paid")
    total_overdue = sum(f["amount"] for f in fees if f.get("status") == "overdue")
    
    # Raggruppa per cliente
    by_client = {}
    for fee in fees:
        cid = fee.get("client_id")
        if cid not in by_client:
            by_client[cid] = {"pending": 0, "paid": 0, "count": 0}
        by_client[cid]["count"] += 1
        if fee.get("status") == "pending":
            by_client[cid]["pending"] += fee["amount"]
        elif fee.get("status") == "paid":
            by_client[cid]["paid"] += fee["amount"]
    
    return {
        "total_pending": total_pending,
        "total_paid": total_paid,
        "total_overdue": total_overdue,
        "total_count": len(fees),
        "clients_count": len(by_client)
    }

@api_router.get("/fees/by-client")
async def get_fees_grouped_by_client(user: dict = Depends(require_commercialista)):
    """Recupera tutti i clienti con i loro onorari raggruppati"""
    # Recupera tutti i clienti
    clients = await db.users.find(
        {"role": "cliente"}, 
        {"_id": 0, "id": 1, "full_name": 1, "email": 1, "tipo_cliente": 1}
    ).to_list(10000)
    
    # Recupera tutti gli onorari
    fees = await db.fees.find({}, {"_id": 0}).to_list(10000)
    
    # Raggruppa per cliente
    fees_by_client = {}
    for fee in fees:
        cid = fee.get("client_id")
        if cid not in fees_by_client:
            fees_by_client[cid] = []
        fees_by_client[cid].append(fee)
    
    # Costruisci risultato
    result = []
    for client in clients:
        client_fees = fees_by_client.get(client["id"], [])
        total_pending = sum(f["amount"] for f in client_fees if f.get("status") == "pending")
        total_paid = sum(f["amount"] for f in client_fees if f.get("status") == "paid")
        
        # Calcola totale Iguala mensile
        iguala_monthly = sum(
            f["amount"] for f in client_fees 
            if f.get("fee_type", "").startswith("iguala_") or f.get("is_recurring")
        )
        
        result.append({
            "id": client["id"],
            "full_name": client.get("full_name", "N/A"),
            "email": client.get("email", ""),
            "tipo_cliente": client.get("tipo_cliente", ""),
            "fees": client_fees,
            "fees_count": len(client_fees),
            "total_pending": total_pending,
            "total_paid": total_paid,
            "iguala_monthly": iguala_monthly
        })
    
    # Ordina per nome
    result.sort(key=lambda x: x["full_name"].lower())
    return result

@api_router.get("/fees/export-excel")
async def export_fees_excel(
    category: Optional[str] = None,
    fee_type: Optional[str] = None,
    user: dict = Depends(require_commercialista)
):
    """Esporta onorari in formato Excel con filtri"""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installato")
    
    # Recupera clienti con filtro categoria
    client_query = {"role": "cliente"}
    if category and category != "all":
        client_query["tipo_cliente"] = category
    
    clients = await db.users.find(
        client_query, 
        {"_id": 0, "id": 1, "full_name": 1, "email": 1, "tipo_cliente": 1}
    ).to_list(10000)
    
    client_ids = [c["id"] for c in clients]
    client_map = {c["id"]: c for c in clients}
    
    # Recupera onorari
    fee_query = {"client_id": {"$in": client_ids}}
    if fee_type and fee_type != "all":
        if fee_type == "iguala":
            fee_query["$or"] = [
                {"fee_type": {"$regex": "^iguala_"}},
                {"is_recurring": True}
            ]
        else:
            fee_query["fee_type"] = fee_type
    
    fees = await db.fees.find(fee_query, {"_id": 0}).to_list(10000)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Onorari"
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F855A", end_color="2F855A", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header
    headers = ["Cliente", "Tipo Cliente", "Email", "Descrizione Onorario", "Tipo Onorario", "Importo (€)", "Stato", "Scadenza", "Mese Rif."]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Dati
    row = 2
    for fee in fees:
        client = client_map.get(fee.get("client_id"), {})
        fee_type_labels = {
            "standard": "Standard",
            "consulenza": "Consulenza",
            "pratica": "Pratica/Procedura",
            "dichiarazione": "Dichiarazione Fiscale",
            "iguala_buste_paga": "Iguala - Buste Paga",
            "iguala_contabilita": "Iguala - Contabilità",
            "iguala_domicilio": "Iguala - Domicilio"
        }
        status_labels = {"pending": "In Attesa", "paid": "Pagato", "overdue": "Scaduto"}
        tipo_cliente_labels = {
            "societa": "Società",
            "autonomo": "Autonomo",
            "vivienda_vacacional": "Vivienda Vacacional",
            "persona_fisica": "Persona Fisica"
        }
        
        ws.cell(row=row, column=1, value=client.get("full_name", "N/A")).border = thin_border
        ws.cell(row=row, column=2, value=tipo_cliente_labels.get(client.get("tipo_cliente", ""), client.get("tipo_cliente", ""))).border = thin_border
        ws.cell(row=row, column=3, value=client.get("email", "")).border = thin_border
        ws.cell(row=row, column=4, value=fee.get("description", "")).border = thin_border
        ws.cell(row=row, column=5, value=fee_type_labels.get(fee.get("fee_type", "standard"), fee.get("fee_type", ""))).border = thin_border
        ws.cell(row=row, column=6, value=fee.get("amount", 0)).border = thin_border
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=status_labels.get(fee.get("status", ""), fee.get("status", ""))).border = thin_border
        ws.cell(row=row, column=8, value=fee.get("due_date", "-") or "-").border = thin_border
        ws.cell(row=row, column=9, value=fee.get("recurring_month", "-") or "-").border = thin_border
        row += 1
    
    # Aggiungi riga totale
    if fees:
        total_row = row + 1
        ws.cell(row=total_row, column=5, value="TOTALE:").font = Font(bold=True)
        total_amount = sum(f.get("amount", 0) for f in fees)
        ws.cell(row=total_row, column=6, value=total_amount).font = Font(bold=True)
        ws.cell(row=total_row, column=6).number_format = '#,##0.00'
    
    # Larghezza colonne
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Salva in buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Genera nome file
    filename = f"onorari"
    if category and category != "all":
        filename += f"_{category}"
    if fee_type and fee_type != "all":
        filename += f"_{fee_type}"
    filename += f"_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/clients/{client_id}/fees", response_model=List[FeeResponse])
async def get_client_fees(client_id: str, user: dict = Depends(require_commercialista)):
    """Recupera tutti gli onorari di un cliente"""
    fees = await db.fees.find({"client_id": client_id}, {"_id": 0}).sort("due_date", -1).to_list(1000)
    return [FeeResponse(**fee) for fee in fees]

@api_router.post("/clients/{client_id}/fees", response_model=FeeResponse)
async def create_fee(client_id: str, fee_data: FeeCreate, user: dict = Depends(require_commercialista)):
    """Crea un nuovo onorario per un cliente"""
    # Verifica che il cliente esista
    client = await db.users.find_one({"id": client_id, "role": "cliente"})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    # Determina se is_recurring in base al fee_type
    is_recurring = fee_data.is_recurring or fee_data.fee_type.startswith("iguala_")
    
    fee_id = str(uuid.uuid4())
    fee = {
        "id": fee_id,
        "client_id": client_id,
        "description": fee_data.description,
        "amount": fee_data.amount,
        "due_date": fee_data.due_date,
        "status": fee_data.status,
        "paid_date": None,
        "notes": fee_data.notes,
        "fee_type": fee_data.fee_type,
        "is_recurring": is_recurring,
        "recurring_month": fee_data.recurring_month,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.fees.insert_one(fee)
    
    await log_activity(
        "creazione_onorario",
        f"Onorario €{fee_data.amount} ({fee_data.fee_type}) creato per cliente {client_id}",
        user["id"]
    )
    
    return FeeResponse(**fee)

@api_router.put("/clients/{client_id}/fees/{fee_id}", response_model=FeeResponse)
async def update_fee(client_id: str, fee_id: str, fee_data: FeeUpdate, user: dict = Depends(require_commercialista)):
    """Aggiorna un onorario"""
    update_dict = {k: v for k, v in fee_data.model_dump().items() if v is not None}
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    # Se stato diventa "paid", imposta paid_date
    if update_dict.get("status") == "paid" and not update_dict.get("paid_date"):
        update_dict["paid_date"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.fees.update_one(
        {"id": fee_id, "client_id": client_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Onorario non trovato")
    
    fee = await db.fees.find_one({"id": fee_id}, {"_id": 0})
    return FeeResponse(**fee)

@api_router.delete("/clients/{client_id}/fees/{fee_id}")
async def delete_fee(client_id: str, fee_id: str, user: dict = Depends(require_commercialista)):
    """Elimina un onorario"""
    result = await db.fees.delete_one({"id": fee_id, "client_id": client_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Onorario non trovato")
    
    await log_activity("eliminazione_onorario", f"Onorario {fee_id} eliminato", user["id"])
    
    return {"message": "Onorario eliminato"}

@api_router.get("/clients/{client_id}/fees/summary")
async def get_fees_summary(client_id: str, user: dict = Depends(require_commercialista)):
    """Ottiene un riepilogo degli onorari di un cliente"""
    fees = await db.fees.find({"client_id": client_id}, {"_id": 0}).to_list(1000)
    
    total_pending = sum(f["amount"] for f in fees if f["status"] == "pending")
    total_paid = sum(f["amount"] for f in fees if f["status"] == "paid")
    total = total_pending + total_paid
    
    return {
        "total": total,
        "total_paid": total_paid,
        "total_pending": total_pending,
        "count_total": len(fees),
        "count_paid": len([f for f in fees if f["status"] == "paid"]),
        "count_pending": len([f for f in fees if f["status"] == "pending"])
    }

# ==================== DIGITAL SIGNATURE ROUTES ====================

@api_router.post("/certificates/upload")
async def upload_certificate(
    certificate_name: str = Form(...),
    certificate_password: str = Form(...),
    file: UploadFile = File(...),
    user: dict = Depends(require_commercialista)
):
    """Carica un certificato .p12 per la firma digitale"""
    if not file.filename.endswith('.p12'):
        raise HTTPException(status_code=400, detail="Il file deve essere in formato .p12")
    
    # Leggi il file
    cert_data = await file.read()
    
    # Verifica che il certificato sia valido
    try:
        from cryptography.hazmat.primitives.serialization import pkcs12
        from cryptography.hazmat.backends import default_backend
        
        private_key, certificate, _ = pkcs12.load_key_and_certificates(
            cert_data,
            certificate_password.encode('utf-8'),
            default_backend()
        )
        
        if not private_key or not certificate:
            raise HTTPException(status_code=400, detail="Certificato non valido")
            
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Errore verifica certificato: {str(e)}")
    
    # Salva il certificato
    result = await save_certificate(cert_data, certificate_name, user["id"])
    
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error"))
    
    # Salva info certificato nel DB (senza password!)
    cert_id = str(uuid.uuid4())
    await db.certificates.insert_one({
        "id": cert_id,
        "name": certificate_name,
        "user_id": user["id"],
        "filename": file.filename,
        "subject": str(certificate.subject) if certificate else None,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    await log_activity("caricamento_certificato", f"Certificato {certificate_name} caricato", user["id"])
    
    return {
        "success": True,
        "message": "Certificato caricato con successo",
        "certificate_id": cert_id,
        "certificate_name": certificate_name
    }

@api_router.get("/certificates")
async def get_certificates(user: dict = Depends(require_commercialista)):
    """Lista i certificati del commercialista"""
    certs = await db.certificates.find({"user_id": user["id"]}, {"_id": 0}).to_list(100)
    return certs

@api_router.delete("/certificates/{cert_name}")
async def remove_certificate(cert_name: str, user: dict = Depends(require_commercialista)):
    """Elimina un certificato"""
    result = await delete_certificate(user["id"], cert_name)
    
    if not result.get("success"):
        raise HTTPException(status_code=404, detail=result.get("error"))
    
    # Rimuovi anche dal DB
    await db.certificates.delete_one({"name": cert_name, "user_id": user["id"]})
    
    return {"message": "Certificato eliminato"}

# ==================== CLIENT CERTIFICATES (Per cliente) ====================

@api_router.post("/clients/{client_id}/certificates")
async def upload_client_certificate(
    client_id: str,
    certificate_name: str = Form(...),
    file: UploadFile = File(...),
    notes: str = Form(None),
    user: dict = Depends(require_commercialista)
):
    """Carica un certificato digitale associato a un cliente specifico (solo admin)"""
    # Verifica che il cliente esista
    client = await db.users.find_one({"id": client_id, "role": "cliente"}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    if not file.filename.endswith('.p12'):
        raise HTTPException(status_code=400, detail="Il file deve essere in formato .p12")
    
    # Leggi il file
    cert_data = await file.read()
    cert_base64 = base64.b64encode(cert_data).decode('utf-8')
    
    # Salva info certificato nel DB
    cert_id = str(uuid.uuid4())
    await db.client_certificates.insert_one({
        "id": cert_id,
        "client_id": client_id,
        "name": certificate_name,
        "filename": file.filename,
        "file_data": cert_base64,
        "notes": notes,
        "uploaded_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    await log_activity("caricamento_certificato_cliente", f"Certificato {certificate_name} caricato per cliente {client.get('full_name')}", user["id"])
    
    return {
        "success": True,
        "message": "Certificato caricato con successo",
        "certificate_id": cert_id,
        "certificate_name": certificate_name
    }

@api_router.get("/clients/{client_id}/certificates")
async def get_client_certificates(client_id: str, user: dict = Depends(get_current_user)):
    """
    Lista i certificati digitali associati a un cliente.
    Accessibile a: admin, cliente (solo i propri), consulente (solo clienti assegnati)
    NON richiede password per la visualizzazione.
    """
    # Verifica permessi
    if user["role"] == "cliente":
        if user["id"] != client_id:
            raise HTTPException(status_code=403, detail="Puoi visualizzare solo i tuoi certificati")
    elif user["role"] == "consulente_lavoro":
        assigned_clients = user.get("assigned_clients", [])
        if client_id not in assigned_clients:
            raise HTTPException(status_code=403, detail="Non hai accesso ai certificati di questo cliente")
    elif user["role"] != "commercialista":
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Recupera certificati del cliente (SENZA file_data per sicurezza nella lista)
    certs = await db.client_certificates.find(
        {"client_id": client_id},
        {"_id": 0, "file_data": 0}  # Esclude file_data dalla lista
    ).to_list(100)
    
    return certs

@api_router.get("/clients/{client_id}/certificates/{cert_id}/download")
async def download_client_certificate(client_id: str, cert_id: str, user: dict = Depends(get_current_user)):
    """
    Scarica un certificato digitale del cliente.
    Accessibile a: admin, cliente (solo i propri), consulente (solo clienti assegnati)
    NON richiede password per il download.
    """
    # Verifica permessi
    if user["role"] == "cliente":
        if user["id"] != client_id:
            raise HTTPException(status_code=403, detail="Puoi scaricare solo i tuoi certificati")
    elif user["role"] == "consulente_lavoro":
        assigned_clients = user.get("assigned_clients", [])
        if client_id not in assigned_clients:
            raise HTTPException(status_code=403, detail="Non hai accesso ai certificati di questo cliente")
    elif user["role"] != "commercialista":
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Recupera certificato con file_data
    cert = await db.client_certificates.find_one(
        {"id": cert_id, "client_id": client_id},
        {"_id": 0}
    )
    
    if not cert:
        raise HTTPException(status_code=404, detail="Certificato non trovato")
    
    return {
        "id": cert["id"],
        "name": cert["name"],
        "filename": cert["filename"],
        "file_data": cert["file_data"],
        "notes": cert.get("notes"),
        "created_at": cert["created_at"]
    }

@api_router.delete("/clients/{client_id}/certificates/{cert_id}")
async def delete_client_certificate(client_id: str, cert_id: str, user: dict = Depends(require_commercialista)):
    """Elimina un certificato digitale del cliente (solo admin)"""
    result = await db.client_certificates.delete_one({"id": cert_id, "client_id": client_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Certificato non trovato")
    
    await log_activity("eliminazione_certificato_cliente", f"Certificato {cert_id} eliminato", user["id"])
    
    return {"message": "Certificato eliminato"}

@api_router.post("/documents/{doc_id}/sign")
async def sign_document(
    doc_id: str,
    certificate_name: str = Form(...),
    certificate_password: str = Form(...),
    reason: str = Form("Documento firmato digitalmente"),
    location: str = Form("Isole Canarie, Spagna"),
    user: dict = Depends(require_commercialista)
):
    """Firma digitalmente un documento PDF"""
    # Recupera documento
    document = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    if not document.get("file_data"):
        raise HTTPException(status_code=400, detail="Documento senza file allegato")
    
    # Recupera certificato
    from pathlib import Path
    cert_path = Path(__file__).parent / "certificates" / user["id"] / f"{certificate_name}.p12"
    
    if not cert_path.exists():
        raise HTTPException(status_code=404, detail="Certificato non trovato")
    
    with open(cert_path, 'rb') as f:
        p12_data = f.read()
    
    # Decodifica PDF
    pdf_data = base64.b64decode(document["file_data"])
    
    # Firma il documento
    result = await sign_pdf_with_p12(
        pdf_data=pdf_data,
        p12_data=p12_data,
        p12_password=certificate_password,
        signer_name=user.get("full_name", "Fiscal Tax Canarie"),
        reason=reason,
        location=location
    )
    
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error"))
    
    # Aggiorna documento con versione firmata
    new_filename = document["file_name"].replace(".pdf", "_FIRMATO.pdf")
    
    await db.documents.update_one(
        {"id": doc_id},
        {"$set": {
            "file_data": result["signed_pdf_data"],
            "file_name": new_filename,
            "signed": True,
            "signed_by": user["id"],
            "signed_at": result["signed_at"],
            "signature_reason": reason,
            "signature_location": location
        }}
    )
    
    await log_activity(
        "firma_documento",
        f"Documento {doc_id} firmato digitalmente",
        user["id"]
    )
    
    return {
        "success": True,
        "message": "Documento firmato con successo",
        "signed_filename": new_filename,
        "signed_at": result["signed_at"]
    }

@api_router.post("/documents/{doc_id}/verify-signature")
async def verify_document_signature(doc_id: str, user: dict = Depends(get_current_user)):
    """Verifica la firma digitale di un documento"""
    document = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    if not document.get("file_data"):
        raise HTTPException(status_code=400, detail="Documento senza file allegato")
    
    pdf_data = base64.b64decode(document["file_data"])
    
    result = await verify_pdf_signature(pdf_data)
    
    return result

# ==================== BACKUP & EXPORT ROUTES ====================

@api_router.get("/backup/client/{client_id}")
async def download_client_backup(client_id: str, user: dict = Depends(require_commercialista)):
    """
    Scarica backup completo di un cliente (ZIP con tutti i documenti)
    """
    result = await create_client_backup(db, client_id)
    
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error", "Errore creazione backup"))
    
    # Log backup
    await log_activity("backup_cliente", f"Backup cliente {client_id} scaricato", user["id"])
    
    return Response(
        content=result["backup_data"],
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{result["filename"]}"',
            "X-Backup-Size": str(result["size_bytes"]),
            "X-Documents-Count": str(result["statistics"]["documents"]),
            "X-Payslips-Count": str(result["statistics"]["payslips"])
        }
    )

@api_router.get("/backup/full")
async def download_full_backup(user: dict = Depends(require_commercialista)):
    """
    Scarica backup completo di TUTTI i dati dello studio (ZIP)
    Include tutti i clienti, documenti, configurazioni, etc.
    """
    result = await create_full_backup(db, user["id"])
    
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error", "Errore creazione backup"))
    
    # Salva log backup nel DB
    await db.backups.insert_one({
        "id": str(uuid.uuid4()),
        "filename": result["filename"],
        "type": "full",
        "size_bytes": result["size_bytes"],
        "statistics": result["statistics"],
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    await log_activity("backup_completo", f"Backup completo scaricato ({result['size_mb']} MB)", user["id"])
    
    return Response(
        content=result["backup_data"],
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{result["filename"]}"',
            "X-Backup-Size": str(result["size_bytes"]),
            "X-Backup-Size-MB": str(result["size_mb"]),
            "X-Clients-Count": str(result["statistics"]["clients"]),
            "X-Documents-Count": str(result["statistics"]["documents"])
        }
    )

@api_router.get("/backup/export-json")
async def export_json(user: dict = Depends(require_commercialista)):
    """
    Esporta tutto il database in formato JSON (senza file binari).
    Utile per analisi dati o migrazione.
    """
    result = await export_database_json(db, user["id"])
    
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error", "Errore export"))
    
    await log_activity("export_json", "Database esportato in JSON", user["id"])
    
    return Response(
        content=result["data"],
        media_type="application/json",
        headers={
            "Content-Disposition": f'attachment; filename="{result["filename"]}"'
        }
    )

@api_router.get("/backup/history")
async def get_backup_history(user: dict = Depends(require_commercialista)):
    """Recupera storico dei backup effettuati"""
    backups = await db.backups.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return backups

@api_router.get("/storage/status")
async def get_storage_status(user: dict = Depends(require_commercialista)):
    """Verifica stato dello storage cloud Backblaze B2"""
    from storage_service import get_storage_stats
    
    storage_enabled = is_storage_enabled()
    
    # Se Backblaze è attivo, ottieni statistiche da B2
    if storage_enabled:
        b2_stats = await get_storage_stats()
        
        # Conta anche file locali in MongoDB non ancora migrati
        local_docs = await db.documents.count_documents({"storage_path": {"$exists": False}, "file_data": {"$exists": True}})
        local_payslips = await db.payslips.count_documents({"storage_path": {"$exists": False}, "file_data": {"$exists": True}})
        
        return {
            "cloud_storage_enabled": True,
            "storage_provider": "Backblaze B2",
            "bucket_name": b2_stats.get("bucket_name", "fiscaltaxcanarie"),
            "statistics": {
                "cloud_documents": b2_stats.get("statistics", {}).get("documents_count", 0),
                "cloud_payslips": b2_stats.get("statistics", {}).get("payslips_count", 0),
                "local_documents_pending": local_docs,
                "local_payslips_pending": local_payslips,
                "total_cloud_files": b2_stats.get("statistics", {}).get("total_files", 0),
                "total_size_mb": b2_stats.get("statistics", {}).get("total_size_mb", 0),
                "total_size_gb": b2_stats.get("statistics", {}).get("total_size_gb", 0)
            },
            "limits": {
                "max_file_size_mb": 5000,  # B2 supporta fino a 5GB
                "storage_limit": "Illimitato (pay-per-use)"
            },
            "migration_pending": local_docs + local_payslips
        }
    
    # Fallback: solo MongoDB
    docs_count = await db.documents.count_documents({})
    payslips_count = await db.payslips.count_documents({})
    
    pipeline = [
        {"$project": {"size": {"$strLenBytes": {"$ifNull": ["$file_data", ""]}}}},
        {"$group": {"_id": None, "total": {"$sum": "$size"}}}
    ]
    
    docs_size = await db.documents.aggregate(pipeline).to_list(1)
    payslips_size = await db.payslips.aggregate(pipeline).to_list(1)
    
    total_docs_bytes = docs_size[0]["total"] if docs_size else 0
    total_payslips_bytes = payslips_size[0]["total"] if payslips_size else 0
    total_bytes = total_docs_bytes + total_payslips_bytes
    
    return {
        "cloud_storage_enabled": False,
        "storage_provider": "MongoDB (locale)",
        "statistics": {
            "documents_count": docs_count,
            "payslips_count": payslips_count,
            "total_files": docs_count + payslips_count,
            "estimated_size_bytes": total_bytes,
            "estimated_size_mb": round(total_bytes / (1024 * 1024), 2),
            "estimated_size_gb": round(total_bytes / (1024 * 1024 * 1024), 3)
        },
        "limits": {
            "max_file_size_mb": 16,
            "recommended_action": "Configura B2_KEY_ID e B2_APPLICATION_KEY per abilitare Backblaze B2"
        }
    }

@api_router.post("/storage/migrate-to-cloud")
async def migrate_to_cloud(
    background_tasks: BackgroundTasks,
    user: dict = Depends(require_commercialista)
):
    """
    Avvia migrazione dei file da MongoDB a cloud storage.
    Operazione in background per non bloccare l'API.
    """
    if not is_storage_enabled():
        raise HTTPException(
            status_code=400, 
            detail="Storage cloud non configurato. Imposta EMERGENT_LLM_KEY nel file .env"
        )
    
    # Conta file da migrare
    docs_to_migrate = await db.documents.count_documents({"storage_path": {"$exists": False}, "file_data": {"$exists": True}})
    payslips_to_migrate = await db.payslips.count_documents({"storage_path": {"$exists": False}, "file_data": {"$exists": True}})
    
    if docs_to_migrate == 0 and payslips_to_migrate == 0:
        return {"message": "Nessun file da migrare. Tutti i file sono già su cloud storage."}
    
    # Avvia migrazione in background
    async def migrate_files():
        migrated = 0
        errors = 0
        
        # Migra documenti
        async for doc in db.documents.find({"storage_path": {"$exists": False}, "file_data": {"$exists": True}}):
            try:
                file_data = base64.b64decode(doc["file_data"])
                result = await cloud_upload(
                    file_data=file_data,
                    client_id=doc.get("client_id", "unknown"),
                    original_filename=doc.get("file_name", "document.pdf"),
                    folder="documents"
                )
                if result.get("success"):
                    await db.documents.update_one(
                        {"id": doc["id"]},
                        {"$set": {"storage_path": result["storage_path"]}, "$unset": {"file_data": ""}}
                    )
                    migrated += 1
            except Exception as e:
                logger.error(f"Errore migrazione documento {doc.get('id')}: {e}")
                errors += 1
        
        # Migra buste paga
        async for ps in db.payslips.find({"storage_path": {"$exists": False}, "file_data": {"$exists": True}}):
            try:
                file_data = base64.b64decode(ps["file_data"])
                result = await cloud_upload(
                    file_data=file_data,
                    client_id=ps.get("client_id", "unknown"),
                    original_filename=ps.get("file_name", "payslip.pdf"),
                    folder="payslips"
                )
                if result.get("success"):
                    await db.payslips.update_one(
                        {"id": ps["id"]},
                        {"$set": {"storage_path": result["storage_path"]}, "$unset": {"file_data": ""}}
                    )
                    migrated += 1
            except Exception as e:
                logger.error(f"Errore migrazione busta paga {ps.get('id')}: {e}")
                errors += 1
        
        # Log risultato
        await db.migrations.insert_one({
            "id": str(uuid.uuid4()),
            "type": "storage_migration",
            "files_migrated": migrated,
            "errors": errors,
            "completed_at": datetime.now(timezone.utc).isoformat()
        })
        
        logger.info(f"Migrazione completata: {migrated} file migrati, {errors} errori")
    
    background_tasks.add_task(migrate_files)
    
    await log_activity("migrazione_storage", f"Avviata migrazione di {docs_to_migrate + payslips_to_migrate} file", user["id"])
    
    return {
        "message": "Migrazione avviata in background",
        "files_to_migrate": {
            "documents": docs_to_migrate,
            "payslips": payslips_to_migrate,
            "total": docs_to_migrate + payslips_to_migrate
        }
    }

# ==================== BANK ENTITIES ENDPOINTS ====================

@api_router.get("/bank-entities")
async def get_bank_entities(user: dict = Depends(require_commercialista)):
    """Lista tutte le entità bancarie disponibili"""
    entities = await db.bank_entities.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return entities

@api_router.post("/bank-entities")
async def create_bank_entity(entity: BankEntity, user: dict = Depends(require_commercialista)):
    """Crea una nuova entità bancaria personalizzata"""
    # Verifica che non esista già
    existing = await db.bank_entities.find_one({"name": {"$regex": f"^{entity.name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Entità bancaria già esistente")
    
    entity_id = str(uuid.uuid4())
    entity_doc = {
        "id": entity_id,
        "name": entity.name,
        "is_default": False,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.bank_entities.insert_one(entity_doc)
    
    return {"success": True, "id": entity_id, "name": entity.name}

@api_router.delete("/bank-entities/{entity_id}")
async def delete_bank_entity(entity_id: str, user: dict = Depends(require_commercialista)):
    """Elimina un'entità bancaria (solo quelle non predefinite)"""
    entity = await db.bank_entities.find_one({"id": entity_id}, {"_id": 0})
    if not entity:
        raise HTTPException(status_code=404, detail="Entità bancaria non trovata")
    
    if entity.get("is_default"):
        raise HTTPException(status_code=400, detail="Non puoi eliminare un'entità bancaria predefinita")
    
    await db.bank_entities.delete_one({"id": entity_id})
    return {"success": True, "message": "Entità bancaria eliminata"}

# ==================== CLIENT BANK CREDENTIALS ENDPOINTS ====================

@api_router.get("/clients/{client_id}/bank-credentials")
async def get_client_bank_credentials(client_id: str, user: dict = Depends(require_commercialista)):
    """Ottiene le credenziali bancarie di un cliente"""
    client = await db.users.find_one({"id": client_id, "role": "cliente"}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    credentials = client.get("bank_credentials", [])
    
    # Arricchisci con i nomi delle banche
    enriched = []
    for cred in credentials:
        bank = await db.bank_entities.find_one({"id": cred.get("bank_entity_id")}, {"_id": 0, "name": 1})
        enriched.append({
            **cred,
            "bank_name": bank["name"] if bank else "Sconosciuta"
        })
    
    return enriched

@api_router.post("/clients/{client_id}/bank-credentials")
async def add_client_bank_credential(
    client_id: str, 
    credential: BankCredential, 
    user: dict = Depends(require_commercialista)
):
    """Aggiunge una credenziale bancaria a un cliente"""
    client = await db.users.find_one({"id": client_id, "role": "cliente"}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    # Verifica che l'entità bancaria esista
    bank = await db.bank_entities.find_one({"id": credential.bank_entity_id}, {"_id": 0})
    if not bank:
        raise HTTPException(status_code=400, detail="Entità bancaria non valida")
    
    cred_id = str(uuid.uuid4())
    new_credential = {
        "id": cred_id,
        "bank_entity_id": credential.bank_entity_id,
        "username": credential.username,
        "password": credential.password,  # In produzione: cifrare!
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.update_one(
        {"id": client_id},
        {"$push": {"bank_credentials": new_credential}}
    )
    
    return {"success": True, "id": cred_id, "bank_name": bank["name"]}

@api_router.put("/clients/{client_id}/bank-credentials/{cred_id}")
async def update_client_bank_credential(
    client_id: str,
    cred_id: str,
    credential: BankCredentialUpdate,
    user: dict = Depends(require_commercialista)
):
    """Aggiorna una credenziale bancaria"""
    update_fields = {k: v for k, v in credential.dict().items() if v is not None}
    if not update_fields:
        raise HTTPException(status_code=400, detail="Nessun campo da aggiornare")
    
    # Aggiorna la credenziale specifica nell'array
    update_query = {f"bank_credentials.$.{k}": v for k, v in update_fields.items()}
    
    result = await db.users.update_one(
        {"id": client_id, "bank_credentials.id": cred_id},
        {"$set": update_query}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Credenziale non trovata")
    
    return {"success": True, "message": "Credenziale aggiornata"}

@api_router.delete("/clients/{client_id}/bank-credentials/{cred_id}")
async def delete_client_bank_credential(
    client_id: str,
    cred_id: str,
    user: dict = Depends(require_commercialista)
):
    """Elimina una credenziale bancaria"""
    result = await db.users.update_one(
        {"id": client_id},
        {"$pull": {"bank_credentials": {"id": cred_id}}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Credenziale non trovata")
    
    return {"success": True, "message": "Credenziale eliminata"}

# ==================== CLIENT ADDITIONAL EMAILS ====================

@api_router.post("/clients/{client_id}/emails")
async def add_client_email(client_id: str, email: EmailStr = Form(...), user: dict = Depends(require_commercialista)):
    """Aggiunge un'email aggiuntiva a un cliente"""
    client = await db.users.find_one({"id": client_id, "role": "cliente"}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    current_emails = client.get("additional_emails", [])
    if email in current_emails or email == client.get("email"):
        raise HTTPException(status_code=400, detail="Email già presente")
    
    await db.users.update_one(
        {"id": client_id},
        {"$push": {"additional_emails": email}}
    )
    
    return {"success": True, "email": email}

@api_router.delete("/clients/{client_id}/emails/{email}")
async def delete_client_email(client_id: str, email: str, user: dict = Depends(require_commercialista)):
    """Rimuove un'email aggiuntiva"""
    result = await db.users.update_one(
        {"id": client_id},
        {"$pull": {"additional_emails": email}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Email non trovata")
    
    return {"success": True, "message": "Email rimossa"}

# ==================== CONSULENTE LAVORO ENDPOINTS ====================

@api_router.post("/consulenti")
async def create_consulente(consulente: ConsulenteCreate, user: dict = Depends(require_commercialista)):
    """Crea un nuovo consulente del lavoro (solo commercialista)"""
    existing = await db.users.find_one({"email": consulente.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email già registrata")
    
    consulente_id = str(uuid.uuid4())
    consulente_doc = {
        "id": consulente_id,
        "email": consulente.email,
        "password": hash_password(consulente.password),
        "full_name": consulente.full_name,
        "role": "consulente_lavoro",
        "stato": "attivo",
        "assigned_clients": [],  # Clienti assegnati
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(consulente_doc)
    
    await log_activity(
        "consulente_creato",
        f"Creato consulente del lavoro: {consulente.full_name}",
        user["id"]
    )
    
    return {"success": True, "id": consulente_id, "email": consulente.email}

@api_router.post("/consulenti/invite")
async def invite_consulente(invite_data: ConsulenteInvite, user: dict = Depends(require_commercialista)):
    """
    Invita un consulente del lavoro via email.
    Il consulente riceverà un link per completare la registrazione.
    """
    # Verifica che l'email non sia già registrata
    existing = await db.users.find_one({"email": invite_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email già registrata nel sistema")
    
    # Verifica che non ci sia già un invito pendente
    existing_invite = await db.invitations.find_one({
        "notification_email": invite_data.email,
        "status": "pending",
        "role": "consulente_lavoro"
    })
    if existing_invite:
        raise HTTPException(status_code=400, detail="Esiste già un invito pendente per questa email")
    
    # Genera token e crea l'invito
    invitation_token = secrets.token_urlsafe(32)
    invite_id = str(uuid.uuid4())
    
    invite_doc = {
        "id": invite_id,
        "notification_email": invite_data.email,
        "suggested_name": invite_data.full_name,
        "role": "consulente_lavoro",  # Identificatore per distinguere dagli inviti clienti
        "invitation_token": invitation_token,
        "invitation_sent_at": datetime.now(timezone.utc).isoformat(),
        "invited_by": user["id"],
        "status": "pending",
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
    }
    await db.invitations.insert_one(invite_doc)
    
    # Costruisci il link di invito
    frontend_url = os.environ.get("FRONTEND_URL", "https://app.fiscaltaxcanarie.com")
    invitation_link = f"{frontend_url}/complete-registration?token={invitation_token}"
    
    # Invia email di invito
    try:
        await send_generic_email(
            to_email=invite_data.email,
            subject="Invito come Consulente del Lavoro - Fiscal Tax Canarie",
            html_body=f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <h2 style="color: #3caca4;">Benvenuto in Fiscal Tax Canarie</h2>
                <p>Ciao {invite_data.full_name},</p>
                <p>Sei stato invitato come <strong>Consulente del Lavoro</strong> nella piattaforma Fiscal Tax Canarie.</p>
                <p>Clicca sul pulsante qui sotto per completare la registrazione e impostare la tua password:</p>
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{invitation_link}" style="background-color: #6366f1; color: white; padding: 14px 28px; text-decoration: none; border-radius: 8px; font-weight: bold;">
                        Completa Registrazione
                    </a>
                </div>
                <p style="color: #666; font-size: 14px;">Questo link scadrà tra 7 giorni.</p>
                <p style="color: #666; font-size: 12px;">Se non hai richiesto questo invito, puoi ignorare questa email.</p>
                <hr style="border: none; border-top: 1px solid #eee; margin: 30px 0;">
                <p style="color: #999; font-size: 12px;">Fiscal Tax Canarie - Gestione Fiscale Professionale</p>
            </div>
            """,
            to_name=invite_data.full_name
        )
        email_sent = True
    except Exception as e:
        logger.error(f"Errore invio email invito consulente: {e}")
        email_sent = False
    
    await log_activity(
        "consulente_invitato",
        f"Invito inviato a consulente del lavoro: {invite_data.email}",
        user["id"]
    )
    
    return {
        "success": True,
        "message": f"Invito inviato a {invite_data.email}" + (" (email inviata)" if email_sent else " (email non inviata - verifica la configurazione)"),
        "invite_id": invite_id,
        "invitation_link": invitation_link,
        "email_sent": email_sent
    }

@api_router.get("/consulenti/invitations")
async def get_consulenti_invitations(user: dict = Depends(require_commercialista)):
    """Lista gli inviti pendenti per consulenti del lavoro"""
    invitations = await db.invitations.find(
        {"invited_by": user["id"], "status": "pending", "role": "consulente_lavoro"},
        {"_id": 0}
    ).to_list(100)
    return invitations

@api_router.post("/consulenti/resend-invite/{invite_id}")
async def resend_consulente_invitation(invite_id: str, user: dict = Depends(require_commercialista)):
    """Reinvia l'invito a un consulente"""
    invitation = await db.invitations.find_one({"id": invite_id, "role": "consulente_lavoro"}, {"_id": 0})
    
    if not invitation:
        raise HTTPException(status_code=404, detail="Invito non trovato")
    
    if invitation.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Questo invito è già stato utilizzato")
    
    # Genera nuovo token
    new_token = secrets.token_urlsafe(32)
    await db.invitations.update_one(
        {"id": invite_id},
        {"$set": {
            "invitation_token": new_token,
            "invitation_sent_at": datetime.now(timezone.utc).isoformat(),
            "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
        }}
    )
    
    frontend_url = os.environ.get("FRONTEND_URL", "https://app.fiscaltaxcanarie.com")
    invitation_link = f"{frontend_url}/complete-registration?token={new_token}"
    
    # Reinvia email
    try:
        await send_generic_email(
            to_email=invitation["notification_email"],
            subject="Invito come Consulente del Lavoro - Fiscal Tax Canarie (Reinvio)",
            html_content=f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <h2 style="color: #3caca4;">Benvenuto in Fiscal Tax Canarie</h2>
                <p>Ciao {invitation.get('suggested_name', '')},</p>
                <p>Ecco il tuo nuovo link di invito come <strong>Consulente del Lavoro</strong>.</p>
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{invitation_link}" style="background-color: #6366f1; color: white; padding: 14px 28px; text-decoration: none; border-radius: 8px; font-weight: bold;">
                        Completa Registrazione
                    </a>
                </div>
                <p style="color: #666; font-size: 14px;">Questo link scadrà tra 7 giorni.</p>
            </div>
            """
        )
    except Exception as e:
        logger.error(f"Errore reinvio email consulente: {e}")
    
    return {"success": True, "message": f"Invito reinviato a {invitation['notification_email']}", "invitation_link": invitation_link}

@api_router.get("/consulenti")
async def get_consulenti(user: dict = Depends(require_commercialista)):
    """Lista tutti i consulenti del lavoro"""
    consulenti = await db.users.find(
        {"role": "consulente_lavoro"},
        {"_id": 0, "password": 0}
    ).to_list(100)
    return consulenti

@api_router.delete("/consulenti/{consulente_id}")
async def delete_consulente(consulente_id: str, user: dict = Depends(require_commercialista)):
    """Elimina un consulente del lavoro"""
    consulente = await db.users.find_one({"id": consulente_id, "role": "consulente_lavoro"}, {"_id": 0})
    if not consulente:
        raise HTTPException(status_code=404, detail="Consulente non trovato")
    
    await db.users.delete_one({"id": consulente_id})
    
    return {"success": True, "message": "Consulente eliminato"}

@api_router.delete("/consulenti/invitations/{invitation_id}")
async def delete_consulente_invitation(invitation_id: str, user: dict = Depends(require_commercialista)):
    """Elimina un invito per consulente del lavoro in attesa"""
    invitation = await db.invitations.find_one({
        "id": invitation_id, 
        "invited_by": user["id"],
        "role": "consulente_lavoro"
    })
    if not invitation:
        raise HTTPException(status_code=404, detail="Invito non trovato")
    
    result = await db.invitations.delete_one({"id": invitation_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Errore durante l'eliminazione")
    
    await log_activity(
        "eliminazione_invito_consulente",
        f"Invito consulente eliminato: {invitation.get('notification_email', 'N/A')}",
        user["id"]
    )
    
    return {"message": "Invito eliminato con successo"}

@api_router.post("/consulenti/{consulente_id}/assign-clients")
async def assign_clients_to_consulente(
    consulente_id: str,
    assignment: ClientAssignment,
    user: dict = Depends(require_commercialista)
):
    """Assegna clienti a un consulente del lavoro"""
    consulente = await db.users.find_one({"id": consulente_id, "role": "consulente_lavoro"}, {"_id": 0})
    if not consulente:
        raise HTTPException(status_code=404, detail="Consulente non trovato")
    
    # Verifica che tutti i client_ids esistano
    for cid in assignment.client_ids:
        client = await db.users.find_one({"id": cid, "role": "cliente"})
        if not client:
            raise HTTPException(status_code=400, detail=f"Cliente {cid} non trovato")
    
    await db.users.update_one(
        {"id": consulente_id},
        {"$set": {"assigned_clients": assignment.client_ids}}
    )
    
    return {"success": True, "assigned_count": len(assignment.client_ids)}

@api_router.get("/consulente/clients")
async def get_consulente_clients(user: dict = Depends(require_consulente_lavoro)):
    """Ottiene la lista dei clienti assegnati al consulente"""
    assigned_ids = user.get("assigned_clients", [])
    if not assigned_ids:
        return []
    
    clients = await db.users.find(
        {"id": {"$in": assigned_ids}, "role": "cliente"},
        {"_id": 0, "password": 0}
    ).to_list(1000)
    
    # Aggiungi conteggi
    for client in clients:
        client["payslips_count"] = await db.payslips.count_documents({"client_id": client["id"]})
    
    return clients

@api_router.get("/consulente/stats")
async def get_consulente_stats(user: dict = Depends(require_consulente_lavoro)):
    """Statistiche per la dashboard del consulente del lavoro"""
    assigned_ids = user.get("assigned_clients", [])
    
    total_payslips = 0
    if assigned_ids:
        total_payslips = await db.payslips.count_documents({"client_id": {"$in": assigned_ids}})
    
    return {
        "clients_assigned": len(assigned_ids),
        "total_payslips": total_payslips
    }

# ==================== EMPLOYEE (DIPENDENTI) ENDPOINTS ====================

async def send_employee_notification_email(subject: str, body: str):
    """Invia email di notifica per operazioni sui dipendenti"""
    try:
        for email in EMPLOYEE_NOTIFICATION_EMAILS:
            await send_generic_email(email, subject, body)
        logger.info(f"Email notifica dipendenti inviata a {len(EMPLOYEE_NOTIFICATION_EMAILS)} destinatari")
    except Exception as e:
        logger.error(f"Errore invio email notifica dipendenti: {e}")

async def create_employee_notification(
    notification_type: str,
    title: str,
    message: str,
    client_id: str,
    employee_id: str = None,
    send_email: bool = True
):
    """Crea una notifica per operazioni sui dipendenti"""
    notification_id = str(uuid.uuid4())
    notification = {
        "id": notification_id,
        "type": notification_type,
        "title": title,
        "message": message,
        "client_id": client_id,
        "employee_id": employee_id,
        "read": False,
        "read_by": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.employee_notifications.insert_one(notification)
    
    # Invia email
    if send_email:
        await send_employee_notification_email(
            f"[Fiscal Tax] {title}",
            f"""
            <h2>{title}</h2>
            <p>{message}</p>
            <p><small>Notifica generata automaticamente da Fiscal Tax Canarie</small></p>
            """
        )
    
    return notification_id

@api_router.delete("/employee-notifications/{notification_id}")
async def delete_employee_notification(notification_id: str, user: dict = Depends(require_commercialista)):
    """Elimina una notifica dipendente"""
    result = await db.employee_notifications.delete_one({"id": notification_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Notifica non trovata")
    return {"success": True, "message": "Notifica eliminata"}

@api_router.delete("/employee-notifications")
async def delete_all_employee_notifications(user: dict = Depends(require_commercialista)):
    """Elimina tutte le notifiche dipendenti"""
    result = await db.employee_notifications.delete_many({})
    return {"success": True, "deleted_count": result.deleted_count}

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, user: dict = Depends(get_current_user)):
    """Elimina un dipendente - accessibile a commercialista, consulente e cliente (per i propri dipendenti)"""
    employee = await db.employees.find_one({"id": employee_id})
    if not employee:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Verifica permessi
    if user["role"] == "cliente":
        # Il cliente può eliminare solo i propri dipendenti
        if employee.get("client_id") != user["id"]:
            raise HTTPException(status_code=403, detail="Non hai i permessi per eliminare questo dipendente")
    elif user["role"] == "consulente_lavoro":
        # Il consulente può eliminare solo dipendenti dei clienti assegnati
        assigned_clients = user.get("assigned_clients", [])
        if employee.get("client_id") not in assigned_clients:
            raise HTTPException(status_code=403, detail="Non hai i permessi per eliminare questo dipendente")
    elif user["role"] != "commercialista":
        raise HTTPException(status_code=403, detail="Accesso non autorizzato")
    
    # Elimina documenti associati al dipendente
    await db.employee_documents.delete_many({"employee_id": employee_id})
    
    # Elimina notifiche associate
    await db.employee_notifications.delete_many({"employee_id": employee_id})
    
    # Elimina il dipendente
    await db.employees.delete_one({"id": employee_id})
    
    await log_activity(
        "eliminazione_dipendente",
        f"Dipendente eliminato: {employee.get('full_name', 'N/A')}",
        user["id"]
    )
    
    return {"success": True, "message": "Dipendente eliminato con successo"}

@api_router.post("/employees/hire-request")
async def request_employee_hire(
    hire_data: EmployeeHireRequest,
    user: dict = Depends(get_current_user)
):
    """Cliente richiede assunzione di un dipendente"""
    if user["role"] != "cliente":
        raise HTTPException(status_code=403, detail="Solo i clienti possono richiedere assunzioni")
    
    # Validazione ore settimanali (max 40)
    if hire_data.weekly_hours and hire_data.weekly_hours > 40:
        raise HTTPException(status_code=400, detail="Le ore settimanali non possono superare 40")
    
    employee_id = str(uuid.uuid4())
    employee_doc = {
        "id": employee_id,
        "client_id": user["id"],
        "full_name": hire_data.full_name,
        "start_date": hire_data.start_date,
        "job_title": hire_data.job_title,
        "work_hours": hire_data.work_hours,
        "work_location": hire_data.work_location,
        "work_days": hire_data.work_days,
        "weekly_hours": hire_data.weekly_hours,
        "notes": hire_data.notes,
        "status": "pending",  # pending, active, terminated
        "termination_date": None,
        "termination_reason": None,
        "documents": [],  # Documenti del dipendente
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.employees.insert_one(employee_doc)
    
    client_name = user.get("full_name", user.get("email"))
    
    # Trova solo i consulenti del lavoro ASSEGNATI a questo cliente
    consulenti = await db.users.find(
        {
            "role": "consulente_lavoro",
            "assigned_clients": user["id"]  # Solo consulenti assegnati a questo cliente
        },
        {"_id": 0, "id": 1, "email": 1, "full_name": 1}
    ).to_list(100)
    
    # Importa la funzione per notifiche email al consulente
    from email_service import notify_consulente_employee_request
    
    # Crea notifica e invia email a ogni consulente del lavoro ASSEGNATO
    notification_message = f"Il cliente {client_name} ha richiesto l'assunzione di {hire_data.full_name} come {hire_data.job_title}. Data inizio: {hire_data.start_date}."
    
    for consulente in consulenti:
        # Crea notifica interna per il consulente
        notification_doc = {
            "id": str(uuid.uuid4()),
            "user_id": consulente["id"],
            "notification_type": "hire_request",
            "title": "Nuova richiesta di assunzione",
            "message": notification_message,
            "client_id": user["id"],
            "employee_id": employee_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.employee_notifications.insert_one(notification_doc)
        
        # Invia email al consulente con template dedicato
        if consulente.get("email"):
            try:
                await notify_consulente_employee_request(
                    consulente_email=consulente["email"],
                    consulente_name=consulente.get("full_name", "Consulente"),
                    client_name=client_name,
                    request_type="assunzione",
                    employee_name=hire_data.full_name,
                    details={
                        "job_title": hire_data.job_title,
                        "start_date": hire_data.start_date,
                        "work_location": hire_data.work_location,
                        "work_hours": hire_data.work_hours,
                        "work_days": hire_data.work_days,
                        "weekly_hours": hire_data.weekly_hours,
                        "notes": hire_data.notes
                    }
                )
                logger.info(f"Email notifica assunzione inviata al consulente {consulente['email']}")
            except Exception as e:
                logger.error(f"Errore invio email al consulente {consulente['email']}: {e}")
    
    # Notifica anche agli admin
    admins = await db.users.find(
        {"role": "commercialista"},
        {"_id": 0, "id": 1}
    ).to_list(100)
    
    for admin in admins:
        notification_doc = {
            "id": str(uuid.uuid4()),
            "user_id": admin["id"],
            "notification_type": "hire_request",
            "title": "Nuova richiesta di assunzione",
            "message": notification_message,
            "client_id": user["id"],
            "employee_id": employee_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.employee_notifications.insert_one(notification_doc)
    
    await log_activity(
        "richiesta_assunzione",
        f"Richiesta assunzione dipendente: {hire_data.full_name}",
        user["id"]
    )
    
    return {"success": True, "employee_id": employee_id, "message": "Richiesta di assunzione inviata"}

@api_router.post("/employees/{employee_id}/documents")
async def upload_employee_document(
    employee_id: str,
    file: UploadFile = File(...),
    document_type: str = Form(...),  # id_document, nie, contract, timesheet, other
    description: str = Form(None),
    user: dict = Depends(get_current_user)
):
    """Carica documento per un dipendente (cliente, consulente o admin)"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Verifica permessi
    can_access = False
    if user["role"] == "commercialista":
        can_access = True
    elif user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        can_access = employee["client_id"] in assigned
    elif user["role"] == "cliente":
        can_access = employee["client_id"] == user["id"]
    
    if not can_access:
        raise HTTPException(status_code=403, detail="Non hai accesso a questo dipendente")
    
    # Leggi file
    file_content = await file.read()
    file_base64 = base64.b64encode(file_content).decode("utf-8")
    
    doc_id = str(uuid.uuid4())
    doc = {
        "id": doc_id,
        "document_type": document_type,
        "description": description,
        "file_name": file.filename,
        "file_data": file_base64,
        "file_type": file.content_type,
        "uploaded_by": user["id"],
        "uploaded_by_role": user["role"],
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.employees.update_one(
        {"id": employee_id},
        {"$push": {"documents": doc}}
    )
    
    # Notifica se caricato dal cliente
    if user["role"] == "cliente":
        client_name = user.get("full_name", user.get("email"))
        doc_type_names = {
            "id_document": "Documento di identità",
            "nie": "NIE",
            "contract": "Contratto",
            "timesheet": "Registro orario",
            "other": "Documento"
        }
        await create_employee_notification(
            "document_upload",
            "Nuovo documento dipendente caricato",
            f"Il cliente {client_name} ha caricato un documento ({doc_type_names.get(document_type, document_type)}) per il dipendente {employee['full_name']}.",
            user["id"],
            employee_id
        )
    
    # Notifica se caricato dal consulente del lavoro (notifica all'admin)
    if user["role"] == "consulente_lavoro":
        consulente_name = user.get("full_name", user.get("email"))
        # Trova il nome del cliente proprietario del dipendente
        client = await db.users.find_one({"id": employee["client_id"]}, {"_id": 0, "full_name": 1})
        client_name = client.get("full_name") if client else "N/A"
        
        doc_type_names = {
            "id_document": "Documento di identità",
            "nie": "NIE",
            "contract": "Contratto",
            "timesheet": "Registro orario",
            "payslip": "Busta paga",
            "other": "Documento"
        }
        await create_employee_notification(
            "consulente_document_upload",
            "Documento caricato dal Consulente",
            f"Il consulente {consulente_name} ha caricato un documento ({doc_type_names.get(document_type, document_type)}) per il dipendente {employee['full_name']} (Cliente: {client_name}).",
            employee["client_id"],
            employee_id,
            send_email=False  # L'admin vedrà la notifica in dashboard
        )
    
    return {"success": True, "document_id": doc_id}

@api_router.post("/employees/{employee_id}/terminate")
async def request_employee_termination(
    employee_id: str,
    termination: EmployeeTerminationRequest,
    user: dict = Depends(get_current_user)
):
    """Cliente richiede licenziamento di un dipendente"""
    if user["role"] != "cliente":
        raise HTTPException(status_code=403, detail="Solo i clienti possono richiedere licenziamenti")
    
    employee = await db.employees.find_one({"id": employee_id, "client_id": user["id"]}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Aggiorna stato a "termination_pending"
    await db.employees.update_one(
        {"id": employee_id},
        {"$set": {
            "status": "termination_pending",
            "termination_date": termination.termination_date,
            "termination_reason": termination.reason,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    client_name = user.get("full_name", user.get("email"))
    notification_message = f"Il cliente {client_name} ha richiesto il licenziamento di {employee['full_name']}. Data cessazione: {termination.termination_date}. Motivo: {termination.reason or 'Non specificato'}."
    
    # Trova solo i consulenti del lavoro ASSEGNATI a questo cliente e invia notifica + email
    consulenti = await db.users.find(
        {
            "role": "consulente_lavoro",
            "assigned_clients": user["id"]  # Solo consulenti assegnati a questo cliente
        },
        {"_id": 0, "id": 1, "email": 1, "full_name": 1}
    ).to_list(100)
    
    # Importa la funzione per notifiche email al consulente
    from email_service import notify_consulente_employee_request
    
    for consulente in consulenti:
        # Crea notifica interna
        notification_doc = {
            "id": str(uuid.uuid4()),
            "user_id": consulente["id"],
            "notification_type": "termination_request",
            "title": "Richiesta di licenziamento",
            "message": notification_message,
            "client_id": user["id"],
            "employee_id": employee_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.employee_notifications.insert_one(notification_doc)
        
        # Invia email al consulente con template dedicato
        if consulente.get("email"):
            try:
                await notify_consulente_employee_request(
                    consulente_email=consulente["email"],
                    consulente_name=consulente.get("full_name", "Consulente"),
                    client_name=client_name,
                    request_type="licenziamento",
                    employee_name=employee['full_name'],
                    details={
                        "termination_date": termination.termination_date,
                        "reason": termination.reason or "Non specificato"
                    }
                )
                logger.info(f"Email notifica licenziamento inviata al consulente {consulente['email']}")
            except Exception as e:
                logger.error(f"Errore invio email al consulente {consulente['email']}: {e}")
    
    # Notifica anche agli admin
    admins = await db.users.find(
        {"role": "commercialista"},
        {"_id": 0, "id": 1}
    ).to_list(100)
    
    for admin in admins:
        notification_doc = {
            "id": str(uuid.uuid4()),
            "user_id": admin["id"],
            "notification_type": "termination_request",
            "title": "Richiesta di licenziamento",
            "message": notification_message,
            "client_id": user["id"],
            "employee_id": employee_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.employee_notifications.insert_one(notification_doc)
    
    return {"success": True, "message": "Richiesta di licenziamento inviata"}

@api_router.put("/employees/{employee_id}")
async def update_employee(
    employee_id: str,
    update_data: EmployeeUpdate,
    user: dict = Depends(require_commercialista_or_consulente)
):
    """Aggiorna dati dipendente (solo consulente o admin)"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Verifica permessi consulente
    if user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        if employee["client_id"] not in assigned:
            raise HTTPException(status_code=403, detail="Non hai accesso a questo dipendente")
    
    update_fields = {k: v for k, v in update_data.dict().items() if v is not None}
    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Assegna consulente se cambia stato
    if "status" in update_fields and user["role"] == "consulente_lavoro":
        update_fields["assigned_consulente"] = user["id"]
        update_fields["assigned_consulente_name"] = user.get("full_name", user.get("email"))
    
    await db.employees.update_one(
        {"id": employee_id},
        {"$set": update_fields}
    )
    
    # Se la modifica NON è fatta dal consulente, notifica i consulenti via email
    if user["role"] == "commercialista":
        # Ottieni info cliente
        client = await db.users.find_one({"id": employee["client_id"]}, {"_id": 0, "full_name": 1, "email": 1})
        client_name = client.get("full_name", client.get("email", "N/D")) if client else "N/D"
        
        consulenti = await db.users.find(
            {"role": "consulente_lavoro"},
            {"_id": 0, "id": 1, "email": 1, "full_name": 1}
        ).to_list(100)
        
        changes = ", ".join([f"{k}: {v}" for k, v in update_fields.items() if k not in ["updated_at"]])
        
        for consulente in consulenti:
            if consulente.get("email"):
                try:
                    email_subject = f"Modifica Dipendente - {employee['full_name']}"
                    email_content = f"""
                    <h2>Modifica Dati Dipendente</h2>
                    <p><strong>Dipendente:</strong> {employee['full_name']}</p>
                    <p><strong>Cliente:</strong> {client_name}</p>
                    <p><strong>Modifiche effettuate:</strong> {changes}</p>
                    <hr>
                    <p>Accedi alla piattaforma per visualizzare i dettagli.</p>
                    """
                    send_generic_email(consulente["email"], email_subject, email_content)
                except Exception as e:
                    print(f"Errore invio email al consulente {consulente['email']}: {e}")
    
    return {"success": True, "message": "Dipendente aggiornato"}

@api_router.get("/employees")
async def get_employees(
    client_id: str = None,
    status: str = None,
    user: dict = Depends(get_current_user)
):
    """Lista dipendenti in base al ruolo"""
    query = {}
    
    if user["role"] == "cliente":
        # Cliente vede solo i propri dipendenti
        query["client_id"] = user["id"]
    elif user["role"] == "consulente_lavoro":
        # Consulente vede dipendenti dei clienti assegnati
        assigned = user.get("assigned_clients", [])
        if client_id and client_id in assigned:
            query["client_id"] = client_id
        else:
            query["client_id"] = {"$in": assigned}
    elif user["role"] == "commercialista":
        # Admin vede tutti, può filtrare per client_id
        if client_id:
            query["client_id"] = client_id
    
    if status:
        query["status"] = status
    
    employees = await db.employees.find(query, {"_id": 0, "documents.file_data": 0}).to_list(500)
    
    # Aggiungi info cliente
    for emp in employees:
        client = await db.users.find_one({"id": emp["client_id"]}, {"_id": 0, "full_name": 1, "email": 1})
        emp["client_name"] = client.get("full_name", client.get("email")) if client else "N/A"
    
    return employees

@api_router.get("/employees/{employee_id}")
async def get_employee_detail(
    employee_id: str,
    user: dict = Depends(get_current_user)
):
    """Dettaglio singolo dipendente"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Verifica permessi
    can_access = False
    if user["role"] == "commercialista":
        can_access = True
    elif user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        can_access = employee["client_id"] in assigned
    elif user["role"] == "cliente":
        can_access = employee["client_id"] == user["id"]
    
    if not can_access:
        raise HTTPException(status_code=403, detail="Non hai accesso a questo dipendente")
    
    # Aggiungi info cliente
    client = await db.users.find_one({"id": employee["client_id"]}, {"_id": 0, "full_name": 1, "email": 1})
    employee["client_name"] = client.get("full_name", client.get("email")) if client else "N/A"
    
    # Rimuovi file_data dai documenti per la risposta (troppo pesante)
    if "documents" in employee:
        for doc in employee["documents"]:
            doc.pop("file_data", None)
    
    return employee

@api_router.get("/employees/{employee_id}/documents/{doc_id}/download")
async def download_employee_document(
    employee_id: str,
    doc_id: str,
    user: dict = Depends(get_current_user)
):
    """Download documento dipendente"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Verifica permessi
    can_access = False
    if user["role"] == "commercialista":
        can_access = True
    elif user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        can_access = employee["client_id"] in assigned
    elif user["role"] == "cliente":
        can_access = employee["client_id"] == user["id"]
    
    if not can_access:
        raise HTTPException(status_code=403, detail="Non hai accesso a questo dipendente")
    
    # Trova documento
    doc = next((d for d in employee.get("documents", []) if d["id"] == doc_id), None)
    if not doc:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    return {
        "file_name": doc["file_name"],
        "file_data": doc["file_data"],
        "file_type": doc["file_type"]
    }

@api_router.get("/employees/{employee_id}/documents/{doc_id}/preview")
async def preview_employee_document(
    employee_id: str,
    doc_id: str,
    token: Optional[str] = None
):
    """Preview documento dipendente inline"""
    from fastapi.responses import Response
    
    # Verifica token
    if not token:
        raise HTTPException(status_code=401, detail="Token non fornito")
    
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Token non valido")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Utente non trovato")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token scaduto")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Token non valido")
    
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Verifica permessi
    can_access = False
    if user["role"] == "commercialista":
        can_access = True
    elif user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        can_access = employee["client_id"] in assigned
    elif user["role"] == "cliente":
        can_access = employee["client_id"] == user["id"]
    
    if not can_access:
        raise HTTPException(status_code=403, detail="Non hai accesso a questo dipendente")
    
    # Trova documento
    doc = next((d for d in employee.get("documents", []) if d["id"] == doc_id), None)
    if not doc:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    # Decodifica il contenuto
    file_content = base64.b64decode(doc["file_data"])
    content_type = doc.get("file_type", "application/octet-stream")
    
    # Determina il content type dal nome file
    file_name = doc.get("file_name", "").lower()
    if file_name.endswith(".pdf"):
        content_type = "application/pdf"
    elif file_name.endswith((".jpg", ".jpeg")):
        content_type = "image/jpeg"
    elif file_name.endswith(".png"):
        content_type = "image/png"
    
    return Response(
        content=file_content,
        media_type=content_type,
        headers={
            "Content-Disposition": f'inline; filename="{doc.get("file_name", "document")}"',
            "Cache-Control": "private, max-age=3600"
        }
    )

@api_router.delete("/employees/{employee_id}/documents/{doc_id}")
async def delete_employee_document(
    employee_id: str,
    doc_id: str,
    user: dict = Depends(require_commercialista_or_consulente)
):
    """Elimina documento dipendente"""
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Verifica permessi consulente
    if user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        if employee["client_id"] not in assigned:
            raise HTTPException(status_code=403, detail="Non hai accesso a questo dipendente")
    
    await db.employees.update_one(
        {"id": employee_id},
        {"$pull": {"documents": {"id": doc_id}}}
    )
    
    return {"success": True, "message": "Documento eliminato"}

# ==================== EMPLOYEE NOTIFICATIONS ====================

@api_router.get("/employee-notifications")
async def get_employee_notifications(
    unread_only: bool = False,
    user: dict = Depends(require_commercialista_or_consulente)
):
    """Lista notifiche dipendenti per admin/consulente"""
    query = {}
    
    if user["role"] == "consulente_lavoro":
        # Consulente vede solo notifiche dei clienti assegnati
        assigned = user.get("assigned_clients", [])
        query["client_id"] = {"$in": assigned}
    
    if unread_only:
        query["$or"] = [
            {"read": False},
            {"read_by": {"$ne": user["id"]}}
        ]
    
    notifications = await db.employee_notifications.find(
        query, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return notifications

@api_router.get("/employee-notifications/count")
async def get_employee_notifications_count(user: dict = Depends(require_commercialista_or_consulente)):
    """Conta notifiche non lette"""
    query = {"read_by": {"$ne": user["id"]}}
    
    if user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        query["client_id"] = {"$in": assigned}
    
    count = await db.employee_notifications.count_documents(query)
    return {"unread_count": count}

@api_router.post("/employee-notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: str,
    user: dict = Depends(require_commercialista_or_consulente)
):
    """Segna notifica come letta"""
    await db.employee_notifications.update_one(
        {"id": notification_id},
        {"$addToSet": {"read_by": user["id"]}}
    )
    return {"success": True}

@api_router.post("/employee-notifications/read-all")
async def mark_all_notifications_read(user: dict = Depends(require_commercialista_or_consulente)):
    """Segna tutte le notifiche come lette"""
    query = {}
    if user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        query["client_id"] = {"$in": assigned}
    
    await db.employee_notifications.update_many(
        query,
        {"$addToSet": {"read_by": user["id"]}}
    )
    return {"success": True}

# ==================== CLIENT NOTIFICATIONS HISTORY ====================

async def log_client_notification(
    client_id: str,
    notification_type: str,
    title: str,
    message: str,
    sent_by: str,
    email_sent: bool = False,
    email_recipients: List[str] = None
):
    """Salva una notifica nella cronologia del cliente"""
    notification_id = str(uuid.uuid4())
    notification = {
        "id": notification_id,
        "client_id": client_id,
        "type": notification_type,  # document, deadline, note, welcome, invite, employee
        "title": title,
        "message": message,
        "sent_by": sent_by,
        "email_sent": email_sent,
        "email_recipients": email_recipients or [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.client_notifications_history.insert_one(notification)
    return notification_id

@api_router.get("/clients/{client_id}/notifications-history")
async def get_client_notifications_history(
    client_id: str,
    limit: int = 50,
    user: dict = Depends(require_commercialista_or_consulente)
):
    """Ottiene la cronologia delle notifiche inviate a un cliente"""
    # Verifica permessi consulente
    if user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        if client_id not in assigned:
            raise HTTPException(status_code=403, detail="Non hai accesso a questo cliente")
    
    notifications = await db.client_notifications_history.find(
        {"client_id": client_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return notifications

@api_router.get("/my-notifications-history")
async def get_my_notifications_history(
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    """Ottiene la cronologia delle notifiche ricevute dal cliente corrente"""
    if user["role"] != "cliente":
        raise HTTPException(status_code=403, detail="Solo i clienti possono accedere a questo endpoint")
    
    notifications = await db.client_notifications_history.find(
        {"client_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return notifications

@api_router.post("/clients/{client_id}/send-notification")
async def send_client_notification(
    client_id: str,
    title: str = Form(...),
    message: str = Form(...),
    send_email: bool = Form(False),
    user: dict = Depends(require_commercialista_or_consulente)
):
    """Invia una notifica personalizzata a un cliente"""
    client = await db.users.find_one({"id": client_id, "role": "cliente"}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Cliente non trovato")
    
    # Verifica permessi consulente
    if user["role"] == "consulente_lavoro":
        assigned = user.get("assigned_clients", [])
        if client_id not in assigned:
            raise HTTPException(status_code=403, detail="Non hai accesso a questo cliente")
    
    email_sent = False
    email_recipients = []
    
    if send_email and client.get("email"):
        try:
            await send_generic_email(
                client["email"],
                f"[Fiscal Tax] {title}",
                f"<h2>{title}</h2><p>{message}</p>"
            )
            email_sent = True
            email_recipients = [client["email"]]
        except Exception as e:
            logger.error(f"Errore invio email: {e}")
    
    # Salva nella cronologia
    await log_client_notification(
        client_id=client_id,
        notification_type="manual",
        title=title,
        message=message,
        sent_by=user["id"],
        email_sent=email_sent,
        email_recipients=email_recipients
    )
    
    return {"success": True, "email_sent": email_sent}

# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
